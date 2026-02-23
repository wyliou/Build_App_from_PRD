"""Tests for autoconvert.merge_tracker.

Verifies FR-010 acceptance criteria:
- Merged cell ranges captured before unmerging.
- All cells unmerged after MergeTracker.__init__.
- Anchor values propagated correctly per field-type rules.
- All lookup methods return correct results for merged and unmerged cells.
"""

from __future__ import annotations

import openpyxl
import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from autoconvert.merge_tracker import MergeTracker
from autoconvert.models import MergeRange


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_sheet() -> Worksheet:
    """Return a fresh worksheet from a new in-memory workbook.

    Returns:
        An openpyxl Worksheet with no data and no merged cells.
    """
    wb: Workbook = openpyxl.Workbook()
    return wb.active  # type: ignore[return-value]


# ---------------------------------------------------------------------------
# MergeTracker.__init__
# ---------------------------------------------------------------------------


class TestMergeTrackerInit:
    """Tests for MergeTracker.__init__ — capture, unmerge, and snapshot."""

    def test_merge_tracker_captures_ranges_before_unmerge(self) -> None:
        """Verify range is captured and sheet is unmerged after __init__.

        After MergeTracker(sheet):
        - is_in_merge(3, 2) returns True (B3 was in B3:D5 merge).
        - sheet has no remaining merged cells (unmerge succeeded).
        """
        sheet = _make_sheet()
        sheet["B3"] = "anchor_value"
        sheet.merge_cells("B3:D5")

        tracker = MergeTracker(sheet)

        # Tracker captured the range.
        assert tracker.is_in_merge(3, 2) is True
        # Sheet has been fully unmerged.
        assert len(list(sheet.merged_cells.ranges)) == 0  # type: ignore[attr-defined]

    def test_merge_tracker_no_merges(self) -> None:
        """Sheet with no merged cells — MergeTracker initialises without error.

        All is_in_merge queries return False.
        """
        sheet = _make_sheet()
        sheet["A1"] = "value"

        tracker = MergeTracker(sheet)

        assert tracker.is_in_merge(1, 1) is False
        assert tracker.is_in_merge(5, 5) is False

    def test_merge_tracker_multiple_ranges(self) -> None:
        """Two non-overlapping merge ranges are both captured.

        Lookup works for cells in each range independently.
        """
        sheet = _make_sheet()
        sheet["B3"] = "first"
        sheet["F10"] = "second"
        sheet.merge_cells("B3:D5")
        sheet.merge_cells("F10:G11")

        tracker = MergeTracker(sheet)

        # Both ranges captured.
        assert tracker.is_in_merge(3, 2) is True   # B3
        assert tracker.is_in_merge(5, 4) is True   # D5
        assert tracker.is_in_merge(10, 6) is True  # F10
        assert tracker.is_in_merge(11, 7) is True  # G11

        # Cells between ranges are not captured.
        assert tracker.is_in_merge(6, 5) is False

        # Sheet fully unmerged.
        assert len(list(sheet.merged_cells.ranges)) == 0  # type: ignore[attr-defined]


# ---------------------------------------------------------------------------
# is_merge_anchor and is_in_merge
# ---------------------------------------------------------------------------


class TestIsMergeAnchorAndIsInMerge:
    """Tests for is_merge_anchor and is_in_merge methods."""

    def test_is_merge_anchor_true_for_topleft(self) -> None:
        """Anchor cell (min_row, min_col) returns True from is_merge_anchor."""
        sheet = _make_sheet()
        sheet["B3"] = "anchor"
        sheet.merge_cells("B3:D5")
        tracker = MergeTracker(sheet)

        # B3 is row=3, col=2 — the top-left (anchor) of the range.
        assert tracker.is_merge_anchor(3, 2) is True

    def test_is_merge_anchor_false_for_non_anchor(self) -> None:
        """Non-anchor cells within the merge range return False."""
        sheet = _make_sheet()
        sheet["B3"] = "anchor"
        sheet.merge_cells("B3:D5")
        tracker = MergeTracker(sheet)

        # Row 4 is not min_row (3) — not the anchor.
        assert tracker.is_merge_anchor(4, 2) is False
        # Col 3 is not min_col (2) — not the anchor.
        assert tracker.is_merge_anchor(3, 3) is False

    def test_is_in_merge_true_for_all_cells_in_range(self) -> None:
        """Every cell within B3:D5 (rows 3-5, cols 2-4) returns True."""
        sheet = _make_sheet()
        sheet["B3"] = "anchor"
        sheet.merge_cells("B3:D5")
        tracker = MergeTracker(sheet)

        assert tracker.is_in_merge(3, 2) is True  # top-left anchor
        assert tracker.is_in_merge(5, 4) is True  # bottom-right
        assert tracker.is_in_merge(4, 3) is True  # middle cell

    def test_is_in_merge_false_outside_range(self) -> None:
        """Cells outside B3:D5 return False from is_in_merge."""
        sheet = _make_sheet()
        sheet["B3"] = "anchor"
        sheet.merge_cells("B3:D5")
        tracker = MergeTracker(sheet)

        assert tracker.is_in_merge(2, 2) is False  # row above
        assert tracker.is_in_merge(6, 2) is False  # row below

    def test_is_merge_anchor_false_not_in_any_merge(self) -> None:
        """Cell not in any merge range returns False from is_merge_anchor."""
        sheet = _make_sheet()
        sheet["B3"] = "anchor"
        sheet.merge_cells("B3:D5")
        tracker = MergeTracker(sheet)

        assert tracker.is_merge_anchor(10, 10) is False


# ---------------------------------------------------------------------------
# get_anchor_value
# ---------------------------------------------------------------------------


class TestGetAnchorValue:
    """Tests for get_anchor_value method."""

    def test_get_anchor_value_for_non_anchor_cell(self) -> None:
        """Non-anchor cell returns the anchor's value (string propagation).

        After unmerging, C6 is empty (None). get_anchor_value returns
        the anchor cell C5's value "ABC".
        """
        sheet = _make_sheet()
        sheet.cell(row=5, column=3).value = "ABC"
        sheet.merge_cells("C5:C7")
        tracker = MergeTracker(sheet)

        # After unmerge, C6 (row=6, col=3) is None.
        assert sheet.cell(row=6, column=3).value is None
        # get_anchor_value transparently returns the anchor value.
        assert tracker.get_anchor_value(sheet, 6, 3) == "ABC"

    def test_get_anchor_value_for_anchor_cell(self) -> None:
        """Anchor cell returns its own value directly."""
        sheet = _make_sheet()
        sheet.cell(row=5, column=3).value = "ABC"
        sheet.merge_cells("C5:C7")
        tracker = MergeTracker(sheet)

        assert tracker.get_anchor_value(sheet, 5, 3) == "ABC"

    def test_get_anchor_value_not_in_merge(self) -> None:
        """Cell not in any merge range returns its own cell value (fall-through)."""
        sheet = _make_sheet()
        sheet.cell(row=10, column=5).value = "standalone"
        sheet.merge_cells("C5:C7")  # unrelated merge
        tracker = MergeTracker(sheet)

        result = tracker.get_anchor_value(sheet, 10, 5)
        assert result == "standalone"

    def test_get_anchor_value_not_in_merge_none_cell(self) -> None:
        """Cell not in any merge with no value returns None (fall-through)."""
        sheet = _make_sheet()
        tracker = MergeTracker(sheet)

        # Cell (10, 5) was never set — should return None.
        result = tracker.get_anchor_value(sheet, 10, 5)
        assert result is None

    def test_get_anchor_value_last_non_anchor_row(self) -> None:
        """Last row of merge range (max_row) also returns anchor value."""
        sheet = _make_sheet()
        sheet.cell(row=5, column=3).value = "ABC"
        sheet.merge_cells("C5:C7")
        tracker = MergeTracker(sheet)

        # C7 is the last row of the merge range — still returns anchor.
        assert tracker.get_anchor_value(sheet, 7, 3) == "ABC"


# ---------------------------------------------------------------------------
# get_merge_range
# ---------------------------------------------------------------------------


class TestGetMergeRange:
    """Tests for get_merge_range method."""

    def test_get_merge_range_returns_range(self) -> None:
        """Cell inside B3:D5 returns the correct MergeRange object."""
        sheet = _make_sheet()
        sheet["B3"] = "anchor"
        sheet.merge_cells("B3:D5")
        tracker = MergeTracker(sheet)

        result = tracker.get_merge_range(4, 3)  # D4 — inside the range

        assert result is not None
        assert result == MergeRange(min_row=3, max_row=5, min_col=2, max_col=4)

    def test_get_merge_range_returns_none_outside(self) -> None:
        """Cell outside any merge range returns None."""
        sheet = _make_sheet()
        sheet["B3"] = "anchor"
        sheet.merge_cells("B3:D5")
        tracker = MergeTracker(sheet)

        result = tracker.get_merge_range(10, 10)

        assert result is None

    def test_get_merge_range_anchor_cell(self) -> None:
        """Anchor cell itself also returns the correct MergeRange."""
        sheet = _make_sheet()
        sheet["B3"] = "anchor"
        sheet.merge_cells("B3:D5")
        tracker = MergeTracker(sheet)

        result = tracker.get_merge_range(3, 2)

        assert result is not None
        assert result.min_row == 3
        assert result.min_col == 2
        assert result.max_row == 5
        assert result.max_col == 4

    def test_get_merge_range_no_merges(self) -> None:
        """Sheet with no merges returns None for every query."""
        sheet = _make_sheet()
        tracker = MergeTracker(sheet)

        assert tracker.get_merge_range(1, 1) is None
        assert tracker.get_merge_range(5, 5) is None


# ---------------------------------------------------------------------------
# is_data_area_merge
# ---------------------------------------------------------------------------


class TestIsDataAreaMerge:
    """Tests for is_data_area_merge method."""

    def test_is_data_area_merge_true(self) -> None:
        """Data-area merge (min_row > header_row) returns True."""
        sheet = _make_sheet()
        sheet["C10"] = "anchor"
        sheet.merge_cells("C10:C12")
        tracker = MergeTracker(sheet)

        # header_row=7, merge starts at row 10 > 7 → data area.
        assert tracker.is_data_area_merge(10, 3, 7) is True

    def test_is_data_area_merge_false_header_area(self) -> None:
        """Header-area merge (min_row <= header_row) returns False."""
        sheet = _make_sheet()
        sheet["C5"] = "anchor"
        sheet.merge_cells("C5:C6")
        tracker = MergeTracker(sheet)

        # header_row=7, merge starts at row 5 <= 7 → header area.
        assert tracker.is_data_area_merge(5, 3, 7) is False

    def test_is_data_area_merge_false_not_in_merge(self) -> None:
        """Cell not in any merge range returns False."""
        sheet = _make_sheet()
        sheet["C10"] = "anchor"
        sheet.merge_cells("C10:C12")
        tracker = MergeTracker(sheet)

        assert tracker.is_data_area_merge(15, 5, 7) is False

    def test_is_data_area_merge_boundary_min_row_equals_header_row(self) -> None:
        """Merge starting exactly at header_row is not a data-area merge."""
        sheet = _make_sheet()
        sheet["C7"] = "anchor"
        sheet.merge_cells("C7:C9")
        tracker = MergeTracker(sheet)

        # min_row=7 == header_row=7 → NOT data area (must be strictly greater).
        assert tracker.is_data_area_merge(7, 3, 7) is False

    def test_is_data_area_merge_non_anchor_in_data_range(self) -> None:
        """Non-anchor cell within a data-area merge also returns True."""
        sheet = _make_sheet()
        sheet["C10"] = "anchor"
        sheet.merge_cells("C10:C12")
        tracker = MergeTracker(sheet)

        # C12 is a non-anchor continuation row — still a data-area merge.
        assert tracker.is_data_area_merge(12, 3, 7) is True


# ---------------------------------------------------------------------------
# get_first_row_of_merge
# ---------------------------------------------------------------------------


class TestGetFirstRowOfMerge:
    """Tests for get_first_row_of_merge method."""

    def test_get_first_row_returns_min_row(self) -> None:
        """Cell in C10:C12 returns min_row=10 regardless of queried row."""
        sheet = _make_sheet()
        sheet["C10"] = "anchor"
        sheet.merge_cells("C10:C12")
        tracker = MergeTracker(sheet)

        # Query from the last row of the merge.
        assert tracker.get_first_row_of_merge(12, 3) == 10

    def test_get_first_row_not_in_merge(self) -> None:
        """Cell not in any merge returns its own row (identity)."""
        sheet = _make_sheet()
        tracker = MergeTracker(sheet)

        assert tracker.get_first_row_of_merge(20, 5) == 20

    def test_get_first_row_anchor_cell_returns_itself(self) -> None:
        """Anchor cell (min_row) returns its own row via get_first_row_of_merge."""
        sheet = _make_sheet()
        sheet["C10"] = "anchor"
        sheet.merge_cells("C10:C12")
        tracker = MergeTracker(sheet)

        # min_row of the range IS row 10, so the result equals the queried row.
        assert tracker.get_first_row_of_merge(10, 3) == 10

    def test_get_first_row_middle_of_merge(self) -> None:
        """Middle row of a large merge returns min_row, not its own row."""
        sheet = _make_sheet()
        sheet["A1"] = "anchor"
        sheet.merge_cells("A1:A10")
        tracker = MergeTracker(sheet)

        # Row 5 is in the middle of A1:A10 — first row is 1.
        assert tracker.get_first_row_of_merge(5, 1) == 1


# ---------------------------------------------------------------------------
# Horizontal merge (brand + brand_type pattern)
# ---------------------------------------------------------------------------


class TestHorizontalMerge:
    """Tests for horizontal merge scenarios (brand+brand_type FR-011 pattern)."""

    def test_horizontal_merge_anchor_value_propagates(self) -> None:
        """Non-anchor column E8 in D8:E8 returns anchor D8's value."""
        sheet = _make_sheet()
        sheet.cell(row=8, column=4).value = "无品牌"  # D8 — anchor
        sheet.merge_cells("D8:E8")
        tracker = MergeTracker(sheet)

        # After unmerge, E8 (row=8, col=5) is None.
        assert sheet.cell(row=8, column=5).value is None
        # get_anchor_value returns the anchor value for the non-anchor column.
        result = tracker.get_anchor_value(sheet, 8, 5)
        assert result == "无品牌"

    def test_horizontal_merge_anchor_detected(self) -> None:
        """D8 is the anchor of D8:E8; E8 is not."""
        sheet = _make_sheet()
        sheet.cell(row=8, column=4).value = "无品牌"
        sheet.merge_cells("D8:E8")
        tracker = MergeTracker(sheet)

        assert tracker.is_merge_anchor(8, 4) is True   # D8 — anchor
        assert tracker.is_merge_anchor(8, 5) is False  # E8 — non-anchor

    def test_horizontal_merge_both_columns_in_merge(self) -> None:
        """Both D8 and E8 are reported as in-merge after a horizontal merge."""
        sheet = _make_sheet()
        sheet.cell(row=8, column=4).value = "TestBrand"
        sheet.merge_cells("D8:E8")
        tracker = MergeTracker(sheet)

        assert tracker.is_in_merge(8, 4) is True
        assert tracker.is_in_merge(8, 5) is True
        # Row above and row below are NOT in the merge.
        assert tracker.is_in_merge(7, 4) is False
        assert tracker.is_in_merge(9, 5) is False

    def test_horizontal_merge_get_merge_range_correct(self) -> None:
        """get_merge_range for E8 returns D8:E8 range (min_row=8, max_row=8)."""
        sheet = _make_sheet()
        sheet.cell(row=8, column=4).value = "无品牌"
        sheet.merge_cells("D8:E8")
        tracker = MergeTracker(sheet)

        rng = tracker.get_merge_range(8, 5)

        assert rng is not None
        assert rng == MergeRange(min_row=8, max_row=8, min_col=4, max_col=5)
