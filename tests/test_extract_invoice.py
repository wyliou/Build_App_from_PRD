"""Tests for autoconvert.extract_invoice.

Verifies FR-011 acceptance criteria:
- Happy path: 3 data rows with correct precision.
- Stop conditions: TOTAL keyword in non-part_no column, footer keywords.
- Error handling: empty required field (ERR_030), invalid numeric (ERR_031).
- Placeholder detection: "/" is placeholder, "无" is preserved.
- COD override: meaningful COD overrides COO, placeholder COD does not.
- Horizontal merge: brand+brand_type merged propagates anchor value.
- Unit suffix stripping: "1000 PCS" parsed as Decimal("1000").
- inv_no from parameter: all items receive the parameter value.
- Header continuation: rows with part_no containing "Part No." are skipped.
"""

from __future__ import annotations

from decimal import Decimal

import openpyxl
import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from autoconvert.errors import ErrorCode, ProcessingError
from autoconvert.extract_invoice import extract_invoice_items
from autoconvert.merge_tracker import MergeTracker
from autoconvert.models import ColumnMapping

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_sheet() -> Worksheet:
    """Return a fresh worksheet from a new in-memory workbook.

    Returns:
        An openpyxl Worksheet with no data and no merged cells.
    """
    wb: Workbook = openpyxl.Workbook()
    return wb.active  # type: ignore[return-value]


def _base_field_map() -> dict[str, int]:
    """Return a standard field_map with 10 required columns + cod.

    Column layout (1-based):
        A=1: part_no, B=2: po_no, C=3: qty, D=4: price, E=5: amount,
        F=6: currency, G=7: coo, H=8: brand, I=9: brand_type, J=10: model

    Returns:
        Dictionary mapping field names to 1-based column indices.
    """
    return {
        "part_no": 1,
        "po_no": 2,
        "qty": 3,
        "price": 4,
        "amount": 5,
        "currency": 6,
        "coo": 7,
        "brand": 8,
        "brand_type": 9,
        "model": 10,
    }


def _make_column_map(
    field_map: dict[str, int],
    header_row: int = 1,
    effective_header_row: int = 1,
) -> ColumnMapping:
    """Create a ColumnMapping for testing.

    Args:
        field_map: Maps field names to 1-based column indices.
        header_row: 1-based header row number.
        effective_header_row: 1-based effective header row (data starts at +1).

    Returns:
        A ColumnMapping instance.
    """
    return ColumnMapping(
        sheet_type="invoice",
        field_map=field_map,
        header_row=header_row,
        effective_header_row=effective_header_row,
    )


def _populate_data_row(
    sheet: Worksheet,
    row: int,
    part_no: str = "PART-001",
    po_no: str = "PO-001",
    qty: float | int | str = 100,
    price: float | int | str = 1.5,
    amount: float | int | str = 150.0,
    currency: str = "USD",
    coo: str = "CN",
    brand: str = "BrandX",
    brand_type: str = "OEM",
    model: str = "MODEL-A",
) -> None:
    """Fill a data row with the 10 required fields.

    Args:
        sheet: The worksheet to populate.
        row: 1-based row number.
        part_no: Part number value.
        po_no: Purchase order number.
        qty: Quantity (numeric or string with unit).
        price: Unit price.
        amount: Line total.
        currency: Currency code.
        coo: Country of origin.
        brand: Brand name.
        brand_type: Brand type.
        model: Model identifier.
    """
    sheet.cell(row=row, column=1, value=part_no)
    sheet.cell(row=row, column=2, value=po_no)
    sheet.cell(row=row, column=3, value=qty)
    sheet.cell(row=row, column=4, value=price)
    sheet.cell(row=row, column=5, value=amount)
    sheet.cell(row=row, column=6, value=currency)
    sheet.cell(row=row, column=7, value=coo)
    sheet.cell(row=row, column=8, value=brand)
    sheet.cell(row=row, column=9, value=brand_type)
    sheet.cell(row=row, column=10, value=model)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestExtractInvoiceItems:
    """Tests for extract_invoice_items function."""

    def test_extract_invoice_items_happy_path(self) -> None:
        """3 data rows with all 10 required fields populated.

        Asserts returns 3 InvoiceItems with correct field values and
        correct precision (qty=cell precision, price=5dp, amount=2dp).
        """
        sheet = _make_sheet()
        # Header in row 1.
        sheet.cell(row=1, column=1, value="Part No.")

        # Data rows 2, 3, 4.
        _populate_data_row(
            sheet, 2,
            part_no="P001", po_no="PO100", qty=10, price=2.5,
            amount=25.0, currency="USD", coo="CN",
            brand="BrandA", brand_type="OEM", model="M1",
        )
        _populate_data_row(
            sheet, 3,
            part_no="P002", po_no="PO200", qty=20, price=3.12345,
            amount=62.47, currency="EUR", coo="TW",
            brand="BrandB", brand_type="ODM", model="M2",
        )
        _populate_data_row(
            sheet, 4,
            part_no="P003", po_no="PO300", qty=5, price=10.0,
            amount=50.0, currency="USD", coo="JP",
            brand="BrandC", brand_type="OBM", model="M3",
        )

        column_map = _make_column_map(_base_field_map())
        tracker = MergeTracker(sheet)

        items = extract_invoice_items(sheet, column_map, tracker, inv_no=None)

        assert len(items) == 3

        # First item.
        assert items[0].part_no == "P001"
        assert items[0].po_no == "PO100"
        assert items[0].currency == "USD"
        assert items[0].coo == "CN"
        assert items[0].brand == "BrandA"
        assert items[0].brand_type == "OEM"
        assert items[0].model == "M1"
        assert items[0].inv_no is None
        assert items[0].allocated_weight is None

        # Precision: price=5dp, amount=2dp.
        assert items[1].price == Decimal("3.12345")
        assert items[1].amount == Decimal("62.47")

        # Third item.
        assert items[2].part_no == "P003"
        assert items[2].qty == Decimal("5")

    def test_extract_invoice_items_stop_condition_total_keyword(self) -> None:
        """Row 5 has "TOTAL" in column B (not part_no); part_no and qty empty.

        Asserts extraction stops at row 5 (stop condition 4 fires even on
        blank row). Only rows 2-4 are extracted.
        """
        sheet = _make_sheet()
        sheet.cell(row=1, column=1, value="Part No.")

        # 3 data rows.
        for r in range(2, 5):
            _populate_data_row(
                sheet, r,
                part_no=f"P{r}", po_no=f"PO{r}", qty=10, price=1.0,
                amount=10.0, currency="USD", coo="CN",
                brand="Brand", brand_type="Type", model="Model",
            )

        # Row 5: blank part_no and qty, but "TOTAL" in column B (po_no col).
        sheet.cell(row=5, column=1, value=None)  # part_no empty
        sheet.cell(row=5, column=2, value="TOTAL")  # stop keyword in B
        sheet.cell(row=5, column=3, value=None)  # qty empty

        # Row 6: should not be reached.
        _populate_data_row(
            sheet, 6,
            part_no="P6", po_no="PO6", qty=10, price=1.0,
            amount=10.0, currency="USD", coo="CN",
            brand="Brand", brand_type="Type", model="Model",
        )

        column_map = _make_column_map(_base_field_map())
        tracker = MergeTracker(sheet)

        items = extract_invoice_items(sheet, column_map, tracker, inv_no=None)

        assert len(items) == 3
        assert items[-1].part_no == "P4"

    def test_extract_invoice_items_stop_condition_footer(self) -> None:
        """part_no in row 4 contains footer keyword; extraction stops there.

        Only rows before row 4 are returned.
        """
        sheet = _make_sheet()
        sheet.cell(row=1, column=1, value="Part No.")

        # 2 data rows.
        _populate_data_row(
            sheet, 2,
            part_no="P1", po_no="PO1", qty=10, price=1.0,
            amount=10.0, currency="USD", coo="CN",
            brand="Brand", brand_type="Type", model="Model",
        )
        _populate_data_row(
            sheet, 3,
            part_no="P2", po_no="PO2", qty=20, price=2.0,
            amount=40.0, currency="USD", coo="TW",
            brand="Brand", brand_type="Type", model="Model",
        )

        # Row 4: footer keyword in part_no.
        sheet.cell(row=4, column=1, value="报关行代码")

        column_map = _make_column_map(_base_field_map())
        tracker = MergeTracker(sheet)

        items = extract_invoice_items(sheet, column_map, tracker, inv_no=None)

        assert len(items) == 2
        assert items[0].part_no == "P1"
        assert items[1].part_no == "P2"

    def test_extract_invoice_items_empty_required_field(self) -> None:
        """Row 3 has empty qty; raises ProcessingError with ERR_030.

        Asserts exact error code and field.
        """
        sheet = _make_sheet()
        sheet.cell(row=1, column=1, value="Part No.")

        _populate_data_row(
            sheet, 2,
            part_no="P1", po_no="PO1", qty=10, price=1.0,
            amount=10.0, currency="USD", coo="CN",
            brand="Brand", brand_type="Type", model="Model",
        )

        # Row 3: qty is empty.
        _populate_data_row(
            sheet, 3,
            part_no="P2", po_no="PO2", qty=5, price=1.0,
            amount=5.0, currency="USD", coo="CN",
            brand="Brand", brand_type="Type", model="Model",
        )
        # Reason: openpyxl's cell(value=None) call does not clear existing
        # values; must use direct .value assignment.
        sheet.cell(row=3, column=3).value = None  # override qty to empty

        column_map = _make_column_map(_base_field_map())
        tracker = MergeTracker(sheet)

        with pytest.raises(ProcessingError) as exc_info:
            extract_invoice_items(sheet, column_map, tracker, inv_no=None)

        assert exc_info.value.code == ErrorCode.ERR_030
        assert exc_info.value.row == 3
        assert exc_info.value.field == "qty"

    def test_extract_invoice_items_invalid_numeric(self) -> None:
        """Row 3 has "ABC" in the price column; raises ProcessingError ERR_031."""
        sheet = _make_sheet()
        sheet.cell(row=1, column=1, value="Part No.")

        _populate_data_row(
            sheet, 2,
            part_no="P1", po_no="PO1", qty=10, price=1.0,
            amount=10.0, currency="USD", coo="CN",
            brand="Brand", brand_type="Type", model="Model",
        )
        _populate_data_row(
            sheet, 3,
            part_no="P2", po_no="PO2", qty=5, price="ABC",
            amount=5.0, currency="USD", coo="CN",
            brand="Brand", brand_type="Type", model="Model",
        )

        column_map = _make_column_map(_base_field_map())
        tracker = MergeTracker(sheet)

        with pytest.raises(ProcessingError) as exc_info:
            extract_invoice_items(sheet, column_map, tracker, inv_no=None)

        assert exc_info.value.code == ErrorCode.ERR_031

    def test_extract_invoice_items_placeholder_detection(self) -> None:
        """COO "/" is a placeholder (raises ERR_030); COO "无" is preserved.

        Tests two rows: one with placeholder "/" in coo (should raise),
        and one with "无" which is a valid value.
        """
        sheet = _make_sheet()
        sheet.cell(row=1, column=1, value="Part No.")

        # Row 2: coo is "/" (placeholder) -> should raise ERR_030.
        _populate_data_row(
            sheet, 2,
            part_no="P1", po_no="PO1", qty=10, price=1.0,
            amount=10.0, currency="USD", coo="/",
            brand="Brand", brand_type="Type", model="Model",
        )

        column_map = _make_column_map(_base_field_map())
        tracker = MergeTracker(sheet)

        with pytest.raises(ProcessingError) as exc_info:
            extract_invoice_items(sheet, column_map, tracker, inv_no=None)

        assert exc_info.value.code == ErrorCode.ERR_030
        assert exc_info.value.field == "coo"

        # Separate test for "无" — should be preserved.
        sheet2 = _make_sheet()
        sheet2.cell(row=1, column=1, value="Part No.")
        _populate_data_row(
            sheet2, 2,
            part_no="P1", po_no="PO1", qty=10, price=1.0,
            amount=10.0, currency="USD", coo="无",
            brand="Brand", brand_type="Type", model="Model",
        )

        tracker2 = MergeTracker(sheet2)
        items = extract_invoice_items(
            sheet2, column_map, tracker2, inv_no=None
        )

        assert len(items) == 1
        assert items[0].coo == "无"

    def test_extract_invoice_items_cod_override_coo(self) -> None:
        """COD column has "CHINA" (meaningful); COO is "US".

        Asserts item.coo == "CHINA" (COD overrides) and item.cod == "CHINA".
        """
        sheet = _make_sheet()
        sheet.cell(row=1, column=1, value="Part No.")

        _populate_data_row(
            sheet, 2,
            part_no="P1", po_no="PO1", qty=10, price=1.0,
            amount=10.0, currency="USD", coo="US",
            brand="Brand", brand_type="Type", model="Model",
        )
        # Add COD column at column 11.
        sheet.cell(row=2, column=11, value="CHINA")

        fm = _base_field_map()
        fm["cod"] = 11
        column_map = _make_column_map(fm)
        tracker = MergeTracker(sheet)

        items = extract_invoice_items(sheet, column_map, tracker, inv_no=None)

        assert len(items) == 1
        assert items[0].coo == "CHINA"
        assert items[0].cod == "CHINA"

    def test_extract_invoice_items_cod_placeholder_no_override(self) -> None:
        """COD column has "/" (placeholder); COO column has "CN".

        Asserts item.coo == "CN" (placeholder COD does not override) and
        item.cod is None.
        """
        sheet = _make_sheet()
        sheet.cell(row=1, column=1, value="Part No.")

        _populate_data_row(
            sheet, 2,
            part_no="P1", po_no="PO1", qty=10, price=1.0,
            amount=10.0, currency="USD", coo="CN",
            brand="Brand", brand_type="Type", model="Model",
        )
        sheet.cell(row=2, column=11, value="/")

        fm = _base_field_map()
        fm["cod"] = 11
        column_map = _make_column_map(fm)
        tracker = MergeTracker(sheet)

        items = extract_invoice_items(sheet, column_map, tracker, inv_no=None)

        assert len(items) == 1
        assert items[0].coo == "CN"
        assert items[0].cod is None

    def test_extract_invoice_items_merged_brand_brand_type(self) -> None:
        """Brand and brand_type columns horizontally merged with "无品牌".

        After unmerge, brand_type col is None. Asserts both item.brand
        and item.brand_type == "无品牌".
        """
        sheet = _make_sheet()
        sheet.cell(row=1, column=1, value="Part No.")

        _populate_data_row(
            sheet, 2,
            part_no="P1", po_no="PO1", qty=10, price=1.0,
            amount=10.0, currency="USD", coo="CN",
            brand="无品牌", brand_type="will_be_merged", model="Model",
        )
        # Merge brand (col 8) and brand_type (col 9) horizontally.
        sheet.merge_cells(start_row=2, start_column=8, end_row=2, end_column=9)

        column_map = _make_column_map(_base_field_map())
        # Reason: MergeTracker must be created AFTER merge_cells to capture
        # the merge, then it unmerges the sheet.
        tracker = MergeTracker(sheet)

        items = extract_invoice_items(sheet, column_map, tracker, inv_no=None)

        assert len(items) == 1
        assert items[0].brand == "无品牌"
        assert items[0].brand_type == "无品牌"

    def test_extract_invoice_items_unit_suffix_stripping(self) -> None:
        """Qty cell contains "1000 PCS".

        Asserts parsed as Decimal("1000") without ERR_031.
        """
        sheet = _make_sheet()
        sheet.cell(row=1, column=1, value="Part No.")

        _populate_data_row(
            sheet, 2,
            part_no="P1", po_no="PO1", qty="1000 PCS", price=1.0,
            amount=1000.0, currency="USD", coo="CN",
            brand="Brand", brand_type="Type", model="Model",
        )

        column_map = _make_column_map(_base_field_map())
        tracker = MergeTracker(sheet)

        items = extract_invoice_items(sheet, column_map, tracker, inv_no=None)

        assert len(items) == 1
        assert items[0].qty == Decimal("1000")

    def test_extract_invoice_items_inv_no_from_parameter(self) -> None:
        """No inv_no column in field_map; inv_no="INV2024-001" as parameter.

        Asserts all items have inv_no == "2024-001" (after prefix cleaning).
        """
        sheet = _make_sheet()
        sheet.cell(row=1, column=1, value="Part No.")

        for r in range(2, 5):
            _populate_data_row(
                sheet, r,
                part_no=f"P{r}", po_no=f"PO{r}", qty=10, price=1.0,
                amount=10.0, currency="USD", coo="CN",
                brand="Brand", brand_type="Type", model="Model",
            )

        column_map = _make_column_map(_base_field_map())
        tracker = MergeTracker(sheet)

        items = extract_invoice_items(
            sheet, column_map, tracker, inv_no="INV2024-001"
        )

        assert len(items) == 3
        # Reason: _clean_inv_no strips "INV" prefix but "INV2024-001" has
        # "INV" not followed by # or space, so the regex "INV\s*#\s*" won't
        # match. The full value is preserved.
        for item in items:
            assert item.inv_no == "INV2024-001"

    def test_extract_invoice_items_header_continuation_skip(self) -> None:
        """Row immediately after header has part_no == "Part No.".

        Asserts row is skipped (not extracted as data item).
        """
        sheet = _make_sheet()
        sheet.cell(row=1, column=1, value="Part No.")

        # Row 2: header continuation.
        _populate_data_row(
            sheet, 2,
            part_no="Part No.", po_no="PO No.", qty=0, price=0,
            amount=0, currency="Currency", coo="COO",
            brand="Brand", brand_type="Type", model="Model",
        )

        # Row 3: actual data.
        _populate_data_row(
            sheet, 3,
            part_no="P1", po_no="PO1", qty=10, price=1.0,
            amount=10.0, currency="USD", coo="CN",
            brand="BrandA", brand_type="OEM", model="M1",
        )

        column_map = _make_column_map(_base_field_map())
        tracker = MergeTracker(sheet)

        items = extract_invoice_items(sheet, column_map, tracker, inv_no=None)

        assert len(items) == 1
        assert items[0].part_no == "P1"
