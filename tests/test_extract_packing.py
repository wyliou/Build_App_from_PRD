"""Tests for autoconvert.extract_packing.

Covers FR-012 through FR-017: extract_packing_items, validate_merged_weights,
detect_total_row, and extract_totals.
"""

from __future__ import annotations

from decimal import Decimal

import openpyxl
import pytest

from autoconvert.errors import ErrorCode, ProcessingError, WarningCode
from autoconvert.extract_packing import (
    detect_total_row,
    extract_packing_items,
    extract_totals,
    validate_merged_weights,
)
from autoconvert.merge_tracker import MergeTracker
from autoconvert.models import ColumnMapping

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_column_map(
    field_map: dict[str, int],
    header_row: int = 1,
    effective_header_row: int | None = None,
) -> ColumnMapping:
    """Build a ColumnMapping for testing.

    Args:
        field_map: Column field map.
        header_row: Header row (1-based).
        effective_header_row: Effective header row, defaults to header_row.

    Returns:
        ColumnMapping instance.
    """
    return ColumnMapping(
        sheet_type="packing",
        field_map=field_map,
        header_row=header_row,
        effective_header_row=effective_header_row or header_row,
    )


def _default_field_map() -> dict[str, int]:
    """Return a standard field map for tests: part_no=1, qty=2, nw=3, gw=4."""
    return {"part_no": 1, "qty": 2, "nw": 3, "gw": 4}


def _build_sheet(
    rows: list[list[object]],
    header_row_data: list[object] | None = None,
    merges: list[str] | None = None,
) -> tuple[openpyxl.Workbook, MergeTracker]:
    """Build an in-memory worksheet with optional merges.

    Row 1 is always the header row unless header_row_data is provided.
    Data rows start at row 2.

    Args:
        rows: List of row data lists (each becomes row 2, 3, ...).
        header_row_data: Optional header row content for row 1.
        merges: Optional list of merge range strings (e.g., "A2:A4").

    Returns:
        Tuple of (Workbook, MergeTracker).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    # Write header row
    if header_row_data:
        for col_idx, val in enumerate(header_row_data, start=1):
            ws.cell(row=1, column=col_idx, value=val)
    else:
        ws.cell(row=1, column=1, value="Part No")
        ws.cell(row=1, column=2, value="QTY")
        ws.cell(row=1, column=3, value="N.W.")
        ws.cell(row=1, column=4, value="G.W.")

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Apply merges BEFORE creating MergeTracker
    if merges:
        for merge_range in merges:
            ws.merge_cells(merge_range)

    # MergeTracker captures merges and unmerges
    tracker = MergeTracker(ws)
    return wb, tracker


# ---------------------------------------------------------------------------
# Tests: extract_packing_items (FR-012)
# ---------------------------------------------------------------------------


class TestExtractPackingItems:
    """Tests for extract_packing_items function."""

    def test_extract_packing_items_happy_path(self) -> None:
        """Three normal rows → 3 PackingItems with correct values."""
        rows = [
            ["ABC-001", 100, 5.5, 6.0],
            ["DEF-002", 200, 10.0, 12.0],
            ["GHI-003", 50, 2.5, 3.0],
        ]
        wb, tracker = _build_sheet(rows)
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        items, last_row = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 3
        assert items[0].part_no == "ABC-001"
        assert items[0].qty == Decimal("100")
        assert items[0].nw == Decimal("5.50000")
        assert items[0].is_first_row_of_merge is True
        assert items[0].row_number == 2
        assert items[2].part_no == "GHI-003"
        assert last_row == 4

    def test_extract_packing_items_stop_before_blank(self) -> None:
        """Keyword row stops extraction before blank check fires."""
        rows = [
            ["ABC-001", 100, 5.5, 6.0],
            ["DEF-002", 200, 10.0, 12.0],
            ["ABC-003", 50, 2.5, 3.0],
            [None, "合计", 18.0, 21.0],  # Stop keyword in col B
        ]
        wb, tracker = _build_sheet(rows)
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        items, last_row = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 3
        assert last_row == 4  # Row 4 is last data row (row 5 has 合计)

    def test_extract_packing_items_merged_nw_non_anchor(self) -> None:
        """Two rows share merged NW cell; non-anchor gets nw=0."""
        rows = [
            ["PART-A", 100, 5.0, 6.0],
            ["PART-A", 50, None, 3.0],  # Non-anchor NW (will be None after unmerge)
        ]
        wb, tracker = _build_sheet(rows, merges=["C2:C3"])
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        items, last_row = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 2
        assert items[0].nw == Decimal("5.00000")
        assert items[0].is_first_row_of_merge is True
        assert items[1].nw == Decimal("0")
        assert items[1].is_first_row_of_merge is False

    def test_extract_packing_items_ditto_mark_nw(self) -> None:
        """Ditto mark in NW column → nw=0, is_first_row_of_merge=False."""
        rows = [
            ["PART-A", 100, 5.0, 6.0],
            ["PART-A", 50, "\u3003", 3.0],  # 〃 ditto mark
        ]
        wb, tracker = _build_sheet(rows)
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        items, _ = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 2
        assert items[1].nw == Decimal("0")
        assert items[1].is_first_row_of_merge is False

    def test_extract_packing_items_implicit_continuation_same_part(self) -> None:
        """Same part_no, second row empty NW → implicit continuation."""
        rows = [
            ["PART-X", 100, 5.0, 6.0],
            ["PART-X", 50, None, 3.0],
        ]
        wb, tracker = _build_sheet(rows)
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        items, _ = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 2
        assert items[1].nw == Decimal("0")
        assert items[1].is_first_row_of_merge is False

    def test_extract_packing_items_vertical_merge_part_no(self) -> None:
        """Vertically merged part_no propagates anchor value to continuation rows."""
        # Part_no merged C2:C4 (col 1 = part_no in our field map,
        # but merge is on col A = col 1)
        # Remap: put part_no in col 3 to test merge on that column
        rows = [
            [100, 6.0, "MERGED-PT", 5.0],  # row 2
            [50, 3.0, None, 2.5],            # row 3 — non-anchor
            [25, 1.5, None, 1.0],            # row 4 — non-anchor
        ]
        field_map = {"part_no": 3, "qty": 1, "nw": 4, "gw": 2}
        wb, tracker = _build_sheet(
            rows,
            header_row_data=["QTY", "G.W.", "Part No", "N.W."],
            merges=["C2:C4"],
        )
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(field_map)

        items, last_row = extract_packing_items(ws, col_map, tracker)

        # All 3 rows should have the merged part_no
        assert len(items) == 3
        assert all(item.part_no == "MERGED-PT" for item in items)
        assert last_row == 4

    def test_extract_packing_items_pallet_row_skipped(self) -> None:
        """Row with 'PLT.' in part_no is skipped but doesn't terminate."""
        rows = [
            ["ABC-001", 100, 5.5, 6.0],
            ["PLT. SUMMARY", 1, 0.5, 1.0],  # Should be skipped
            ["DEF-002", 200, 10.0, 12.0],
        ]
        wb, tracker = _build_sheet(rows)
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        items, _ = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 2
        assert items[0].part_no == "ABC-001"
        assert items[1].part_no == "DEF-002"

    def test_extract_packing_items_empty_required_field(self) -> None:
        """Empty part_no with no merge and no previous part → ERR_030."""
        rows = [
            [None, 100, 5.0, 0],  # Empty part_no, gw=0 so not an implicit total
        ]
        wb, tracker = _build_sheet(rows)
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        with pytest.raises(ProcessingError) as exc_info:
            extract_packing_items(ws, col_map, tracker)

        assert exc_info.value.code == ErrorCode.ERR_030


# ---------------------------------------------------------------------------
# Tests: validate_merged_weights (FR-013)
# ---------------------------------------------------------------------------


class TestValidateMergedWeights:
    """Tests for validate_merged_weights function."""

    def test_validate_merged_weights_same_part_ok(self) -> None:
        """Same part_no sharing merged NW → no error."""
        rows = [
            ["PART-A", 100, 5.0, 6.0],
            ["PART-A", 50, None, 3.0],
        ]
        wb, tracker = _build_sheet(rows, merges=["C2:C3"])
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        items, _ = extract_packing_items(ws, col_map, tracker)
        # Should not raise
        validate_merged_weights(items, tracker, col_map)

    def test_validate_merged_weights_different_parts_error(self) -> None:
        """Different part_no sharing merged NW → ERR_046."""
        rows = [
            ["PART-A", 100, 5.0, 6.0],
            ["PART-B", 50, None, 3.0],
        ]
        wb, tracker = _build_sheet(rows, merges=["C2:C3"])
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        items, _ = extract_packing_items(ws, col_map, tracker)

        with pytest.raises(ProcessingError) as exc_info:
            validate_merged_weights(items, tracker, col_map)

        assert exc_info.value.code == ErrorCode.ERR_046

    def test_validate_merged_weights_no_merges(self) -> None:
        """No merged cells → validation passes (no exception)."""
        rows = [
            ["PART-A", 100, 5.0, 6.0],
            ["PART-B", 200, 10.0, 12.0],
        ]
        wb, tracker = _build_sheet(rows)
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        items, _ = extract_packing_items(ws, col_map, tracker)
        # Should not raise
        result = validate_merged_weights(items, tracker, col_map)
        assert result is None


# ---------------------------------------------------------------------------
# Tests: detect_total_row (FR-014)
# ---------------------------------------------------------------------------


class TestDetectTotalRow:
    """Tests for detect_total_row function."""

    def test_detect_total_row_strategy1_keyword(self) -> None:
        """Keyword '合计' found in search range → returns that row."""
        rows = [
            ["ABC-001", 100, 5.5, 6.0],     # row 2
            ["DEF-002", 200, 10.0, 12.0],    # row 3
            [None, None, None, None],         # row 4 blank
            ["合计", None, 15.5, 18.0],       # row 5 keyword
        ]
        wb, tracker = _build_sheet(rows)
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result = detect_total_row(ws, 3, col_map, tracker)
        assert result == 5

    def test_detect_total_row_strategy2_implicit(self) -> None:
        """No keyword row; implicit total row (empty part_no + NW>0 + GW>0)."""
        rows = [
            ["ABC-001", 100, 5.5, 6.0],      # row 2
            ["DEF-002", 200, 10.0, 12.0],     # row 3
            [None, None, None, None],          # row 4 blank
            [None, None, 15.5, 18.0],          # row 5 implicit total
        ]
        wb, tracker = _build_sheet(rows)
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result = detect_total_row(ws, 3, col_map, tracker)
        assert result == 5

    def test_detect_total_row_strategy2_excludes_merge_continuations(self) -> None:
        """Row with empty part_no due to merge is excluded from implicit detection."""
        # part_no merged A2:A3, so row 3 has empty part_no due to merge
        rows = [
            ["PART-A", 100, 5.5, 6.0],     # row 2 (merge anchor)
            [None, 50, 10.0, 12.0],         # row 3 (merge non-anchor — not a total)
            [None, None, None, None],        # row 4 blank
            [None, None, 15.5, 18.0],        # row 5 actual implicit total
        ]
        wb, tracker = _build_sheet(rows, merges=["A2:A3"])
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        # last_data_row = 3 (row 3 is still a data row, not a total)
        result = detect_total_row(ws, 3, col_map, tracker)
        assert result == 5  # Should skip row 3 (merge) and find row 5

    def test_detect_total_row_not_found(self) -> None:
        """No keyword and no implicit pattern → ERR_032."""
        rows = [
            ["ABC-001", 100, 5.5, 6.0],
        ]
        wb, tracker = _build_sheet(rows)
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        with pytest.raises(ProcessingError) as exc_info:
            detect_total_row(ws, 2, col_map, tracker)

        assert exc_info.value.code == ErrorCode.ERR_032


# ---------------------------------------------------------------------------
# Tests: extract_totals (FR-015, FR-016, FR-017)
# ---------------------------------------------------------------------------


class TestExtractTotals:
    """Tests for extract_totals function."""

    def _build_totals_sheet(
        self,
        total_row_data: list[object],
        extra_rows: list[list[object]] | None = None,
    ) -> openpyxl.Workbook:
        """Build a sheet with a data row and total row for totals extraction.

        Args:
            total_row_data: Data for the total row (row 3).
            extra_rows: Additional rows after total (rows 4, 5, ...).

        Returns:
            openpyxl Workbook.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None

        # Row 1: header
        for col, val in enumerate(["Part No", "QTY", "N.W.", "G.W."], start=1):
            ws.cell(row=1, column=col, value=val)
        # Row 2: data row
        for col, val in enumerate(["ABC", 100, 5.0, 6.0], start=1):
            ws.cell(row=2, column=col, value=val)
        # Row 3: total row
        for col, val in enumerate(total_row_data, start=1):
            ws.cell(row=3, column=col, value=val)
        # Extra rows after total
        if extra_rows:
            for row_idx, row_data in enumerate(extra_rows, start=4):
                for col, val in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col, value=val)
        return wb

    def test_extract_totals_total_nw_unit_stripping(self) -> None:
        """NW cell '10.5 KGS' → unit stripped, precision detected."""
        wb = self._build_totals_sheet(["TOTAL", 100, "10.50 KGS", 12.0])
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result, _warnings = extract_totals(ws, 3, col_map)

        assert result.total_nw == Decimal("10.5")

    def test_extract_totals_total_gw_packaging_weight(self) -> None:
        """Both +1 and +2 rows have numeric GW → use +2 row value."""
        wb = self._build_totals_sheet(
            ["TOTAL", 100, 15.5, 100.0],
            extra_rows=[
                [None, None, None, 5.0],    # row 4: +1 (pallet weight)
                [None, None, None, 105.0],   # row 5: +2 (final total)
            ],
        )
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result, _warnings = extract_totals(ws, 3, col_map)

        assert result.total_gw == Decimal("105")

    def test_extract_totals_total_gw_no_packaging(self) -> None:
        """Only +1 has numeric GW → keep primary total_gw."""
        wb = self._build_totals_sheet(
            ["TOTAL", 100, 15.5, 100.0],
            extra_rows=[
                [None, None, None, 5.0],    # row 4: +1 (just pallet subtotal)
                [None, None, None, None],    # row 5: +2 non-numeric/empty
            ],
        )
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result, _warnings = extract_totals(ws, 3, col_map)

        assert result.total_gw == Decimal("100")

    def test_extract_totals_packets_priority1_jianshu_label(self) -> None:
        """件数 label in row+1, value '55 件' in adjacent cell → 55."""
        wb = self._build_totals_sheet(
            ["TOTAL", 100, 15.5, 18.0],
            extra_rows=[
                ["件数", "55 件", None, None],  # row 4
            ],
        )
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result, _warnings = extract_totals(ws, 3, col_map)

        assert result.total_packets == 55

    def test_extract_totals_packets_priority1_unit_stripped(self) -> None:
        """Adjacent cell '3 件' stripped to '3' and parsed as 3."""
        wb = self._build_totals_sheet(
            ["TOTAL", 100, 15.5, 18.0],
            extra_rows=[
                ["件數", "3 件", None, None],  # row 4
            ],
        )
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result, _warnings = extract_totals(ws, 3, col_map)

        assert result.total_packets == 3

    def test_extract_totals_packets_priority2_plt(self) -> None:
        """PLT.G with numeric value in row total_row-1."""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        # Row 1: header
        for col, val in enumerate(["Part No", "QTY", "N.W.", "G.W."], start=1):
            ws.cell(row=1, column=col, value=val)
        # Row 2: PLT.G indicator with number
        ws.cell(row=2, column=1, value=5)
        ws.cell(row=2, column=2, value="PLT.G")
        # Row 3: total row
        ws.cell(row=3, column=1, value="TOTAL")
        ws.cell(row=3, column=2, value=100)
        ws.cell(row=3, column=3, value=15.5)
        ws.cell(row=3, column=4, value=18.0)

        col_map = _make_column_map(_default_field_map())
        result, _warnings = extract_totals(ws, 3, col_map)

        assert result.total_packets == 5

    def test_extract_totals_packets_priority3_unit_suffix_ctns(self) -> None:
        """'55 CTNS' in below-total row → 55."""
        wb = self._build_totals_sheet(
            ["TOTAL", 100, 15.5, 18.0],
            extra_rows=[
                ["55 CTNS", None, None, None],  # row 4
            ],
        )
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result, _warnings = extract_totals(ws, 3, col_map)

        assert result.total_packets == 55

    def test_extract_totals_packets_priority3_breakdown(self) -> None:
        """'348（256胶框+92纸箱）' → leading number 348."""
        wb = self._build_totals_sheet(
            ["TOTAL", 100, 15.5, 18.0],
            extra_rows=[
                ["\u0033\u0034\u0038\uff08256\u80f6\u6846+92\u7eb8\u7bb1\uff09", None, None, None],
            ],
        )
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result, _warnings = extract_totals(ws, 3, col_map)

        assert result.total_packets == 348

    def test_extract_totals_packets_priority3_embedded_chinese(self) -> None:
        """'共7托' → 7."""
        wb = self._build_totals_sheet(
            ["TOTAL", 100, 15.5, 18.0],
            extra_rows=[
                ["\u51717\u6258", None, None, None],  # 共7托
            ],
        )
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result, _warnings = extract_totals(ws, 3, col_map)

        assert result.total_packets == 7

    def test_extract_totals_packets_priority3_pallet_range(self) -> None:
        """'PLT#1(1~34)' → 1."""
        wb = self._build_totals_sheet(
            ["TOTAL", 100, 15.5, 18.0],
            extra_rows=[
                ["PLT#1(1~34)", None, None, None],  # row 4
            ],
        )
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result, _warnings = extract_totals(ws, 3, col_map)

        assert result.total_packets == 1

    def test_extract_totals_packets_pallet_wins_over_box(self) -> None:
        """'共7托（172件）' → 7 (pallet count, not 172)."""
        wb = self._build_totals_sheet(
            ["TOTAL", 100, 15.5, 18.0],
            extra_rows=[
                ["\u51717\u6258\uff08172\u4ef6\uff09", None, None, None],
            ],
        )
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result, _warnings = extract_totals(ws, 3, col_map)

        assert result.total_packets == 7

    def test_extract_totals_packets_not_found(self) -> None:
        """No pattern matches → total_packets=None, ATT_002 warning logged."""
        wb = self._build_totals_sheet(
            ["TOTAL", 100, 15.5, 18.0],
            extra_rows=[
                ["some random text", None, None, None],
            ],
        )
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        result, att_warnings = extract_totals(ws, 3, col_map)

        assert result.total_packets is None
        assert len(att_warnings) == 1
        assert att_warnings[0].code == WarningCode.ATT_002

    def test_extract_totals_invalid_total_nw(self) -> None:
        """Non-numeric NW in total row → ERR_033."""
        wb = self._build_totals_sheet(["TOTAL", 100, "abc", 18.0])
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        with pytest.raises(ProcessingError) as exc_info:
            extract_totals(ws, 3, col_map)

        assert exc_info.value.code == ErrorCode.ERR_033

    def test_extract_totals_invalid_total_gw(self) -> None:
        """Non-numeric GW in total row → ERR_034."""
        wb = self._build_totals_sheet(["TOTAL", 100, 15.5, "not_a_number"])
        ws = wb.active
        assert ws is not None
        col_map = _make_column_map(_default_field_map())

        with pytest.raises(ProcessingError) as exc_info:
            extract_totals(ws, 3, col_map)

        assert exc_info.value.code == ErrorCode.ERR_034
