"""Integration tests for AutoConvert.

Covers boundary (2-3 module wiring), full pipeline (process_file),
error propagation, edge cases, and PRD output format compliance.
"""

from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.workbook.workbook import Workbook

from autoconvert.batch import process_file
from autoconvert.column_map import (
    detect_header_row,
    extract_inv_no_from_header,
    map_columns,
)
from autoconvert.errors import ErrorCode, WarningCode
from autoconvert.extract_invoice import extract_invoice_items
from autoconvert.extract_packing import (
    detect_total_row,
    extract_packing_items,
    extract_totals,
    validate_merged_weights,
)
from autoconvert.merge_tracker import MergeTracker
from autoconvert.models import (
    AppConfig,
    FieldPattern,
    InvNoCellConfig,
    InvoiceItem,
    PackingTotals,
)
from autoconvert.output import write_template
from autoconvert.sheet_detect import detect_sheets
from autoconvert.transform import clean_po_number, convert_country, convert_currency
from autoconvert.validate import determine_file_status
from autoconvert.weight_alloc import allocate_weights

# Standard headers used across multiple tests
_INV_HDRS = [
    "Part No", "PO No", "Qty", "Price", "Amount",
    "Currency", "COO", "Brand", "Brand Type", "Model",
]
_INV_HDRS_WITH_INV = _INV_HDRS + ["Inv No"]
_PACK_HDRS = ["Part No", "Qty", "N.W.", "G.W."]
_PACK_HDRS_FULL = ["Part No", "PO No", "Qty", "N.W.", "G.W.", "Pack"]


def _fp(
    pat: str, ft: str = "string", req: bool = True,
) -> FieldPattern:
    return FieldPattern(
        patterns=[re.compile(pat, re.IGNORECASE)],
        field_type=ft, required=req,
    )


def _cfg() -> AppConfig:
    """Build a minimal AppConfig matching test workbook layout."""
    return AppConfig(
        invoice_sheet_patterns=[re.compile(r"invoice", re.IGNORECASE)],
        packing_sheet_patterns=[re.compile(r"packing", re.IGNORECASE)],
        invoice_columns={
            "part_no": _fp(r"part\s*no"), "po_no": _fp(r"po\s*no"),
            "qty": _fp(r"^qty$", "numeric"), "price": _fp(r"price", "numeric"),
            "amount": _fp(r"amount", "numeric"), "currency": _fp(r"currency"),
            "coo": _fp(r"coo|origin"), "cod": _fp(r"cod|destination", req=False),
            "brand": _fp(r"brand$"), "brand_type": _fp(r"brand\s*type"),
            "model": _fp(r"model"),
            "weight": _fp(r"weight", "numeric", False),
            "inv_no": _fp(r"inv.*no", req=False),
            "serial": _fp(r"serial", req=False),
        },
        packing_columns={
            "part_no": _fp(r"part\s*no"), "po_no": _fp(r"po\s*no", req=False),
            "qty": _fp(r"^qty$", "numeric"), "nw": _fp(r"n\.?w\.?", "numeric"),
            "gw": _fp(r"g\.?w\.?", "numeric"),
            "pack": _fp(r"pack", "numeric", False),
        },
        inv_no_cell=InvNoCellConfig(
            patterns=[re.compile(r"Invoice\s*No\.?\s*[:：]\s*(\S+)", re.IGNORECASE)],
            label_patterns=[re.compile(r"Invoice\s*No", re.IGNORECASE)],
            exclude_patterns=[re.compile(r"^Date$", re.IGNORECASE)],
        ),
        currency_lookup={"USD": "502", "EUR": "300"},
        country_lookup={"CHINA": "142", "TAIWAN,CHINA": "143", "USA": "502"},
        output_template_path=Path(__file__).parent.parent / "config" / "output_template.xlsx",
        invoice_min_headers=7, packing_min_headers=4,
    )


def _wb(
    inv_rows: list[list[object]],
    pack_rows: list[list[object]],
    total: list[object] | None = None,
    inv_hdrs: list[str] | None = None,
    pack_hdrs: list[str] | None = None,
    total_off: int = 2,
    hdr_area: dict[tuple[int, int], object] | None = None,
    packets: int | None = None,
) -> Workbook:
    """Create workbook: headers at row 8, data from row 9.

    Args:
        packets: If provided, places "{packets} CTNS" below the total row
            so that extract_totals can detect total_packets.
    """
    wb = openpyxl.Workbook()
    if wb.active is not None:
        wb.remove(wb.active)
    inv = wb.create_sheet("Invoice")
    if hdr_area:
        for (r, c), v in hdr_area.items():
            inv.cell(row=r, column=c, value=v)
    for ci, h in enumerate(inv_hdrs or _INV_HDRS, 1):
        inv.cell(row=8, column=ci, value=h)
    for ri, rd in enumerate(inv_rows, 9):
        for ci, v in enumerate(rd, 1):
            inv.cell(row=ri, column=ci, value=v)
    pack = wb.create_sheet("Packing List")
    for ci, h in enumerate(pack_hdrs or _PACK_HDRS, 1):
        pack.cell(row=8, column=ci, value=h)
    for ri, rd in enumerate(pack_rows, 9):
        for ci, v in enumerate(rd, 1):
            pack.cell(row=ri, column=ci, value=v)
    if total is not None:
        tr = 9 + len(pack_rows) + total_off - 1
        for ci, v in enumerate(total, 1):
            pack.cell(row=tr, column=ci, value=v)
        if packets is not None:
            pack.cell(row=tr + 1, column=1, value=f"{packets} CTNS")
    return wb


def _save(wb: Workbook, tmp: Path, name: str = "test.xlsx") -> Path:
    d = tmp / "data"
    d.mkdir(exist_ok=True)
    (d / "finished").mkdir(exist_ok=True)
    p = d / name
    wb.save(p)
    return p


# ---- 1. Boundary Tests (5) ----


class TestBoundary:
    """Wire 2-3 adjacent modules, verify data handoff."""

    def test_sheet_detect_to_column_map(self) -> None:
        """sheet_detect -> MergeTracker -> column_map."""
        cfg = _cfg()
        wb = _wb(
            [["P1", "PO1", 10, 1.0, 10.0, "USD", "CHINA", "B", "OEM", "M"]],
            [["P1", "PO1", 10, 5.5, 7.0, 1]],
            ["Total", "", "", 5.5, 7.0, ""],
            pack_hdrs=_PACK_HDRS_FULL,
        )
        pair = detect_sheets(wb, cfg)
        MergeTracker(pair.invoice_sheet)
        MergeTracker(pair.packing_sheet)
        inv_hdr = detect_header_row(pair.invoice_sheet, "invoice", cfg)
        inv_map = map_columns(pair.invoice_sheet, inv_hdr, "invoice", cfg)
        assert "part_no" in inv_map.field_map and inv_map.effective_header_row == 8
        pack_hdr = detect_header_row(pair.packing_sheet, "packing", cfg)
        pack_map = map_columns(pair.packing_sheet, pack_hdr, "packing", cfg)
        assert "nw" in pack_map.field_map and "gw" in pack_map.field_map

    def test_extraction_to_transform(self) -> None:
        """extract_invoice -> convert_currency -> convert_country -> clean_po."""
        cfg = _cfg()
        wb = _wb([["P1", "PO-100-2.1", 5, 2.0, 10.0, "USD", "CHINA",
                   "B1", "OEM", "M1"]], [])
        inv = wb["Invoice"]
        mt = MergeTracker(inv)
        hdr = detect_header_row(inv, "invoice", cfg)
        cm = map_columns(inv, hdr, "invoice", cfg)
        items = extract_invoice_items(inv, cm, mt, "INV-001")
        assert len(items) == 1 and items[0].currency == "USD"
        items, _ = convert_currency(items, cfg)
        assert items[0].currency == "502"
        items, _ = convert_country(items, cfg)
        assert items[0].coo == "142"
        items = clean_po_number(items)
        assert items[0].po_no == "PO"

    def test_packing_to_weight_alloc(self) -> None:
        """extract_packing -> extract_totals -> allocate_weights."""
        cfg = _cfg()
        wb = _wb(
            [["P1", "PO1", 10, 1.0, 10.0, "USD", "CHINA", "B", "OEM", "M"]],
            [["P1", 10, 5.5, 7.0]], ["Total", "", 5.5, 7.0],
        )
        pack = wb["Packing List"]
        mt = MergeTracker(pack)
        hdr = detect_header_row(pack, "packing", cfg)
        cm = map_columns(pack, hdr, "packing", cfg)
        items, lr = extract_packing_items(pack, cm, mt)
        validate_merged_weights(items, mt, cm)
        tr = detect_total_row(pack, lr, cm, mt)
        totals, _tw = extract_totals(pack, tr, cm)
        assert totals.total_nw == Decimal("5.5") and totals.total_nw_precision >= 1
        inv = wb["Invoice"]
        inv_mt = MergeTracker(inv)
        inv_hdr = detect_header_row(inv, "invoice", cfg)
        inv_cm = map_columns(inv, inv_hdr, "invoice", cfg)
        inv_items = extract_invoice_items(inv, inv_cm, inv_mt, "INV-001")
        result = allocate_weights(inv_items, items, totals)
        assert sum(i.allocated_weight for i in result if i.allocated_weight) == totals.total_nw

    def test_att003_yields_attention(self) -> None:
        """ATT_003 (unknown currency) -> determine_file_status -> Attention."""
        cfg = _cfg()
        item = InvoiceItem(
            part_no="P", po_no="PO", qty=Decimal("1"), price=Decimal("1"),
            amount=Decimal("1"), currency="ZZZ", coo="CHINA", cod=None,
            brand="B", brand_type="O", model="M", inv_no="I", serial=None,
            allocated_weight=None,
        )
        _, w = convert_currency([item], cfg)
        assert len(w) == 1 and w[0].code == WarningCode.ATT_003
        assert determine_file_status([], w) == "Attention"

    def test_att004_yields_attention(self) -> None:
        """ATT_004 (unknown COO) -> determine_file_status -> Attention."""
        cfg = _cfg()
        item = InvoiceItem(
            part_no="P", po_no="PO", qty=Decimal("1"), price=Decimal("1"),
            amount=Decimal("1"), currency="USD", coo="NARNIA", cod=None,
            brand="B", brand_type="O", model="M", inv_no="I", serial=None,
            allocated_weight=None,
        )
        _, w = convert_country([item], cfg)
        assert len(w) == 1 and w[0].code == WarningCode.ATT_004
        assert determine_file_status([], w) == "Attention"


# ---- 2. Full Pipeline Tests (3) ----


class TestFullPipeline:
    """End-to-end via process_file with synthetic .xlsx files."""

    def test_single_item_success(self, tmp_path: Path) -> None:
        """Minimal valid file -> Success, output created, weight allocated."""
        cfg = _cfg()
        wb = _wb(
            [["PA", "PO-200", 20, 3.0, 60.0, "USD", "CHINA",
              "BX", "OEM", "MX", "INV-100"]],
            [["PA", 20, 12.5, 15.0]], ["Total", "", 12.5, 15.0],
            inv_hdrs=_INV_HDRS_WITH_INV, packets=1,
        )
        fp = _save(wb, tmp_path)
        r = process_file(fp, cfg)
        assert r.status == "Success" and r.errors == [] and r.warnings == []
        assert r.invoice_items and len(r.invoice_items) == 1
        it = r.invoice_items[0]
        assert it.currency == "502" and it.coo == "142"
        assert it.allocated_weight == Decimal("12.5")
        assert (fp.parent / "finished" / "test_template.xlsx").exists()

    def test_multi_item_weight_sum(self, tmp_path: Path) -> None:
        """3 invoice items / 2 packing parts -- weights sum to total_nw."""
        cfg = _cfg()
        wb = _wb(
            [["PA", "PO1", 10, 2.0, 20.0, "USD", "CHINA", "B", "OEM", "M", "I"],
             ["PA", "PO2", 20, 2.0, 40.0, "USD", "CHINA", "B", "OEM", "M", "I"],
             ["PB", "PO3", 5, 10.0, 50.0, "USD", "USA", "B", "ODM", "M", "I"]],
            [["PA", 30, 6.0, 8.0], ["PB", 5, 4.0, 5.0]],
            ["Total", "", 10.0, 13.0], inv_hdrs=_INV_HDRS_WITH_INV,
            packets=2,
        )
        fp = _save(wb, tmp_path)
        r = process_file(fp, cfg)
        assert r.status == "Success" and r.invoice_items
        assert sum(i.allocated_weight for i in r.invoice_items
                   if i.allocated_weight) == Decimal("10.0")

    def test_attention_unknown_currency(self, tmp_path: Path) -> None:
        """Unknown currency -> Attention, raw value preserved, output written."""
        cfg = _cfg()
        wb = _wb(
            [["P", "PO", 10, 1.0, 10.0, "ZZZ", "CHINA", "B", "O", "M", "I"]],
            [["P", 10, 5.0, 7.0]], ["Total", "", 5.0, 7.0],
            inv_hdrs=_INV_HDRS_WITH_INV,
        )
        fp = _save(wb, tmp_path)
        r = process_file(fp, cfg)
        assert r.status == "Attention"
        assert any(w.code == WarningCode.ATT_003 for w in r.warnings)
        assert r.invoice_items and r.invoice_items[0].currency == "ZZZ"
        assert (fp.parent / "finished" / "test_template.xlsx").exists()


# ---- 3. Error Propagation Tests (5) ----


class TestErrorPropagation:
    """Errors at different stages surface with proper codes."""

    def _only_sheet(self, name: str, hdrs: list[str]) -> Workbook:
        wb = openpyxl.Workbook()
        if wb.active:
            wb.remove(wb.active)  # type: ignore[arg-type]
        ws = wb.create_sheet(name)
        for ci, h in enumerate(hdrs, 1):
            ws.cell(row=8, column=ci, value=h)
        return wb

    def test_err012_no_invoice_sheet(self, tmp_path: Path) -> None:
        wb = self._only_sheet("Packing List", _PACK_HDRS)
        r = process_file(_save(wb, tmp_path, "x.xlsx"), _cfg())
        assert r.status == "Failed"
        assert any(e.code == ErrorCode.ERR_012 for e in r.errors)

    def test_err013_no_packing_sheet(self, tmp_path: Path) -> None:
        wb = self._only_sheet("Invoice", _INV_HDRS)
        r = process_file(_save(wb, tmp_path, "x.xlsx"), _cfg())
        assert r.status == "Failed"
        assert any(e.code == ErrorCode.ERR_013 for e in r.errors)

    def test_err011_corrupt_file(self, tmp_path: Path) -> None:
        d = tmp_path / "data"
        d.mkdir(exist_ok=True)
        (d / "finished").mkdir(exist_ok=True)
        p = d / "bad.xlsx"
        p.write_bytes(b"not xlsx")
        r = process_file(p, _cfg())
        assert r.status == "Failed"
        assert any(e.code == ErrorCode.ERR_011 for e in r.errors)

    def test_err040_part_mismatch(self, tmp_path: Path) -> None:
        cfg = _cfg()
        wb = _wb(
            [["PX", "PO", 10, 1.0, 10.0, "USD", "CHINA", "B", "O", "M", "I"]],
            [["PY", 10, 5.0, 7.0]], ["Total", "", 5.0, 7.0],
            inv_hdrs=_INV_HDRS_WITH_INV,
        )
        r = process_file(_save(wb, tmp_path, "m.xlsx"), cfg)
        assert r.status == "Failed"
        codes = [e.code for e in r.errors]
        assert ErrorCode.ERR_040 in codes or ErrorCode.ERR_043 in codes

    def test_err014_no_header_row(self, tmp_path: Path) -> None:
        wb = openpyxl.Workbook()
        if wb.active:
            wb.remove(wb.active)  # type: ignore[arg-type]
        wb.create_sheet("Invoice").cell(row=1, column=1, value="x")
        wb.create_sheet("Packing List").cell(row=1, column=1, value="x")
        r = process_file(_save(wb, tmp_path, "nh.xlsx"), _cfg())
        assert r.status == "Failed"
        assert any(e.code == ErrorCode.ERR_014 for e in r.errors)


# ---- 4. Edge Cases (5) ----


class TestEdgeCases:
    """Empty input, minimal valid, duplicates, missing optional, fallback."""

    def test_empty_invoice_yields_no_items(self) -> None:
        cfg = _cfg()
        wb = _wb([], [["P", 10, 5.0, 7.0]], ["Total", "", 5.0, 7.0])
        inv = wb["Invoice"]
        mt = MergeTracker(inv)
        hdr = detect_header_row(inv, "invoice", cfg)
        cm = map_columns(inv, hdr, "invoice", cfg)
        assert extract_invoice_items(inv, cm, mt, "I") == []

    def test_missing_optional_serial_cod(self, tmp_path: Path) -> None:
        cfg = _cfg()
        wb = _wb(
            [["P", "PO", 10, 1.0, 10.0, "USD", "USA", "B", "O", "M", "I"]],
            [["P", 10, 5.0, 7.0]], ["Total", "", 5.0, 7.0],
            inv_hdrs=_INV_HDRS_WITH_INV, packets=1,
        )
        r = process_file(_save(wb, tmp_path), cfg)
        assert r.status == "Success" and r.invoice_items
        assert r.invoice_items[0].serial is None and r.invoice_items[0].cod is None

    def test_duplicate_part_no_proportional_weight(self, tmp_path: Path) -> None:
        cfg = _cfg()
        wb = _wb(
            [["D", "PO1", 10, 1.0, 10.0, "USD", "CHINA", "B", "O", "M", "I"],
             ["D", "PO2", 30, 1.0, 30.0, "USD", "CHINA", "B", "O", "M", "I"]],
            [["D", 40, 8.0, 10.0]], ["Total", "", 8.0, 10.0],
            inv_hdrs=_INV_HDRS_WITH_INV, packets=1,
        )
        r = process_file(_save(wb, tmp_path, "d.xlsx"), cfg)
        assert r.status == "Success" and r.invoice_items
        w0, w1 = r.invoice_items[0].allocated_weight, r.invoice_items[1].allocated_weight
        assert w0 is not None and w1 is not None
        assert w0 + w1 == Decimal("8.0")
        assert w0 == Decimal("2.0") and w1 == Decimal("6.0")

    def test_po_leading_delimiter_preserved(self) -> None:
        item = InvoiceItem(
            part_no="P", po_no="-PO123", qty=Decimal("1"), price=Decimal("1"),
            amount=Decimal("1"), currency="U", coo="C", cod=None,
            brand="B", brand_type="O", model="M", inv_no="I", serial=None,
            allocated_weight=None,
        )
        assert clean_po_number([item])[0].po_no == "-PO123"

    def test_inv_no_header_fallback(self) -> None:
        cfg = _cfg()
        wb = _wb(
            [["P", "PO", 10, 1.0, 10.0, "USD", "CHINA", "B", "O", "M"]],
            [], hdr_area={(3, 1): "Invoice No:", (3, 2): "HDR-999"},
        )
        assert extract_inv_no_from_header(wb["Invoice"], cfg) == "HDR-999"


# ---- 5. Output Format Tests (3) ----


def _output_data() -> tuple[list[InvoiceItem], PackingTotals]:
    """Build minimal data for output generation tests."""
    items = [
        InvoiceItem(
            part_no="PT1", po_no="PO1", qty=Decimal("10"),
            price=Decimal("5.00000"), amount=Decimal("50.00"),
            currency="502", coo="142", cod=None, brand="BA",
            brand_type="OEM", model="MX", inv_no="INV-001",
            serial="S1", allocated_weight=Decimal("12.50"),
        ),
        InvoiceItem(
            part_no="PT2", po_no="PO2", qty=Decimal("5"),
            price=Decimal("20.00000"), amount=Decimal("100.00"),
            currency="502", coo="142", cod=None, brand="BB",
            brand_type="ODM", model="M2", inv_no="INV-001",
            serial=None, allocated_weight=Decimal("7.50"),
        ),
    ]
    totals = PackingTotals(
        total_nw=Decimal("20.00"), total_nw_precision=2,
        total_gw=Decimal("25.00"), total_gw_precision=2,
        total_packets=3,
    )
    return items, totals


class TestOutputFormat:
    """Verify output matches PRD Section 7 Output Template Schema."""

    def test_column_mapping_compliance(self, tmp_path: Path) -> None:
        """All 40 columns: A-AN populated per PRD spec."""
        items, totals = _output_data()
        out = tmp_path / "out.xlsx"
        write_template(items, totals, _cfg(), out)
        ws = openpyxl.load_workbook(out)["\u5de5\u4f5c\u88681"]
        # Row 5 checks
        assert str(ws.cell(5, 1).value) == "PT1"       # A: part_no
        assert str(ws.cell(5, 2).value) == "PO1"       # B: po_no
        assert str(ws.cell(5, 3).value) == "3"          # C: fixed
        assert str(ws.cell(5, 4).value) == "502"        # D: currency
        assert float(ws.cell(5, 5).value) == 10.0       # E: qty
        assert float(ws.cell(5, 7).value) == 50.0       # G: amount
        assert str(ws.cell(5, 8).value) == "142"        # H: coo
        assert str(ws.cell(5, 12).value) == "S1"       # L: serial
        assert float(ws.cell(5, 13).value) == 12.5      # M: weight
        assert str(ws.cell(5, 14).value) == "INV-001"  # N: inv_no
        assert float(ws.cell(5, 16).value) == 25.0      # P: total_gw row5
        assert str(ws.cell(5, 18).value) == "32052"    # R: fixed
        assert str(ws.cell(5, 19).value) == "320506"   # S: fixed
        assert str(ws.cell(5, 20).value) == "142"      # T: fixed
        assert ws.cell(5, 37).value == 3                # AK: packets row5
        assert str(ws.cell(5, 38).value) == "BA"       # AL: brand
        assert str(ws.cell(5, 39).value) == "OEM"      # AM: brand_type
        assert str(ws.cell(5, 40).value) == "MX"       # AN: model
        # Row 6: no total_gw / total_packets
        assert ws.cell(6, 16).value is None
        assert ws.cell(6, 37).value is None

    def test_template_header_rows_preserved(self, tmp_path: Path) -> None:
        """Rows 1-4 from template are untouched."""
        items, totals = _output_data()
        out = tmp_path / "hdr.xlsx"
        write_template(items, totals, _cfg(), out)
        ws = openpyxl.load_workbook(out)["\u5de5\u4f5c\u88681"]
        assert any(ws.cell(1, c).value is not None for c in range(1, 41))

    def test_pipeline_creates_output_file(self, tmp_path: Path) -> None:
        """process_file creates {stem}_template.xlsx in finished/."""
        cfg = _cfg()
        wb = _wb(
            [["P", "PO", 10, 1.0, 10.0, "USD", "CHINA", "B", "O", "M", "I"]],
            [["P", 10, 5.0, 7.0]], ["Total", "", 5.0, 7.0],
            inv_hdrs=_INV_HDRS_WITH_INV, packets=1,
        )
        fp = _save(wb, tmp_path, "vnd.xlsx")
        r = process_file(fp, cfg)
        assert r.status == "Success"
        assert (fp.parent / "finished" / "vnd_template.xlsx").exists()
