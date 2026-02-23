"""Tests for autoconvert.config module.

Covers FR-002 acceptance criteria: existence checks, YAML validation,
regex compilation, lookup table parsing, duplicate rejection, and
template structure validation.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
import pytest
import yaml

from autoconvert.config import load_config
from autoconvert.errors import ConfigError, ErrorCode

# ---------------------------------------------------------------------------
# Helpers for building synthetic config files in tmp_path
# ---------------------------------------------------------------------------

# Minimal valid field_patterns.yaml content with all required structure.
_INVOICE_FIELD_NAMES = [
    "part_no",
    "po_no",
    "qty",
    "price",
    "amount",
    "currency",
    "coo",
    "cod",
    "brand",
    "brand_type",
    "model",
    "weight",
    "inv_no",
    "serial",
]

_PACKING_FIELD_NAMES = ["part_no", "po_no", "qty", "nw", "gw", "pack"]


def _make_field_entry(
    pattern: str = "^test$",
    field_type: str = "string",
    required: bool = True,
) -> dict[str, Any]:
    """Build a single field definition dict for YAML."""
    return {
        "patterns": [pattern],
        "type": field_type,
        "required": required,
    }


def _make_valid_yaml_data() -> dict[str, Any]:
    """Build a complete valid YAML data dict matching the spec."""
    invoice_columns: dict[str, Any] = {
        "min_header_matches": 7,
    }
    for name in _INVOICE_FIELD_NAMES:
        invoice_columns[name] = _make_field_entry(
            pattern=f"^{name}$", field_type="string", required=(name != "cod")
        )

    packing_columns: dict[str, Any] = {
        "min_header_matches": 3,
    }
    for name in _PACKING_FIELD_NAMES:
        packing_columns[name] = _make_field_entry(
            pattern=f"^{name}$",
            field_type="numeric" if name in ("qty", "nw", "gw", "pack") else "string",
            required=(name != "pack"),
        )

    return {
        "invoice_sheet": {"patterns": ["^invoice"]},
        "packing_sheet": {"patterns": ["^packing"]},
        "invoice_columns": invoice_columns,
        "packing_columns": packing_columns,
        "inv_no_cell": {
            "patterns": [r"INV\s*NO\.?\s*[:：]\s*(\S+)"],
            "label_patterns": [r"^INV\s*NO"],
            "exclude_patterns": [r"^invoice\s*no"],
        },
    }


def _write_yaml(config_dir: Path, data: dict[str, Any]) -> Path:
    """Write field_patterns.yaml to the config directory."""
    yaml_path = config_dir / "field_patterns.yaml"
    with open(yaml_path, "w", encoding="utf-8") as f:
        yaml.dump(data, f, allow_unicode=True, default_flow_style=False)
    return yaml_path


def _write_lookup_xlsx(
    config_dir: Path,
    filename: str,
    sheet_name: str,
    rows: list[tuple[str, Any]],
) -> Path:
    """Write a two-column lookup xlsx file (Source_Value, Target_Code)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name  # type: ignore[union-attr]
    ws.cell(row=1, column=1, value="Source_Value")  # type: ignore[union-attr]
    ws.cell(row=1, column=2, value="Target_Code")  # type: ignore[union-attr]
    for idx, (src, tgt) in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=src)  # type: ignore[union-attr]
        ws.cell(row=idx, column=2, value=tgt)  # type: ignore[union-attr]
    path = config_dir / filename
    wb.save(path)
    wb.close()
    return path


def _write_template_xlsx(
    config_dir: Path,
    sheet_name: str = "\u5de5\u4f5c\u88681",
    num_columns: int = 40,
    num_rows: int = 4,
) -> Path:
    """Write a minimal output_template.xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name  # type: ignore[union-attr]
    # Populate cells so max_column and max_row are set.
    for row in range(1, num_rows + 1):
        for col in range(1, num_columns + 1):
            ws.cell(row=row, column=col, value=f"R{row}C{col}")  # type: ignore[union-attr]
    path = config_dir / "output_template.xlsx"
    wb.save(path)
    wb.close()
    return path


def _create_full_config_dir(tmp_path: Path) -> Path:
    """Create a complete valid config directory with all four files."""
    config_dir = tmp_path / "config"
    config_dir.mkdir()

    _write_yaml(config_dir, _make_valid_yaml_data())
    _write_lookup_xlsx(
        config_dir,
        "currency_rules.xlsx",
        "Currency_Rules",
        [("USD", 502), ("EUR", 978)],
    )
    _write_lookup_xlsx(
        config_dir,
        "country_rules.xlsx",
        "Country_Rules",
        [("CHINA", 142), ("JAPAN", 116), ("Taiwan, China", 143)],
    )
    _write_template_xlsx(config_dir)
    return config_dir


# ===========================================================================
# Existence check tests
# ===========================================================================


class TestLoadConfigExistence:
    """Tests for missing config file detection (ERR_001)."""

    def test_load_config_missing_yaml(self, tmp_path: Path) -> None:
        """Config dir with no field_patterns.yaml raises ERR_001."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        # Create the other three files but NOT the YAML.
        _write_lookup_xlsx(
            config_dir, "currency_rules.xlsx", "Currency_Rules", [("USD", 502)]
        )
        _write_lookup_xlsx(
            config_dir, "country_rules.xlsx", "Country_Rules", [("CN", 142)]
        )
        _write_template_xlsx(config_dir)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_001
        assert "field_patterns.yaml" in exc_info.value.message
        assert exc_info.value.path is not None

    def test_load_config_missing_currency_rules(self, tmp_path: Path) -> None:
        """YAML present but currency_rules.xlsx absent raises ERR_001."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _write_yaml(config_dir, _make_valid_yaml_data())
        _write_lookup_xlsx(
            config_dir, "country_rules.xlsx", "Country_Rules", [("CN", 142)]
        )
        _write_template_xlsx(config_dir)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_001
        assert "currency_rules.xlsx" in exc_info.value.message

    def test_load_config_missing_country_rules(self, tmp_path: Path) -> None:
        """YAML and currency present but country_rules.xlsx absent raises ERR_001."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _write_yaml(config_dir, _make_valid_yaml_data())
        _write_lookup_xlsx(
            config_dir, "currency_rules.xlsx", "Currency_Rules", [("USD", 502)]
        )
        _write_template_xlsx(config_dir)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_001
        assert "country_rules.xlsx" in exc_info.value.message

    def test_load_config_missing_template(self, tmp_path: Path) -> None:
        """All three present but output_template.xlsx absent raises ERR_001."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _write_yaml(config_dir, _make_valid_yaml_data())
        _write_lookup_xlsx(
            config_dir, "currency_rules.xlsx", "Currency_Rules", [("USD", 502)]
        )
        _write_lookup_xlsx(
            config_dir, "country_rules.xlsx", "Country_Rules", [("CN", 142)]
        )

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_001
        assert "output_template.xlsx" in exc_info.value.message


# ===========================================================================
# YAML validation tests
# ===========================================================================


class TestLoadConfigYAML:
    """Tests for YAML structural validation (ERR_002, ERR_004)."""

    def test_load_config_invalid_regex(self, tmp_path: Path) -> None:
        """Malformed regex pattern string raises ERR_002 with pattern info."""
        config_dir = _create_full_config_dir(tmp_path)

        # Inject an invalid regex into the YAML.
        yaml_data = _make_valid_yaml_data()
        yaml_data["invoice_sheet"]["patterns"] = ["[invalid"]
        _write_yaml(config_dir, yaml_data)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_002
        assert "[invalid" in exc_info.value.message

    def test_load_config_missing_yaml_key(self, tmp_path: Path) -> None:
        """YAML missing 'invoice_columns' key raises ERR_004."""
        config_dir = _create_full_config_dir(tmp_path)

        yaml_data = _make_valid_yaml_data()
        del yaml_data["invoice_columns"]
        _write_yaml(config_dir, yaml_data)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_004
        assert "invoice_columns" in exc_info.value.message

    def test_load_config_malformed_column_entry(self, tmp_path: Path) -> None:
        """invoice_columns.part_no missing 'required' key raises ERR_004."""
        config_dir = _create_full_config_dir(tmp_path)

        yaml_data = _make_valid_yaml_data()
        # Remove the 'required' key from part_no.
        del yaml_data["invoice_columns"]["part_no"]["required"]
        _write_yaml(config_dir, yaml_data)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_004
        assert "required" in exc_info.value.message
        assert "part_no" in exc_info.value.message


# ===========================================================================
# Lookup table validation tests
# ===========================================================================


class TestLoadConfigLookup:
    """Tests for lookup table loading and validation (ERR_003)."""

    def test_load_config_duplicate_currency(self, tmp_path: Path) -> None:
        """Duplicate Source_Value in currency_rules.xlsx raises ERR_003."""
        config_dir = _create_full_config_dir(tmp_path)

        # Rewrite currency_rules with a duplicate.
        _write_lookup_xlsx(
            config_dir,
            "currency_rules.xlsx",
            "Currency_Rules",
            [("USD", 502), ("usd", 501)],  # "usd" normalizes to "USD" = dup
        )

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_003
        assert "USD" in exc_info.value.message

    def test_load_config_duplicate_country(self, tmp_path: Path) -> None:
        """Duplicate Source_Value in country_rules.xlsx raises ERR_003."""
        config_dir = _create_full_config_dir(tmp_path)

        # Rewrite country_rules with a duplicate after normalization.
        _write_lookup_xlsx(
            config_dir,
            "country_rules.xlsx",
            "Country_Rules",
            [("CHINA", 142), ("China", 142)],
        )

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_003
        assert "CHINA" in exc_info.value.message

    def test_load_config_country_int_target_code(self, tmp_path: Path) -> None:
        """Integer Target_Code (e.g., 142) is loaded successfully as str '142'."""
        config_dir = _create_full_config_dir(tmp_path)

        # Rewrite country_rules with integer Target_Code values.
        _write_lookup_xlsx(
            config_dir,
            "country_rules.xlsx",
            "Country_Rules",
            [("CHINA", 142), ("JAPAN", 116)],
        )

        cfg = load_config(config_dir)

        assert cfg.country_lookup["CHINA"] == "142"
        assert cfg.country_lookup["JAPAN"] == "116"
        # Verify they are str type, not int.
        assert isinstance(cfg.country_lookup["CHINA"], str)


# ===========================================================================
# Template validation tests
# ===========================================================================


class TestLoadConfigTemplate:
    """Tests for output template structural validation (ERR_005)."""

    def test_load_config_template_missing_sheet(self, tmp_path: Path) -> None:
        """Template exists but has no sheet named '工作表1' raises ERR_005."""
        config_dir = _create_full_config_dir(tmp_path)

        # Rewrite template with wrong sheet name.
        _write_template_xlsx(config_dir, sheet_name="Sheet1")

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_005
        assert "\u5de5\u4f5c\u88681" in exc_info.value.message

    def test_load_config_template_too_few_columns(self, tmp_path: Path) -> None:
        """Sheet '工作表1' with fewer than 40 columns raises ERR_005."""
        config_dir = _create_full_config_dir(tmp_path)

        # Rewrite template with only 10 columns.
        _write_template_xlsx(config_dir, num_columns=10)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_005
        assert "40" in exc_info.value.message


# ===========================================================================
# Happy path tests
# ===========================================================================


class TestLoadConfigHappyPath:
    """Tests for successful config loading."""

    def test_load_config_happy_path(self, tmp_path: Path) -> None:
        """Valid config dir with all four files returns AppConfig."""
        config_dir = _create_full_config_dir(tmp_path)
        cfg = load_config(config_dir)

        # Thresholds.
        assert cfg.invoice_min_headers == 7
        assert cfg.packing_min_headers == 4

        # Invoice sheet patterns are compiled re.Pattern objects.
        assert len(cfg.invoice_sheet_patterns) > 0
        assert isinstance(cfg.invoice_sheet_patterns[0], re.Pattern)

        # Currency lookup has UPPERCASE keys.
        assert "USD" in cfg.currency_lookup
        assert isinstance(cfg.currency_lookup["USD"], str)

        # All 14 invoice fields present.
        assert len(cfg.invoice_columns) == 14

        # All 6 packing fields present.
        assert len(cfg.packing_columns) == 6

        # InvNoCellConfig has all three pattern lists.
        assert len(cfg.inv_no_cell.patterns) > 0
        assert len(cfg.inv_no_cell.label_patterns) > 0
        assert len(cfg.inv_no_cell.exclude_patterns) > 0

        # output_template_path is set.
        assert cfg.output_template_path.exists()

    def test_load_config_lookup_normalization(self, tmp_path: Path) -> None:
        """Source_Value 'usd' (lowercase) is stored as key 'USD' (uppercase)."""
        config_dir = _create_full_config_dir(tmp_path)

        # Rewrite currency with lowercase source.
        _write_lookup_xlsx(
            config_dir,
            "currency_rules.xlsx",
            "Currency_Rules",
            [("usd", "502")],
        )

        cfg = load_config(config_dir)

        assert "USD" in cfg.currency_lookup
        assert cfg.currency_lookup["USD"] == "502"

    def test_load_config_comma_normalization(self, tmp_path: Path) -> None:
        """Source_Value 'Taiwan, China' normalizes key to 'TAIWAN,CHINA'."""
        config_dir = _create_full_config_dir(tmp_path)

        # Rewrite country with comma-space source.
        _write_lookup_xlsx(
            config_dir,
            "country_rules.xlsx",
            "Country_Rules",
            [("Taiwan, China", 143)],
        )

        cfg = load_config(config_dir)

        assert "TAIWAN,CHINA" in cfg.country_lookup
        assert cfg.country_lookup["TAIWAN,CHINA"] == "143"
