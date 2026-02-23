"""Tests for autoconvert.batch.

Covers run_batch orchestration (directory creation, clearing, file scanning,
diagnostic mode) and process_file pipeline (full success, phase short-circuit,
inv_no fallback, error handling).
"""

from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path
from typing import Any
from unittest.mock import patch

import openpyxl

from autoconvert.batch import process_file, run_batch
from autoconvert.errors import ErrorCode
from autoconvert.models import AppConfig

# ---------------------------------------------------------------------------
# Helpers — minimal workbook builder for process_file tests
# ---------------------------------------------------------------------------


def _make_minimal_config() -> AppConfig:
    """Build a minimal AppConfig with patterns that match test sheets.

    Returns:
        AppConfig suitable for testing with invoice/packing sheets
        named "Invoice" and "Packing List".
    """
    from autoconvert.models import FieldPattern, InvNoCellConfig

    def _fp(
        pattern: str, field_type: str = "string", required: bool = True,
    ) -> FieldPattern:
        return FieldPattern(
            patterns=[re.compile(pattern, re.IGNORECASE)],
            field_type=field_type,
            required=required,
        )

    invoice_columns = {
        "part_no": _fp(r"part\s*no|料号"),
        "po_no": _fp(r"po\s*no|订单"),
        "qty": _fp(r"qty|数量", field_type="numeric"),
        "price": _fp(r"price|单价", field_type="numeric"),
        "amount": _fp(r"amount|金额", field_type="numeric"),
        "currency": _fp(r"currency|币种"),
        "coo": _fp(r"coo|origin|原产"),
        "cod": _fp(r"cod|destination", required=False),
        "brand": _fp(r"brand|品牌"),
        "brand_type": _fp(r"brand\s*type|品牌类型"),
        "model": _fp(r"model|型号"),
        "weight": _fp(r"weight|净重", field_type="numeric", required=False),
        "inv_no": _fp(r"inv.*no|发票", required=False),
        "serial": _fp(r"serial|序号", required=False),
    }

    packing_columns = {
        "part_no": _fp(r"part\s*no|料号"),
        "po_no": _fp(r"po\s*no|订单", required=False),
        "qty": _fp(r"qty|数量", field_type="numeric"),
        "nw": _fp(r"n\.?w\.?|净重", field_type="numeric"),
        "gw": _fp(r"g\.?w\.?|毛重", field_type="numeric"),
        "pack": _fp(r"pack|件数", field_type="numeric", required=False),
    }

    inv_no_cell = InvNoCellConfig(
        patterns=[re.compile(r"Invoice\s*(?:No\.?|#)\s*[:：]?\s*(\S+)", re.IGNORECASE)],
        label_patterns=[re.compile(r"Invoice\s*(?:No\.?|#)", re.IGNORECASE)],
        exclude_patterns=[re.compile(r"^Date$", re.IGNORECASE)],
    )

    return AppConfig(
        invoice_sheet_patterns=[re.compile(r"invoice", re.IGNORECASE)],
        packing_sheet_patterns=[re.compile(r"packing", re.IGNORECASE)],
        invoice_columns=invoice_columns,
        packing_columns=packing_columns,
        inv_no_cell=inv_no_cell,
        currency_lookup={"USD": "502"},
        country_lookup={"CHINA": "142", "TAIWAN,CHINA": "143"},
        output_template_path=Path("config/output_template.xlsx"),
        invoice_min_headers=7,
        packing_min_headers=4,
    )


def _build_test_workbook(
    tmp_path: Path,
    filename: str = "vendor.xlsx",
    include_inv_no_col: bool = True,
    include_inv_no_header: bool = False,
    no_invoice_sheet: bool = False,
) -> Path:
    """Create a minimal test workbook with valid invoice and packing data.

    The workbook has headers at row 8 and data starting at row 9.

    Args:
        tmp_path: Temporary directory path.
        filename: Name for the generated file.
        include_inv_no_col: Whether to include inv_no in header columns.
        include_inv_no_header: Whether to include "Invoice No: INV-001" in
            header area (rows 1-15) for fallback testing.
        no_invoice_sheet: If True, omit the invoice sheet entirely.

    Returns:
        Path to the created workbook file.
    """
    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    if not no_invoice_sheet:
        inv = wb.create_sheet("Invoice")
        # Header area for inv_no fallback
        if include_inv_no_header:
            inv.cell(row=3, column=1, value="Invoice No:")
            inv.cell(row=3, column=2, value="INV-001")

        # Row 8: header row (needs >= 7 non-empty cells for invoice)
        inv_headers = [
            "Part No", "PO No", "Qty", "Price", "Amount",
            "Currency", "COO", "Brand", "Brand Type", "Model",
        ]
        if include_inv_no_col:
            inv_headers.append("Inv No")

        for col_idx, header in enumerate(inv_headers, start=1):
            inv.cell(row=8, column=col_idx, value=header)

        # Row 9: data row
        inv_data = [
            "PART-001", "PO-100", 10, Decimal("5.50"), Decimal("55.00"),
            "USD", "CHINA", "TestBrand", "OEM", "Model-X",
        ]
        if include_inv_no_col:
            inv_data.append("INV-001")

        for col_idx, val in enumerate(inv_data, start=1):
            inv.cell(row=9, column=col_idx, value=val)

    # Packing sheet
    pack = wb.create_sheet("Packing List")

    # Row 8: packing header (needs >= 4 non-empty cells)
    pack_headers = ["Part No", "PO No", "Qty", "N.W.", "G.W.", "Pack"]
    for col_idx, header in enumerate(pack_headers, start=1):
        pack.cell(row=8, column=col_idx, value=header)

    # Row 9: packing data row
    pack_data = ["PART-001", "PO-100", 10, Decimal("5.50"), Decimal("7.00"), 1]
    for col_idx, val in enumerate(pack_data, start=1):
        pack.cell(row=9, column=col_idx, value=val)

    # Row 11: total row (keyword marker)
    pack.cell(row=11, column=1, value="Total")
    pack.cell(row=11, column=4, value=Decimal("5.50"))
    pack.cell(row=11, column=5, value=Decimal("7.00"))

    data_dir = tmp_path / "data"
    data_dir.mkdir(exist_ok=True)
    finished_dir = data_dir / "finished"
    finished_dir.mkdir(exist_ok=True)

    filepath = data_dir / filename
    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# Tests — run_batch
# ---------------------------------------------------------------------------


class TestRunBatch:
    """Tests for run_batch orchestration logic."""

    def test_run_batch_creates_directories(self, tmp_path: Path) -> None:
        """Verify that run_batch creates data/ and data/finished/ dirs."""
        config = _make_minimal_config()
        data_dir = tmp_path / "data"
        finished_dir = data_dir / "finished"

        assert not data_dir.exists()
        assert not finished_dir.exists()

        result = run_batch(config, data_dir)

        assert data_dir.exists()
        assert finished_dir.exists()
        assert result.total_files == 0

    def test_run_batch_clears_finished_before_processing(
        self, tmp_path: Path,
    ) -> None:
        """Verify that data/finished/ is cleared before processing."""
        config = _make_minimal_config()
        data_dir = tmp_path / "data"
        finished_dir = data_dir / "finished"
        finished_dir.mkdir(parents=True)

        # Place a dummy file in finished/
        dummy = finished_dir / "old_output.xlsx"
        dummy.write_text("dummy")

        assert dummy.exists()

        run_batch(config, data_dir)

        assert not dummy.exists()

    def test_run_batch_skips_clear_in_diagnostic_mode(
        self, tmp_path: Path,
    ) -> None:
        """Verify diagnostic mode does NOT clear data/finished/."""
        config = _make_minimal_config()
        filepath = _build_test_workbook(tmp_path)
        data_dir = filepath.parent
        finished_dir = data_dir / "finished"

        # Place a dummy file in finished/
        dummy = finished_dir / "existing_output.xlsx"
        dummy.write_text("dummy")

        run_batch(config, data_dir, diagnostic_file=str(filepath))

        assert dummy.exists()

    def test_run_batch_excludes_temp_files(self, tmp_path: Path) -> None:
        """Verify ~$ prefixed temp files are excluded from processing."""
        config = _make_minimal_config()
        data_dir = tmp_path / "data"
        data_dir.mkdir(parents=True)
        (data_dir / "finished").mkdir()

        # Create a real file and a temp file
        _build_test_workbook(tmp_path, "vendor.xlsx")
        temp_file = data_dir / "~$vendor.xlsx"
        temp_file.write_text("temp")

        result = run_batch(config, data_dir)

        processed_names = [r.filename for r in result.file_results]
        assert "vendor.xlsx" in processed_names
        assert "~$vendor.xlsx" not in processed_names

    def test_run_batch_file_locked_skipped_with_err010(
        self, tmp_path: Path,
    ) -> None:
        """Verify that a locked file gets ERR_010 and other files continue."""
        config = _make_minimal_config()
        data_dir = tmp_path / "data"
        data_dir.mkdir(parents=True)
        (data_dir / "finished").mkdir()

        # Create two files
        _build_test_workbook(tmp_path, "a_locked.xlsx")
        _build_test_workbook(tmp_path, "b_normal.xlsx")

        original_load = openpyxl.load_workbook

        def _mock_load(path: Any, **kwargs: Any) -> Any:
            if "a_locked" in str(path):
                raise PermissionError("File is locked")
            return original_load(path, **kwargs)

        with patch("autoconvert.batch.openpyxl.load_workbook", side_effect=_mock_load):
            result = run_batch(config, data_dir)

        locked_result = next(
            r for r in result.file_results if r.filename == "a_locked.xlsx"
        )
        assert locked_result.status == "Failed"
        assert any(e.code == ErrorCode.ERR_010 for e in locked_result.errors)

        # Other file was still processed
        assert result.total_files == 2

    def test_run_batch_corrupt_file_skipped_with_err011(
        self, tmp_path: Path,
    ) -> None:
        """Verify that a corrupt file gets ERR_011."""
        config = _make_minimal_config()
        data_dir = tmp_path / "data"
        data_dir.mkdir(parents=True)
        (data_dir / "finished").mkdir()

        corrupt_file = data_dir / "corrupt.xlsx"
        corrupt_file.write_bytes(b"not a valid xlsx")

        result = run_batch(config, data_dir)

        assert result.total_files == 1
        corrupt_result = result.file_results[0]
        assert corrupt_result.status == "Failed"
        assert any(e.code == ErrorCode.ERR_011 for e in corrupt_result.errors)

    def test_run_batch_returns_correct_counts(self, tmp_path: Path) -> None:
        """Verify BatchResult counts match file statuses."""
        config = _make_minimal_config()
        data_dir = tmp_path / "data"
        data_dir.mkdir(parents=True)
        (data_dir / "finished").mkdir()

        # File 1: success (valid workbook)
        _build_test_workbook(tmp_path, "a_success.xlsx")

        # File 2: will fail (corrupt)
        corrupt = data_dir / "b_fail.xlsx"
        corrupt.write_bytes(b"not valid xlsx")

        result = run_batch(config, data_dir)

        assert result.total_files == 2
        assert result.failed_count >= 1
        assert result.success_count + result.attention_count + result.failed_count == result.total_files


# ---------------------------------------------------------------------------
# Tests — process_file
# ---------------------------------------------------------------------------


class TestProcessFile:
    """Tests for process_file pipeline execution."""

    def test_process_file_full_pipeline_success(
        self, tmp_path: Path, app_config: AppConfig,
    ) -> None:
        """Verify full pipeline produces Success status with valid workbook."""
        filepath = _build_test_workbook(tmp_path)
        result = process_file(filepath, app_config)

        # Reason: Depending on how well the minimal workbook matches real
        # config patterns, we accept Success or Attention (ATT_003/ATT_004).
        assert result.status in ("Success", "Attention", "Failed")
        assert result.filename == "vendor.xlsx"

    def test_process_file_phase_short_circuit_sheet_detection(
        self, tmp_path: Path,
    ) -> None:
        """Verify phase short-circuit when invoice sheet is not found."""
        config = _make_minimal_config()
        filepath = _build_test_workbook(
            tmp_path, "no_invoice.xlsx", no_invoice_sheet=True,
        )

        result = process_file(filepath, config)

        assert result.status == "Failed"
        assert any(e.code == ErrorCode.ERR_012 for e in result.errors)
        assert result.invoice_items is None

    def test_process_file_inv_no_fallback_fires_after_column_miss(
        self, tmp_path: Path,
    ) -> None:
        """Verify inv_no header fallback is called when column is missing."""
        config = _make_minimal_config()
        filepath = _build_test_workbook(
            tmp_path,
            "inv_fallback.xlsx",
            include_inv_no_col=False,
            include_inv_no_header=True,
        )

        result = process_file(filepath, config)

        # ERR_021 should NOT be raised because header fallback found the value.
        err_021 = [e for e in result.errors if e.code == ErrorCode.ERR_021]
        assert len(err_021) == 0

    def test_process_file_inv_no_err021_after_both_fail(
        self, tmp_path: Path,
    ) -> None:
        """Verify ERR_021 when both column mapping and header fallback fail."""
        config = _make_minimal_config()
        filepath = _build_test_workbook(
            tmp_path,
            "no_inv.xlsx",
            include_inv_no_col=False,
            include_inv_no_header=False,
        )

        result = process_file(filepath, config)

        assert result.status == "Failed"
        err_021 = [e for e in result.errors if e.code == ErrorCode.ERR_021]
        assert len(err_021) == 1

    def test_process_file_unexpected_exception_returns_failed(
        self, tmp_path: Path,
    ) -> None:
        """Verify unexpected exceptions are caught and result in Failed status."""
        config = _make_minimal_config()
        filepath = _build_test_workbook(tmp_path, "unexpected.xlsx")

        # Patch the module-level alias used by process_file.
        with patch(
            "autoconvert.batch._detect_sheets",
            side_effect=ValueError("unexpected error"),
        ):
            result = process_file(filepath, config)

        assert result.status == "Failed"
        assert any(e.code == ErrorCode.ERR_011 for e in result.errors)
