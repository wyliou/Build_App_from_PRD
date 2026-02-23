"""Tests for autoconvert.output.

Covers FR-029 (template population) and FR-030 (output file save).
All tests use in-memory workbooks — no real corpus files are read.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from autoconvert.errors import ErrorCode, ProcessingError
from autoconvert.models import AppConfig, InvNoCellConfig, InvoiceItem, PackingTotals
from autoconvert.output import write_template

# ---------------------------------------------------------------------------
# Helpers / fixtures
# ---------------------------------------------------------------------------

_EMPTY_INV_NO_CELL = InvNoCellConfig(
    patterns=[],
    label_patterns=[],
    exclude_patterns=[],
)

_SHEET_NAME = "工作表1"


def _make_app_config(template_path: Path) -> AppConfig:
    """Build a minimal AppConfig pointing to the given template path.

    Args:
        template_path: Path to the output template xlsx file.

    Returns:
        AppConfig with all required fields populated.
    """
    return AppConfig(
        invoice_sheet_patterns=[],
        packing_sheet_patterns=[],
        invoice_columns={},
        packing_columns={},
        inv_no_cell=_EMPTY_INV_NO_CELL,
        currency_lookup={},
        country_lookup={},
        output_template_path=template_path,
        invoice_min_headers=7,
        packing_min_headers=4,
    )


def _make_template_workbook() -> openpyxl.Workbook:
    """Create an in-memory workbook that mimics the real output template.

    Contains sheet '工作表1' with 4 header rows (1-4) populated with
    placeholder text so that header preservation tests can verify them.

    Returns:
        An openpyxl Workbook ready to use as a mock template.
    """
    wb = openpyxl.Workbook()
    # Rename default sheet
    ws = wb.active
    ws.title = _SHEET_NAME  # type: ignore[union-attr]
    # Populate 4 header rows with recognizable placeholder values
    for row in range(1, 5):
        for col in range(1, 41):
            ws.cell(row=row, column=col).value = f"H{row}C{col}"
    return wb


def _make_invoice_item(
    *,
    part_no: str = "P001",
    po_no: str = "PO12345",
    qty: str = "10",
    price: str = "5.50",
    amount: str = "55.00",
    currency: str = "502",
    coo: str = "116",
    brand: str = "TestBrand",
    brand_type: str = "0",
    model: str = "ModelX",
    inv_no: str | None = "INV-001",
    serial: str | None = "1",
    allocated_weight: str | None = "2.50000",
) -> InvoiceItem:
    """Construct an InvoiceItem with sensible defaults.

    Args:
        part_no: Part number string.
        po_no: Purchase order number.
        qty: Quantity as string (converted to Decimal).
        price: Unit price as string (converted to Decimal).
        amount: Line total as string (converted to Decimal).
        currency: Currency code string.
        coo: Country-of-origin code string.
        brand: Brand name.
        brand_type: Brand type code.
        model: Model identifier.
        inv_no: Invoice number (may be None).
        serial: Serial number (may be None).
        allocated_weight: Net weight string (may be None).

    Returns:
        InvoiceItem instance.
    """
    return InvoiceItem(
        part_no=part_no,
        po_no=po_no,
        qty=Decimal(qty),
        price=Decimal(price),
        amount=Decimal(amount),
        currency=currency,
        coo=coo,
        cod=None,
        brand=brand,
        brand_type=brand_type,
        model=model,
        inv_no=inv_no,
        serial=serial,
        allocated_weight=Decimal(allocated_weight) if allocated_weight is not None else None,
    )


def _make_packing_totals(
    *,
    total_gw: str = "100.0",
    total_nw: str = "90.0",
    total_packets: int | None = 5,
) -> PackingTotals:
    """Construct a PackingTotals with sensible defaults.

    Args:
        total_gw: Total gross weight string.
        total_nw: Total net weight string.
        total_packets: Packet count, or None.

    Returns:
        PackingTotals instance.
    """
    return PackingTotals(
        total_nw=Decimal(total_nw),
        total_nw_precision=2,
        total_gw=Decimal(total_gw),
        total_gw_precision=2,
        total_packets=total_packets,
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_write_template_success_all_columns(tmp_path: Path) -> None:
    """Happy path: 2 items, all columns verified for rows 5 and 6."""
    # Build in-memory template and save it to tmp_path
    template_path = tmp_path / "template.xlsx"
    _make_template_workbook().save(template_path)

    output_path = tmp_path / "result.xlsx"
    config = _make_app_config(template_path)

    item1 = _make_invoice_item(
        part_no="PART-A",
        po_no="PO-001",
        qty="10",
        price="5.00",
        amount="50.00",
        currency="502",
        coo="116",
        brand="BrandA",
        brand_type="0",
        model="ModelA",
        inv_no="INV-100",
        serial="1",
        allocated_weight="3.00000",
    )
    item2 = _make_invoice_item(
        part_no="PART-B",
        po_no="PO-002",
        qty="20",
        price="2.50",
        amount="50.00",
        currency="303",
        coo="142",
        brand="BrandB",
        brand_type="1",
        model="ModelB",
        inv_no="INV-100",
        serial="2",
        allocated_weight="7.00000",
    )
    totals = _make_packing_totals(total_gw="10.5", total_packets=3)

    write_template([item1, item2], totals, config, output_path)

    assert output_path.exists()
    wb = openpyxl.load_workbook(output_path)
    ws = wb[_SHEET_NAME]

    # Row 5 — item1
    assert ws.cell(row=5, column=1).value == "PART-A"   # A part_no
    assert ws.cell(row=5, column=2).value == "PO-001"   # B po_no
    assert ws.cell(row=5, column=3).value == "3"         # C fixed
    assert ws.cell(row=5, column=4).value == "502"       # D currency
    assert ws.cell(row=5, column=5).value == pytest.approx(10.0)  # E qty
    assert ws.cell(row=5, column=6).value == pytest.approx(5.0)   # F price
    assert ws.cell(row=5, column=7).value == pytest.approx(50.0)  # G amount
    assert ws.cell(row=5, column=8).value == "116"       # H coo
    assert ws.cell(row=5, column=12).value == "1"        # L serial
    assert ws.cell(row=5, column=13).value == pytest.approx(3.0)  # M net_weight
    assert ws.cell(row=5, column=14).value == "INV-100"  # N inv_no
    assert ws.cell(row=5, column=16).value == pytest.approx(10.5)  # P total_gw
    assert ws.cell(row=5, column=18).value == "32052"   # R fixed
    assert ws.cell(row=5, column=19).value == "320506"  # S fixed
    assert ws.cell(row=5, column=20).value == "142"     # T fixed
    assert ws.cell(row=5, column=37).value == 3         # AK total_packets
    assert ws.cell(row=5, column=38).value == "BrandA"  # AL brand
    assert ws.cell(row=5, column=39).value == "0"       # AM brand_type
    assert ws.cell(row=5, column=40).value == "ModelA"  # AN model

    # Row 6 — item2
    assert ws.cell(row=6, column=1).value == "PART-B"
    assert ws.cell(row=6, column=4).value == "303"
    assert ws.cell(row=6, column=8).value == "142"
    assert ws.cell(row=6, column=38).value == "BrandB"
    assert ws.cell(row=6, column=40).value == "ModelB"
    # P and AK must be empty in row 6
    assert ws.cell(row=6, column=16).value is None  # P empty
    assert ws.cell(row=6, column=37).value is None  # AK empty


def test_write_template_fixed_columns_all_rows(tmp_path: Path) -> None:
    """Fixed columns C, R, S, T must contain correct strings in all 3 data rows."""
    template_path = tmp_path / "template.xlsx"
    _make_template_workbook().save(template_path)
    output_path = tmp_path / "result.xlsx"
    config = _make_app_config(template_path)

    items = [
        _make_invoice_item(part_no=f"P{i}", serial=str(i)) for i in range(3)
    ]
    totals = _make_packing_totals()

    write_template(items, totals, config, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[_SHEET_NAME]

    for row in (5, 6, 7):
        assert ws.cell(row=row, column=3).value == "3", f"C in row {row}"
        assert ws.cell(row=row, column=18).value == "32052", f"R in row {row}"
        assert ws.cell(row=row, column=19).value == "320506", f"S in row {row}"
        assert ws.cell(row=row, column=20).value == "142", f"T in row {row}"


def test_write_template_total_gw_and_packets_row5_only(tmp_path: Path) -> None:
    """total_gw (col P) and total_packets (col AK) are written to row 5 only."""
    template_path = tmp_path / "template.xlsx"
    _make_template_workbook().save(template_path)
    output_path = tmp_path / "result.xlsx"
    config = _make_app_config(template_path)

    items = [_make_invoice_item(serial=str(i)) for i in range(3)]
    totals = _make_packing_totals(total_gw="50.0", total_packets=10)

    write_template(items, totals, config, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[_SHEET_NAME]

    # Row 5 — both should be populated
    assert ws.cell(row=5, column=16).value == pytest.approx(50.0)
    assert ws.cell(row=5, column=37).value == 10

    # Rows 6 and 7 — both must be empty (None)
    for row in (6, 7):
        assert ws.cell(row=row, column=16).value is None, f"P in row {row} must be None"
        assert ws.cell(row=row, column=37).value is None, f"AK in row {row} must be None"


def test_write_template_null_total_packets(tmp_path: Path) -> None:
    """When total_packets is None, column AK in row 5 must remain None (not written)."""
    template_path = tmp_path / "template.xlsx"
    _make_template_workbook().save(template_path)
    output_path = tmp_path / "result.xlsx"
    config = _make_app_config(template_path)

    items = [_make_invoice_item()]
    totals = _make_packing_totals(total_packets=None)

    write_template(items, totals, config, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[_SHEET_NAME]

    assert ws.cell(row=5, column=37).value is None


def test_write_template_template_load_failed_raises_err051(tmp_path: Path) -> None:
    """A non-existent template path must raise ProcessingError with code ERR_051."""
    missing_path = tmp_path / "does_not_exist.xlsx"
    config = _make_app_config(missing_path)
    output_path = tmp_path / "result.xlsx"
    items = [_make_invoice_item()]
    totals = _make_packing_totals()

    with pytest.raises(ProcessingError) as exc_info:
        write_template(items, totals, config, output_path)

    assert exc_info.value.code == ErrorCode.ERR_051


def test_write_template_save_failure_raises_err052(tmp_path: Path) -> None:
    """A PermissionError during save must raise ProcessingError with code ERR_052."""
    template_path = tmp_path / "template.xlsx"
    _make_template_workbook().save(template_path)
    output_path = tmp_path / "result.xlsx"
    config = _make_app_config(template_path)

    items = [_make_invoice_item()]
    totals = _make_packing_totals()

    with patch.object(openpyxl.Workbook, "save", side_effect=PermissionError("locked")):
        with pytest.raises(ProcessingError) as exc_info:
            write_template(items, totals, config, output_path)

    assert exc_info.value.code == ErrorCode.ERR_052


def test_write_template_attention_passthrough_raw_currency_coo(tmp_path: Path) -> None:
    """Raw (unconverted) currency and COO strings are written as-is to D and H."""
    template_path = tmp_path / "template.xlsx"
    _make_template_workbook().save(template_path)
    output_path = tmp_path / "result.xlsx"
    config = _make_app_config(template_path)

    item = _make_invoice_item(currency="USD-UNKNOWN", coo="XYZ-UNKNOWN")
    totals = _make_packing_totals()

    write_template([item], totals, config, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[_SHEET_NAME]

    assert ws.cell(row=5, column=4).value == "USD-UNKNOWN"   # D currency
    assert ws.cell(row=5, column=8).value == "XYZ-UNKNOWN"   # H coo


def test_write_template_preserves_header_rows(tmp_path: Path) -> None:
    """Rows 1-4 in the output must be identical to the template's rows 1-4."""
    template_path = tmp_path / "template.xlsx"
    original_wb = _make_template_workbook()
    original_wb.save(template_path)

    # Capture original header values
    original_ws = original_wb[_SHEET_NAME]
    original_headers: dict[tuple[int, int], object] = {}
    for row in range(1, 5):
        for col in range(1, 41):
            original_headers[(row, col)] = original_ws.cell(row=row, column=col).value

    output_path = tmp_path / "result.xlsx"
    config = _make_app_config(template_path)
    items = [_make_invoice_item()]
    totals = _make_packing_totals()

    write_template(items, totals, config, output_path)

    result_wb = openpyxl.load_workbook(output_path)
    result_ws = result_wb[_SHEET_NAME]

    for (row, col), expected in original_headers.items():
        actual = result_ws.cell(row=row, column=col).value
        assert actual == expected, (
            f"Header mismatch at row={row}, col={col}: "
            f"expected {expected!r}, got {actual!r}"
        )
