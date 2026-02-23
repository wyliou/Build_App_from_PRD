"""Shared test fixtures for AutoConvert test suite.

Provides session-scoped config fixture, function-scoped workbook factory,
and temporary output directory fixture.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Callable

import openpyxl
import pytest

from autoconvert.config import load_config
from autoconvert.models import AppConfig


@pytest.fixture(scope="session")
def app_config() -> AppConfig:
    """Load application config once per test session.

    Reads from the project's config/ directory (same one used by the
    actual application). Returns the fully validated AppConfig.

    Returns:
        AppConfig instance with all patterns compiled and lookups built.
    """
    config_dir = Path(__file__).parent.parent / "config"
    return load_config(config_dir)


@pytest.fixture()
def make_workbook() -> Callable[..., openpyxl.Workbook]:
    """Factory fixture for creating in-memory openpyxl workbooks.

    Creates minimal workbooks for test scenarios. Each workbook is
    function-scoped (fresh instance per test).

    Returns:
        Callable that creates a Workbook with sheet names and optional data.
        Signature: make_workbook(sheets: list[str] | None = None) -> Workbook
    """

    def _make_workbook(
        sheets: list[str] | None = None,
        data: dict[str, list[list[Any]]] | None = None,
    ) -> openpyxl.Workbook:
        """Create an in-memory workbook with specified sheet names and data.

        Args:
            sheets: List of sheet names. If None, creates one default sheet.
            data: Dict mapping sheet names to 2D lists of cell values.

        Returns:
            openpyxl.Workbook with the specified structure.
        """
        wb = openpyxl.Workbook()
        # Remove default sheet if custom sheets are provided
        if sheets:
            default_sheet = wb.active
            if default_sheet is not None and default_sheet.title not in sheets:
                wb.remove(default_sheet)
            for sheet_name in sheets:
                ws = wb.create_sheet(title=sheet_name)
                # Populate with data if provided
                if data and sheet_name in data:
                    for row_idx, row_data in enumerate(data[sheet_name], start=1):
                        for col_idx, cell_value in enumerate(row_data, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=cell_value)
        else:
            # Just populate the default sheet with data if provided
            ws = wb.active
            if ws is not None and data and "Sheet" in data:
                for row_idx, row_data in enumerate(data["Sheet"], start=1):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)

        return wb

    return _make_workbook


@pytest.fixture()
def tmp_output_dir(tmp_path: Path) -> Path:
    """Provide a clean temporary output directory for each test.

    Args:
        tmp_path: pytest built-in tmp_path fixture.

    Returns:
        Path to a temporary output directory.
    """
    output_dir = tmp_path / "finished"
    output_dir.mkdir()
    return output_dir
