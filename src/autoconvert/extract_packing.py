"""Packing data extraction for AutoConvert.

Extracts packing items with merge-aware continuation, ditto mark handling,
total row detection, and totals extraction.
Implements FR-012, FR-013, FR-014, FR-015, FR-016, FR-017.
"""

from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation

from openpyxl.worksheet.worksheet import Worksheet

from autoconvert.errors import ErrorCode, ProcessingError
from autoconvert.merge_tracker import MergeTracker
from autoconvert.models import ColumnMapping, PackingItem
from autoconvert.utils import (
    DITTO_MARKS,
    detect_cell_precision,
    is_cell_empty,
    is_stop_keyword,
    parse_numeric,
    round_half_up,
    strip_unit_suffix,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# FR-012 — Extract Packing Items
# ---------------------------------------------------------------------------


def _read_numeric_field(
    value: object,
    field_name: str,
    row: int,
) -> Decimal:
    """Parse a cell value as Decimal with unit-suffix stripping.

    Args:
        value: Raw cell value from openpyxl.
        field_name: Field name for error context.
        row: 1-based row number for error context.

    Returns:
        Decimal representation of the value.

    Raises:
        ProcessingError: ERR_031 when the value cannot be parsed.
    """
    if isinstance(value, str):
        cleaned = strip_unit_suffix(value.strip())
        try:
            return Decimal(cleaned)
        except (InvalidOperation, ValueError, TypeError) as exc:
            raise ProcessingError(
                code=ErrorCode.ERR_031,
                message=(
                    f"Invalid numeric value '{value}' at row {row}, "
                    f"field '{field_name}'"
                ),
                row=row,
                field=field_name,
            ) from exc
    return parse_numeric(value, field_name, row)


def _check_stop_conditions(
    sheet: Worksheet,
    row: int,
    part_no_col: int,
    nw_col: int,
    gw_col: int,
    merge_tracker: MergeTracker,
    has_first_data_row: bool,
) -> bool:
    """Check whether the current row triggers extraction termination.

    Implements the three stop conditions from FR-012 in order:
    1. Keyword stop (any cell A-J contains total/合计/总计/小计)
    2. Blank row (all key columns empty, after first data row)
    3. Implicit total row (empty part_no + NW>0 + GW>0, excluding merges)

    Args:
        sheet: The packing worksheet.
        row: Current 1-based row number.
        part_no_col: 1-based column index for part_no.
        nw_col: 1-based column index for nw.
        gw_col: 1-based column index for gw.
        merge_tracker: MergeTracker for merge-aware checks.
        has_first_data_row: Whether we have already extracted at least one row.

    Returns:
        True if extraction should stop at this row.
    """
    # Stop condition 1: keyword in columns A-J
    for col in range(1, 11):
        cell_val = sheet.cell(row=row, column=col).value
        if isinstance(cell_val, str) and is_stop_keyword(cell_val):
            return True

    # Read key column values for conditions 2 and 3
    part_raw = sheet.cell(row=row, column=part_no_col).value
    nw_raw = sheet.cell(row=row, column=nw_col).value
    gw_raw = sheet.cell(row=row, column=gw_col).value

    part_empty = is_cell_empty(part_raw)
    nw_empty = is_cell_empty(nw_raw)
    gw_empty = is_cell_empty(gw_raw)

    # Determine if nw_raw and gw_raw are ditto marks (treated as empty for stop)
    nw_is_ditto = isinstance(nw_raw, str) and nw_raw.strip() in DITTO_MARKS
    gw_is_ditto = isinstance(gw_raw, str) and gw_raw.strip() in DITTO_MARKS

    # Stop condition 3: implicit total row (checked before blank to handle
    # the case where empty part_no + NW>0 + GW>0 is an implicit total)
    # part_no empty AND nw > 0 AND gw > 0, excluding merge continuations
    if part_empty and not nw_empty and not nw_is_ditto and not gw_empty and not gw_is_ditto:
        # Exclude rows where part_no is empty due to vertical merge
        if merge_tracker.is_in_merge(row, part_no_col) and not merge_tracker.is_merge_anchor(row, part_no_col):
            # Merge continuation — not a total row
            pass
        else:
            # Check if nw and gw are both numeric and positive
            try:
                nw_val = _read_numeric_field(nw_raw, "nw", row)
                gw_val = _read_numeric_field(gw_raw, "gw", row)
                if nw_val > 0 and gw_val > 0:
                    return True
            except ProcessingError:
                pass  # Not numeric — not an implicit total

    # Re-check stop condition 2 (blank row) after implicit total check
    if has_first_data_row and part_empty and (nw_empty or nw_is_ditto) and (gw_empty or gw_is_ditto):
        # Exclude merge continuations for part_no
        if merge_tracker.is_in_merge(row, part_no_col) and not merge_tracker.is_merge_anchor(row, part_no_col):
            return False  # Not blank — merge continuation
        return True

    return False


def extract_packing_items(
    sheet: Worksheet,
    column_map: ColumnMapping,
    merge_tracker: MergeTracker,
) -> tuple[list[PackingItem], int]:
    """Extract packing items (part_no, qty, nw) from the packing sheet.

    Processes rows starting at effective_header_row + 1. Checks stop
    conditions BEFORE blank check (differs from invoice extraction).
    Handles merged NW/QTY cells, implicit continuation (same part_no,
    empty NW), ditto marks (NW=0), and vertically merged part_no.
    Skips pallet/header-continuation rows.

    Args:
        sheet: The packing worksheet (already unmerged by MergeTracker).
        column_map: Column mapping with field_map and effective_header_row.
        merge_tracker: MergeTracker with pre-unmerge merge info.

    Returns:
        Tuple of (list of PackingItem, last_data_row).

    Raises:
        ProcessingError: ERR_030 for empty required fields,
            ERR_031 for invalid numeric values.
    """
    field_map = column_map.field_map
    part_no_col = field_map["part_no"]
    qty_col = field_map["qty"]
    nw_col = field_map["nw"]
    gw_col = field_map["gw"]
    start_row = column_map.effective_header_row + 1

    items: list[PackingItem] = []
    last_data_row = start_row
    has_first_data_row = False
    prev_part_no: str | None = None

    for row in range(start_row, sheet.max_row + 1):  # type: ignore[operator]
        # --- STOP CONDITIONS FIRST (CRITICAL ordering) ---
        if _check_stop_conditions(
            sheet, row, part_no_col, nw_col, gw_col,
            merge_tracker, has_first_data_row,
        ):
            break

        # --- Read raw values ---
        part_raw = sheet.cell(row=row, column=part_no_col).value
        qty_raw = sheet.cell(row=row, column=qty_col).value
        nw_raw = sheet.cell(row=row, column=nw_col).value

        # --- Part_no handling (merge propagation) ---
        part_empty = is_cell_empty(part_raw)
        is_part_merge_continuation = (
            merge_tracker.is_in_merge(row, part_no_col)
            and not merge_tracker.is_merge_anchor(row, part_no_col)
        )

        if part_empty and is_part_merge_continuation:
            # Propagate anchor value for vertically merged part_no
            part_no = str(
                merge_tracker.get_anchor_value(sheet, row, part_no_col)
            ).strip()
        elif part_empty:
            # Truly empty part_no — not a merge continuation
            # Check if this is an implicit continuation (same part_no as prev)
            # If not, it's an error
            if prev_part_no is not None:
                # Could be a row we skip (e.g., PO-reference row filtered by
                # qty=0 AND nw=0 later). But if nw is not empty and not zero,
                # this is an error.
                pass
            # We'll handle below after reading nw/qty
            part_no = ""
        else:
            part_no = str(part_raw).strip()

        # --- FILTERING: header continuation ---
        if part_no and "part no" in part_no.lower():
            continue

        # --- FILTERING: pallet/summary rows ---
        # Reason: Pallet/summary rows can appear with pallet keywords in
        # part_no OR with non-numeric text in NW (e.g., "7 Pallets", "棧板",
        # "12PLT", "KGS"). Both forms must be skipped.
        _PALLET_KEYWORDS = ("plt.", "pallet", "pallets", "棧板", "栈板", "plt")
        part_lower = part_no.lower() if part_no else ""
        if part_lower and any(kw in part_lower for kw in _PALLET_KEYWORDS):
            continue
        # Also skip rows where the NW cell is a string that cannot be parsed
        # and contains pallet/unit-label text — these are summary rows.
        if isinstance(nw_raw, str) and not is_cell_empty(nw_raw):
            nw_stripped_lower = nw_raw.strip().lower()
            if any(kw in nw_stripped_lower for kw in _PALLET_KEYWORDS):
                continue
            # Reason: Pure unit labels like "KGS" that slipped through header
            # detection indicate a sub-header or label row, not data.
            if nw_stripped_lower in ("kgs", "kgs.", "kg", "lbs", "lb"):
                continue

        # --- NW handling (merge, continuation, ditto) ---
        nw_empty = is_cell_empty(nw_raw)
        nw_is_ditto = isinstance(nw_raw, str) and nw_raw.strip() in DITTO_MARKS
        is_nw_merge_non_anchor = (
            merge_tracker.is_in_merge(row, nw_col)
            and not merge_tracker.is_merge_anchor(row, nw_col)
        )

        is_first_row_of_merge = True
        nw: Decimal

        if nw_is_ditto:
            # Ditto mark: NW=0, not first row of merge
            nw = Decimal("0")
            is_first_row_of_merge = False
        elif is_nw_merge_non_anchor:
            # Non-anchor row of merged NW cell
            nw = Decimal("0")
            is_first_row_of_merge = False
        elif nw_empty:
            # Implicit continuation: empty NW, not in merge
            # Check if same part_no as previous item
            if prev_part_no is not None and part_no == prev_part_no:
                nw = Decimal("0")
                is_first_row_of_merge = False
            elif part_empty and prev_part_no is not None:
                # Empty part_no with empty nw — likely PO-reference row
                nw = Decimal("0")
                is_first_row_of_merge = False
            else:
                # Truly empty NW that is not continuation — error
                raise ProcessingError(
                    code=ErrorCode.ERR_030,
                    message=(
                        f"Empty required field 'nw' at row {row}"
                    ),
                    row=row,
                    field="nw",
                )
        else:
            # Normal NW value
            nw = _read_numeric_field(nw_raw, "nw", row)
            nw = round_half_up(nw, 5)

        # --- QTY handling (merge, continuation) ---
        qty_empty = is_cell_empty(qty_raw)
        is_qty_merge_non_anchor = (
            merge_tracker.is_in_merge(row, qty_col)
            and not merge_tracker.is_merge_anchor(row, qty_col)
        )

        qty: Decimal

        if is_qty_merge_non_anchor:
            qty = Decimal("0")
        elif qty_empty:
            # Implicit continuation for qty
            if prev_part_no is not None and (
                part_no == prev_part_no or part_empty
            ):
                qty = Decimal("0")
            else:
                raise ProcessingError(
                    code=ErrorCode.ERR_030,
                    message=(
                        f"Empty required field 'qty' at row {row}"
                    ),
                    row=row,
                    field="qty",
                )
        else:
            qty = _read_numeric_field(qty_raw, "qty", row)
            precision = detect_cell_precision(
                qty_raw, sheet.cell(row=row, column=qty_col).number_format
            )
            qty = round_half_up(qty, precision)

        # --- Handle truly empty part_no (not merge continuation) ---
        if part_no == "" and part_empty:
            if prev_part_no is not None:
                # Implicit continuation — use previous part_no
                part_no = prev_part_no
            else:
                raise ProcessingError(
                    code=ErrorCode.ERR_030,
                    message=(
                        f"Empty required field 'part_no' at row {row}"
                    ),
                    row=row,
                    field="part_no",
                )

        # --- FILTERING: qty=0 AND nw=0 rows ---
        if qty == Decimal("0") and nw == Decimal("0"):
            continue

        # --- Build PackingItem ---
        item = PackingItem(
            part_no=part_no,
            qty=qty,
            nw=nw,
            is_first_row_of_merge=is_first_row_of_merge,
            row_number=row,
        )
        items.append(item)
        last_data_row = row
        has_first_data_row = True
        prev_part_no = part_no

    logger.info(
        "Packing sheet extracted %d items (rows %d-%d)",
        len(items),
        start_row,
        last_data_row,
    )

    return items, last_data_row


# ---------------------------------------------------------------------------
# FR-013 — Validate Merged Weight Cells
# ---------------------------------------------------------------------------


def validate_merged_weights(
    packing_items: list[PackingItem],
    merge_tracker: MergeTracker,
    column_map: ColumnMapping,
) -> None:
    """Validate that no merged NW/QTY cell is shared by different part_no values.

    For each merge range in the NW or QTY columns (data area only), checks
    that all PackingItem records sharing that merged cell have the same
    part_no. Raises ERR_046 if different part_no values share a merged cell.

    Args:
        packing_items: Extracted packing items from FR-012.
        merge_tracker: MergeTracker with pre-unmerge merge info.
        column_map: Column mapping with field_map and header_row.

    Raises:
        ProcessingError: ERR_046 when different parts share a merged weight cell.
    """
    field_map = column_map.field_map
    nw_col = field_map["nw"]
    qty_col = field_map["qty"]
    header_row = column_map.header_row

    # Build a lookup from row_number -> part_no
    row_to_part: dict[int, str] = {
        item.row_number: item.part_no for item in packing_items
    }

    # Check each item's NW and QTY columns for merge ranges
    checked_ranges: set[tuple[int, int, int, int]] = set()

    for item in packing_items:
        for col in (nw_col, qty_col):
            mr = merge_tracker.get_merge_range(item.row_number, col)
            if mr is None:
                continue
            # Only data-area merges (min_row > header_row)
            if mr.min_row <= header_row:
                continue
            range_key = (mr.min_row, mr.max_row, mr.min_col, mr.max_col)
            if range_key in checked_ranges:
                continue
            checked_ranges.add(range_key)

            # Collect all part_no values for items in this merge range
            parts_in_range: set[str] = set()
            for r in range(mr.min_row, mr.max_row + 1):
                if r in row_to_part:
                    parts_in_range.add(row_to_part[r])

            if len(parts_in_range) > 1:
                raise ProcessingError(
                    code=ErrorCode.ERR_046,
                    message=(
                        f"Different parts share merged weight cell "
                        f"(rows {mr.min_row}-{mr.max_row}): "
                        f"{sorted(parts_in_range)}"
                    ),
                    row=mr.min_row,
                    field="nw" if col == nw_col else "qty",
                )


# ---------------------------------------------------------------------------
# FR-014 through FR-017 — Total row detection and extraction
# Delegated to extract_totals_helpers.py to stay under 500-line limit.
# ---------------------------------------------------------------------------

from autoconvert.extract_totals_helpers import (  # noqa: E402
    detect_total_row,
    extract_totals,
)

__all__ = [
    "extract_packing_items",
    "validate_merged_weights",
    "detect_total_row",
    "extract_totals",
]
