"""Column mapping and header detection for AutoConvert.

Detects header rows using a three-tier priority system, maps
required/optional columns to their positions, and extracts invoice
numbers from the header area. Implements FR-007, FR-008, FR-009.

Error codes owned by this module:
    ERR_014 (HEADER_ROW_NOT_FOUND) — raised by detect_header_row
    ERR_020 (REQUIRED_COLUMN_MISSING) — raised by map_columns
"""

from __future__ import annotations

import logging
import re
from collections.abc import Mapping

from openpyxl.worksheet.worksheet import Worksheet

from autoconvert.errors import ErrorCode, ProcessingError
from autoconvert.models import AppConfig, ColumnMapping, FieldPattern
from autoconvert.utils import normalize_header

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_HEADER_KEYWORDS: frozenset[str] = frozenset({
    "qty", "n.w.", "g.w.", "part no", "amount", "price", "quantity",
    "weight", "品牌", "料号", "数量", "单价", "金额", "净重", "毛重",
    "原产", "country", "origin", "brand", "model", "description",
    "unit", "currency", "coo",
})
"""Header keywords for Tier-0 classification (FR-007)."""

_METADATA_MARKERS: tuple[str, ...] = (
    "Tel:", "Fax:", "Cust ID:", "Contact:", "Address:",
)
"""Metadata marker substrings for Tier-2 demotion (FR-007)."""

# Reason: First alternative matches pure numbers/decimals like "123", "45.67".
# Second alternative matches alphanumeric codes containing at least one digit
# (e.g., "ABC-123", "PT001") but NOT pure alphabetic words like "Price".
_NUMERIC_RE: re.Pattern[str] = re.compile(
    r"^[\d]+\.?[\d]*$|^(?=[A-Za-z0-9\-]*\d)[A-Za-z0-9\-]+$"
)

_HEADER_SCAN_ROW_START = 7
_HEADER_SCAN_ROW_END = 30
_HEADER_SCAN_COL_COUNT = 13
_MAP_COL_COUNT = 20
_INV_NO_SCAN_ROW_END = 15
_DATA_LIKE_THRESHOLD = 3


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def detect_header_row(
    sheet: Worksheet, sheet_type: str, config: AppConfig,
) -> int:
    """Scan rows 7-30 to find the header row using a three-tier priority system.

    Tier 0 = keyword-containing non-data rows (highest priority).
    Tier 1 = other qualifying rows (not metadata, not data-like).
    Tier 2 = metadata or data-like rows (lowest priority).

    Args:
        sheet: An openpyxl Worksheet (already unmerged by MergeTracker).
        sheet_type: Either "invoice" or "packing".
        config: Application configuration.

    Returns:
        1-based row number of the detected header row.

    Raises:
        ProcessingError: ERR_014 if no qualifying row is found.
    """
    threshold = (
        config.invoice_min_headers
        if sheet_type == "invoice"
        else config.packing_min_headers
    )
    best_tier = 3  # worse than any valid tier
    best_row: int | None = None

    for row in range(_HEADER_SCAN_ROW_START, _HEADER_SCAN_ROW_END + 1):
        cells = _collect_row_cells(sheet, row, _HEADER_SCAN_COL_COUNT)
        filtered = [c for c in cells if not c.startswith("Unnamed:")]
        if len(filtered) < threshold:
            continue

        has_meta = _has_metadata_markers(filtered)
        num_count = _count_numeric_cells(filtered)
        has_kw = _has_header_keywords(filtered)

        if has_kw and num_count < 2:
            tier = 0
        elif has_meta or num_count >= _DATA_LIKE_THRESHOLD:
            tier = 2
        else:
            tier = 1

        if tier < best_tier or (tier == best_tier and best_row is None):
            best_tier = tier
            best_row = row

    if best_row is None:
        raise ProcessingError(
            code=ErrorCode.ERR_014,
            message=(
                f"Header row not found in {sheet_type} sheet: "
                f"no row in range 7-30 meets the minimum threshold "
                f"of {threshold} non-empty cells"
            ),
        )
    logger.info(
        "Detected header row %d (tier %d) for %s sheet",
        best_row, best_tier, sheet_type,
    )
    return best_row


def map_columns(
    sheet: Worksheet, header_row: int, sheet_type: str, config: AppConfig,
) -> ColumnMapping:
    """Map column headers to field names using regex patterns from config.

    Scans up to 20 columns in the detected header row, normalizes each
    cell via normalize_header(), and matches against field patterns.
    Handles two-row merged headers via sub-header fallback and currency
    data-row fallback for invoice sheets.

    Args:
        sheet: An openpyxl Worksheet (already unmerged).
        header_row: 1-based header row number from detect_header_row.
        sheet_type: Either "invoice" or "packing".
        config: Application configuration with compiled field patterns.

    Returns:
        ColumnMapping with field_map, header_row, and effective_header_row.

    Raises:
        ProcessingError: ERR_020 listing all missing required fields
            after all fallback attempts.
    """
    field_defs = (
        config.invoice_columns if sheet_type == "invoice"
        else config.packing_columns
    )

    # Step 1: Primary header row scan (columns 1-20).
    field_map: dict[str, int] = {}
    _scan_row_for_fields(sheet, header_row, field_defs, field_map)

    # Step 2: Sub-header fallback if required fields remain unmapped.
    effective_header_row = header_row
    if _get_missing_required(field_defs, field_map):
        sub_row = header_row + 1
        sub_cells = _collect_row_cells(sheet, sub_row, _MAP_COL_COUNT)
        if _count_numeric_cells(sub_cells) < _DATA_LIKE_THRESHOLD:
            remaining = {k: v for k, v in field_defs.items() if k not in field_map}
            sub_map: dict[str, int] = {}
            _scan_row_for_fields(sheet, sub_row, remaining, sub_map)
            if sub_map:
                field_map.update(sub_map)
                effective_header_row = sub_row
                logger.debug(
                    "Sub-header fallback advanced effective_header_row to %d",
                    sub_row,
                )

    # Step 3: Currency data-row fallback (invoice only).
    if sheet_type == "invoice" and "currency" not in field_map:
        _currency_data_row_fallback(sheet, header_row, field_map, config)

    # Step 4: Check for missing required fields after all fallbacks.
    still_missing = _get_missing_required(field_defs, field_map)
    if still_missing:
        raise ProcessingError(
            code=ErrorCode.ERR_020,
            message=(
                f"Required columns missing in {sheet_type} sheet: "
                f"{', '.join(sorted(still_missing))}"
            ),
        )

    logger.info(
        "Mapped %d columns for %s sheet (header=%d, effective=%d)",
        len(field_map), sheet_type, header_row, effective_header_row,
    )
    return ColumnMapping(
        sheet_type=sheet_type,
        field_map=field_map,
        header_row=header_row,
        effective_header_row=effective_header_row,
    )


def extract_inv_no_from_header(
    sheet: Worksheet, config: AppConfig,
) -> str | None:
    """Search rows 1-15 for invoice number using capture-group and label patterns.

    Tries capture-group patterns first, then label patterns with adjacent-cell
    lookup (right up to +3 cols, below row+1 and row+2). Filters false
    positives via exclude patterns and cleans "INV#"/"NO." prefixes.

    Args:
        sheet: An openpyxl Worksheet for the invoice sheet.
        config: Application configuration with inv_no_cell patterns.

    Returns:
        Extracted and cleaned invoice number string, or None if not found.
        Does NOT raise ERR_021 (that is the batch orchestrator's responsibility).
    """
    inv_cfg = config.inv_no_cell

    # Pass 1: Capture-group patterns.
    for row in range(1, _INV_NO_SCAN_ROW_END + 1):
        for col in range(1, _MAP_COL_COUNT + 1):
            cell_str = _read_cell_str(sheet, row, col)
            if not cell_str:
                continue
            for pat in inv_cfg.patterns:
                m = pat.search(cell_str)
                if m and m.group(1):
                    candidate = m.group(1).strip()
                    if _is_excluded(candidate, inv_cfg.exclude_patterns):
                        continue
                    return _clean_inv_no_prefix(candidate)

    # Pass 2: Label patterns with adjacent-cell lookup.
    for row in range(1, _INV_NO_SCAN_ROW_END + 1):
        for col in range(1, _MAP_COL_COUNT + 1):
            cell_str = _read_cell_str(sheet, row, col)
            if not cell_str:
                continue
            if not any(p.search(cell_str) for p in inv_cfg.label_patterns):
                continue

            # Search adjacent right (up to +3 columns).
            for offset in range(1, 4):
                cand = _read_cell_str(sheet, row, col + offset)
                if cand and not _is_excluded(cand, inv_cfg.exclude_patterns):
                    return _clean_inv_no_prefix(cand)

            # Search below: row+1 and row+2.
            for row_off in (1, 2):
                cand = _read_cell_str(sheet, row + row_off, col)
                if cand and not _is_excluded(cand, inv_cfg.exclude_patterns):
                    return _clean_inv_no_prefix(cand)

    logger.debug("No invoice number found in header area rows 1-15")
    return None


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _collect_row_cells(
    sheet: Worksheet, row: int, max_cols: int,
) -> list[str]:
    """Read non-empty cell string values from a row."""
    result: list[str] = []
    for col in range(1, max_cols + 1):
        raw = sheet.cell(row=row, column=col).value
        if raw is None:
            continue
        text = str(raw).strip()
        if text:
            result.append(text)
    return result


def _has_metadata_markers(cells: list[str]) -> bool:
    """Return True if any cell contains a metadata marker substring."""
    return any(marker in c for c in cells for marker in _METADATA_MARKERS)


def _count_numeric_cells(cells: list[str]) -> int:
    """Count cells matching the numeric/alphanumeric-code pattern."""
    return sum(1 for c in cells if _NUMERIC_RE.match(c))


def _has_header_keywords(cells: list[str]) -> bool:
    """Return True if any cell (lowercased) contains a header keyword."""
    return any(kw in c.lower() for c in cells for kw in _HEADER_KEYWORDS)


def _scan_row_for_fields(
    sheet: Worksheet,
    row: int,
    field_defs: Mapping[str, FieldPattern],
    field_map: dict[str, int],
) -> None:
    """Scan a row, match cells against field patterns. First match per field wins."""
    for col in range(1, _MAP_COL_COUNT + 1):
        raw = sheet.cell(row=row, column=col).value
        if raw is None:
            continue
        text = str(raw).strip()
        if not text:
            continue
        normalized = normalize_header(text)
        for field_name, fp in field_defs.items():
            if field_name in field_map:
                continue
            for pat in fp.patterns:
                if pat.search(normalized):
                    field_map[field_name] = col
                    break


def _get_missing_required(
    field_defs: Mapping[str, FieldPattern], field_map: dict[str, int],
) -> list[str]:
    """Return sorted list of required field names not present in field_map."""
    return sorted(
        name for name, fp in field_defs.items()
        if fp.required and name not in field_map
    )


def _currency_data_row_fallback(
    sheet: Worksheet,
    header_row: int,
    field_map: dict[str, int],
    config: AppConfig,
) -> None:
    """Scan data rows for embedded currency values (invoice only).

    Scans rows header_row+1 through header_row+4 for cells matching
    currency patterns. Shifts affected price/amount mappings when a
    currency value occupies their mapped column.

    Args:
        sheet: Worksheet to scan.
        header_row: 1-based detected header row.
        field_map: Mutable field_map to update.
        config: AppConfig with currency field patterns.
    """
    currency_patterns = config.invoice_columns["currency"].patterns

    for row in range(header_row + 1, header_row + 5):
        found = False
        for col in range(1, _MAP_COL_COUNT + 1):
            raw = sheet.cell(row=row, column=col).value
            if raw is None:
                continue
            text = str(raw).strip()
            if not text:
                continue
            if not any(p.search(text) for p in currency_patterns):
                continue

            # Shift price/amount if currency occupies their column.
            for shift_field in ("price", "amount"):
                if shift_field in field_map and field_map[shift_field] == col:
                    field_map[shift_field] = col + 1

            if "currency" not in field_map:
                field_map["currency"] = col
            found = True
            # Reason: Continue scanning remaining columns in the same row
            # for multi-column currency handling.

        if found:
            break


def _is_excluded(
    candidate: str,
    exclude_patterns: list[re.Pattern[str]],  # type: ignore[type-arg]
) -> bool:
    """Return True if candidate matches any exclude pattern."""
    return any(p.search(candidate) for p in exclude_patterns)


def _clean_inv_no_prefix(value: str) -> str:
    """Remove 'INV#' and 'NO.' prefixes from an invoice number."""
    cleaned = value
    if cleaned.upper().startswith("INV#"):
        cleaned = cleaned[4:]
    if cleaned.upper().startswith("NO."):
        cleaned = cleaned[3:]
    return cleaned.strip()


def _read_cell_str(sheet: Worksheet, row: int, col: int) -> str | None:
    """Read a cell as a stripped string, returning None if empty."""
    raw = sheet.cell(row=row, column=col).value
    if raw is None:
        return None
    text = str(raw).strip()
    return text if text else None
