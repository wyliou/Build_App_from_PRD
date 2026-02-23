"""Merged cell tracking and value propagation for AutoConvert.

Captures merged cell ranges before unmerging, then provides O(1) methods
to query merge status and propagate anchor values. Implements FR-010.

Key invariant: __init__ captures ALL merged ranges before any unmerging.
After __init__ completes the sheet has no merged cells, but this module
retains the pre-unmerge snapshot for downstream lookup.
"""

from __future__ import annotations

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from autoconvert.models import MergeRange


class MergeTracker:
    """Tracks merged cell ranges captured before sheet unmerging.

    On construction, snapshots all merged cell ranges from the provided
    worksheet, builds an O(1) cell-to-range lookup, then unmerges every
    cell in the sheet. Downstream extractors query this object to
    determine merge membership and retrieve anchor values.

    Attributes:
        _ranges: All captured MergeRange objects.
        _cell_to_range: Maps every (row, col) tuple within a merge range
            to its owning MergeRange for O(1) lookup.
    """

    def __init__(self, sheet: Worksheet) -> None:
        """Capture all merge ranges and unmerge the sheet.

        Steps (order is critical):
        1. Snapshot all current merged cell ranges as MergeRange objects.
        2. Build _cell_to_range lookup for every cell in each range.
        3. Unmerge all cells from the snapshot (never iterate the live
           collection while modifying it).

        After __init__ completes:
        - The sheet has no remaining merged cells.
        - This tracker holds the pre-unmerge snapshot for all queries.

        Args:
            sheet: The openpyxl Worksheet to process. Modified in-place
                (all merged cells are unmerged).
        """
        self._ranges: list[MergeRange] = []
        self._cell_to_range: dict[tuple[int, int], MergeRange] = {}

        # Step 1 & 2: snapshot ranges and build lookup BEFORE any unmerging.
        # sheet.merged_cells.ranges needs type: ignore due to incomplete stubs
        # (openpyxl MergedCellRange attributes are not fully typed in typeshed).
        raw_ranges = list(sheet.merged_cells.ranges)  # type: ignore[attr-defined]

        for rng in raw_ranges:
            # Reason: MergedCellRange has 1-based min/max attributes matching
            # openpyxl's native convention — store directly without adjustment.
            merge_range = MergeRange(
                min_row=rng.min_row,  # type: ignore[attr-defined]
                max_row=rng.max_row,  # type: ignore[attr-defined]
                min_col=rng.min_col,  # type: ignore[attr-defined]
                max_col=rng.max_col,  # type: ignore[attr-defined]
            )
            self._ranges.append(merge_range)

            # Populate every (row, col) cell within this range in the lookup.
            for r in range(merge_range.min_row, merge_range.max_row + 1):
                for c in range(merge_range.min_col, merge_range.max_col + 1):
                    self._cell_to_range[(r, c)] = merge_range

        # Step 3: unmerge from the snapshot list (not from the live collection)
        # to avoid mutating the iterable during iteration.
        for rng in raw_ranges:
            sheet.unmerge_cells(str(rng))

    def is_merge_anchor(self, row: int, col: int) -> bool:
        """Return True if (row, col) is the top-left cell of a merge range.

        The anchor cell is where openpyxl retains the merged value after
        unmerging. Non-anchor cells become empty after unmerge.

        Args:
            row: 1-based row index.
            col: 1-based column index.

        Returns:
            True if the cell is the anchor (min_row, min_col) of its
            merge range; False if it is a non-anchor cell in a range or
            if it is not in any merge range.
        """
        merge_range = self._cell_to_range.get((row, col))
        if merge_range is None:
            return False
        return row == merge_range.min_row and col == merge_range.min_col

    def is_in_merge(self, row: int, col: int) -> bool:
        """Return True if (row, col) falls within any captured merge range.

        Includes both anchor and non-anchor cells. Returns False for cells
        that were never part of a merge.

        Args:
            row: 1-based row index.
            col: 1-based column index.

        Returns:
            True if the cell is within any merged range (anchor or not);
            False otherwise.
        """
        return (row, col) in self._cell_to_range

    def get_anchor_value(self, sheet: Worksheet, row: int, col: int) -> Any:
        """Return the anchor cell's value for any cell in a merge range.

        Transparently returns the merged value regardless of whether the
        queried cell is the anchor or a non-anchor continuation cell.
        For cells not in any merge range, reads and returns the cell value
        directly (fall-through to normal cell read).

        After unmerging, non-anchor cells become empty (None). This method
        compensates by reading from the anchor cell (min_row, min_col).

        Args:
            sheet: The worksheet (must be the same sheet passed to __init__).
            row: 1-based row index.
            col: 1-based column index.

        Returns:
            The anchor cell's value if the cell is in a merge range, or
            the cell's own value if not in any merge range.
        """
        merge_range = self._cell_to_range.get((row, col))
        if merge_range is None:
            # Not in any merge — read cell directly.
            return sheet.cell(row=row, column=col).value
        # For both anchor and non-anchor cells, return the anchor's value.
        # The anchor is always at (min_row, min_col) of the range.
        return sheet.cell(row=merge_range.min_row, column=merge_range.min_col).value

    def get_merge_range(self, row: int, col: int) -> MergeRange | None:
        """Return the MergeRange containing (row, col), or None.

        Used by validate_merged_weights (FR-013) to find all rows sharing
        a merged NW cell, then checks their part_no values.

        Args:
            row: 1-based row index.
            col: 1-based column index.

        Returns:
            The MergeRange object that contains the cell, or None if the
            cell is not part of any merge range.
        """
        return self._cell_to_range.get((row, col))

    def is_data_area_merge(self, row: int, col: int, header_row: int) -> bool:
        """Return True if (row, col) is in a data-area merge range.

        A data-area merge starts after the header row (min_row > header_row).
        Header-area merges (min_row <= header_row) are formatting-only and
        return False. Cells not in any merge also return False.

        Used by extract_packing (FR-012, FR-014) to distinguish data
        continuation rows from total-row candidates.

        Args:
            row: 1-based row index.
            col: 1-based column index.
            header_row: 1-based row number of the sheet header row.

        Returns:
            True if the cell is in a merge range whose min_row > header_row;
            False if not in a merge range or if merge starts in header area.
        """
        merge_range = self._cell_to_range.get((row, col))
        if merge_range is None:
            return False
        return merge_range.min_row > header_row

    def get_first_row_of_merge(self, row: int, col: int) -> int:
        """Return the first row (min_row) of the merge range containing (row, col).

        Used by extract_packing to determine is_first_row_of_merge flag:
        ``is_first_row_of_merge = (get_first_row_of_merge(row, nw_col) == row)``.

        Args:
            row: 1-based row index.
            col: 1-based column index.

        Returns:
            min_row of the containing merge range if the cell is in a merge;
            row itself if the cell is not in any merge range.
        """
        merge_range = self._cell_to_range.get((row, col))
        if merge_range is None:
            return row
        return merge_range.min_row
