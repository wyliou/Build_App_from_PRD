"""Invoice data extraction for AutoConvert.

Extracts 13 per-item fields from invoice sheets with precision detection,
placeholder handling, COD override, and stop conditions. Implements FR-011.
"""

from __future__ import annotations

import logging
import re
from decimal import Decimal

from openpyxl.worksheet.worksheet import Worksheet

from autoconvert.errors import ErrorCode, ProcessingError
from autoconvert.merge_tracker import MergeTracker
from autoconvert.models import ColumnMapping, InvoiceItem
from autoconvert.utils import (
    FOOTER_KEYWORDS,
    detect_cell_precision,
    is_placeholder,
    is_stop_keyword,
    parse_numeric,
    round_half_up,
    strip_unit_suffix,
)

logger = logging.getLogger(__name__)

# Regex for stripping INV# / NO. prefixes from invoice numbers.
_INV_PREFIX_RE: re.Pattern[str] = re.compile(
    r"^(?:INV\s*#\s*|NO\.\s*)", re.IGNORECASE
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _read_string_field(
    sheet: Worksheet, row: int, col: int, merge_tracker: MergeTracker,
) -> str | None:
    """Read a string field from the sheet with merge propagation.

    Args:
        sheet: The worksheet to read from.
        row: 1-based row index.
        col: 1-based column index.
        merge_tracker: MergeTracker for merge-aware reading.

    Returns:
        Stripped string value, or None if empty.
    """
    if merge_tracker.is_in_merge(row, col) and not merge_tracker.is_merge_anchor(row, col):
        raw = merge_tracker.get_anchor_value(sheet, row, col)
    else:
        raw = sheet.cell(row=row, column=col).value

    if raw is None:
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        # Reason: Numeric part numbers (e.g., 91600.0) need string conversion.
        raw_str = str(int(raw)) if isinstance(raw, float) and raw == int(raw) else str(raw)
        return raw_str.strip() or None
    if not isinstance(raw, str):
        return str(raw).strip() or None
    return raw.strip() or None


def _require_string(
    sheet: Worksheet, row: int, col: int,
    field_name: str, merge_tracker: MergeTracker,
) -> str:
    """Read a required string field; raise ERR_030 if empty or placeholder.

    Args:
        sheet: The worksheet to read from.
        row: 1-based row index.
        col: 1-based column index.
        field_name: Field name for error context.
        merge_tracker: MergeTracker for merge-aware reading.

    Returns:
        Non-empty, non-placeholder string value.

    Raises:
        ProcessingError: ERR_030 if the field is empty or a placeholder.
    """
    val = _read_string_field(sheet, row, col, merge_tracker)
    if val is None:
        raise ProcessingError(
            code=ErrorCode.ERR_030,
            message=f"Empty required field '{field_name}' at row {row}",
            row=row, field=field_name,
        )
    if is_placeholder(val):
        raise ProcessingError(
            code=ErrorCode.ERR_030,
            message=f"Placeholder value '{val}' in required field '{field_name}' at row {row}",
            row=row, field=field_name,
        )
    return val


def _read_brand_type(
    sheet: Worksheet, row: int, brand_type_col: int,
    merge_tracker: MergeTracker, brand_value: str,
) -> str:
    """Read brand_type with horizontal merge handling.

    When brand and brand_type are horizontally merged, the non-anchor
    column is None after unmerging. Both get the anchor value.

    Args:
        sheet: The worksheet to read from.
        row: 1-based row index.
        brand_type_col: 1-based column index for brand_type.
        merge_tracker: MergeTracker instance.
        brand_value: Already-resolved brand value for fallback.

    Returns:
        The brand_type string value.

    Raises:
        ProcessingError: ERR_030 if brand_type is empty after resolution.
    """
    if merge_tracker.is_in_merge(row, brand_type_col):
        if not merge_tracker.is_merge_anchor(row, brand_type_col):
            # Reason: Non-anchor of horizontal merge — propagate anchor value.
            anchor_val = merge_tracker.get_anchor_value(sheet, row, brand_type_col)
            if anchor_val is not None and isinstance(anchor_val, str) and anchor_val.strip():
                return anchor_val.strip()
            return brand_value

    raw = sheet.cell(row=row, column=brand_type_col).value
    if raw is None or (isinstance(raw, str) and raw.strip() == ""):
        raise ProcessingError(
            code=ErrorCode.ERR_030,
            message=f"Empty required field 'brand_type' at row {row}",
            row=row, field="brand_type",
        )
    if isinstance(raw, str):
        val = raw.strip()
        if is_placeholder(val):
            raise ProcessingError(
                code=ErrorCode.ERR_030,
                message=f"Placeholder value '{val}' in required field 'brand_type' at row {row}",
                row=row, field="brand_type",
            )
        return val
    return str(raw).strip()


def _read_numeric_field(
    sheet: Worksheet, row: int, col: int,
    field_name: str, precision: int | None, merge_tracker: MergeTracker,
) -> Decimal:
    """Read and round a numeric field from the sheet.

    Args:
        sheet: The worksheet to read from.
        row: 1-based row index.
        col: 1-based column index.
        field_name: Field name for error context.
        precision: Fixed decimal places, or None to detect from cell format.
        merge_tracker: MergeTracker instance.

    Returns:
        Decimal value rounded to the appropriate precision.

    Raises:
        ProcessingError: ERR_030 for empty fields, ERR_031 for invalid values.
    """
    cell = sheet.cell(row=row, column=col)
    raw = cell.value

    # Non-anchor merged numeric cell: empty after unmerge.
    if merge_tracker.is_in_merge(row, col) and not merge_tracker.is_merge_anchor(row, col):
        raise ProcessingError(
            code=ErrorCode.ERR_030,
            message=f"Empty required field '{field_name}' at row {row} (non-anchor of merged cell)",
            row=row, field=field_name,
        )

    if raw is None or (isinstance(raw, str) and raw.strip() == ""):
        raise ProcessingError(
            code=ErrorCode.ERR_030,
            message=f"Empty required field '{field_name}' at row {row}",
            row=row, field=field_name,
        )

    if isinstance(raw, str):
        raw = strip_unit_suffix(raw)

    value = parse_numeric(raw, field_name, row)

    if precision is None:
        detected = detect_cell_precision(cell.value, cell.number_format)
        return round_half_up(value, detected)
    return round_half_up(value, precision)


def _clean_inv_no(value: str) -> str:
    """Remove INV# and NO. prefixes from an invoice number.

    Args:
        value: The raw invoice number string.

    Returns:
        Cleaned invoice number without prefixes.
    """
    return _INV_PREFIX_RE.sub("", value).strip()


def _scan_stop_keywords(sheet: Worksheet, row: int) -> bool:
    """Scan columns A-J (1-10) for stop keywords.

    Args:
        sheet: The worksheet to scan.
        row: 1-based row index.

    Returns:
        True if any string cell in columns A-J contains a stop keyword.
    """
    for col in range(1, 11):
        raw = sheet.cell(row=row, column=col).value
        if raw is not None and isinstance(raw, str) and is_stop_keyword(raw):
            return True
    return False


def _check_part_no_stops(part_no: str | None) -> bool:
    """Check if part_no triggers stop conditions 2 or 3.

    Args:
        part_no: The stripped part_no string, or None.

    Returns:
        True if a stop condition is triggered.
    """
    if part_no is None:
        return False
    lowered = part_no.lower()
    if "total" in lowered:
        return True
    return any(kw in part_no for kw in FOOTER_KEYWORDS)


def _read_optional_string(
    sheet: Worksheet, row: int, col: int | None, merge_tracker: MergeTracker,
) -> str | None:
    """Read an optional string field; return None if empty/placeholder.

    Args:
        sheet: The worksheet to read from.
        row: 1-based row index.
        col: 1-based column index, or None if column absent.
        merge_tracker: MergeTracker instance.

    Returns:
        String value, or None if column absent, empty, or placeholder.
    """
    if col is None:
        return None
    val = _read_string_field(sheet, row, col, merge_tracker)
    if val is not None and not is_placeholder(val):
        return val
    return None


# ---------------------------------------------------------------------------
# Main extraction function
# ---------------------------------------------------------------------------


def extract_invoice_items(
    sheet: Worksheet,
    column_map: ColumnMapping,
    merge_tracker: MergeTracker,
    inv_no: str | None,
) -> list[InvoiceItem]:
    """Extract all invoice line items from the sheet.

    Starts at column_map.effective_header_row + 1. Applies stop conditions
    (blank+zero-qty, stop keywords in A-J, footer keywords). Handles merged
    cells via MergeTracker for string fields. Applies COD override on COO.
    Detects and skips placeholder values. Strips unit suffixes from numeric
    fields. Applies ROUND_HALF_UP precision rules (qty=cell-precision,
    price=5dp, amount=2dp). Raises ProcessingError(ERR_030) for empty
    required fields and ProcessingError(ERR_031) for invalid numeric values.

    Args:
        sheet: The invoice worksheet (already unmerged by MergeTracker).
        column_map: Mapping of field names to 1-based column indices.
        merge_tracker: MergeTracker for merge-aware cell reading.
        inv_no: Invoice number resolved by the batch orchestrator, or None.

    Returns:
        List of InvoiceItem objects, one per extracted data row.

    Raises:
        ProcessingError: ERR_030 for empty required fields.
        ProcessingError: ERR_031 for invalid numeric values.
    """
    fm = column_map.field_map
    start_row = column_map.effective_header_row + 1
    max_row = sheet.max_row or start_row
    items: list[InvoiceItem] = []
    first_data_extracted = False

    # Required column indices.
    c_part = fm["part_no"]
    c_po = fm["po_no"]
    c_qty = fm["qty"]
    c_price = fm["price"]
    c_amount = fm["amount"]
    c_curr = fm["currency"]
    c_coo = fm["coo"]
    c_brand = fm["brand"]
    c_btype = fm["brand_type"]
    c_model = fm["model"]
    # Optional column indices.
    c_cod = fm.get("cod")
    c_inv = fm.get("inv_no")
    c_serial = fm.get("serial")

    for row in range(start_row, max_row + 1):
        # Step a: Read raw part_no and qty for stop/blank checks.
        part_no_raw = _read_string_field(sheet, row, c_part, merge_tracker)
        qty_raw = sheet.cell(row=row, column=c_qty).value

        # Step b: ALWAYS scan A-J for stop keywords first.
        if _scan_stop_keywords(sheet, row):
            logger.debug("Stop keyword in columns A-J at row %d", row)
            break

        # Step c: Check part_no stop conditions (2, 3).
        if _check_part_no_stops(part_no_raw):
            logger.debug("Stop condition (part_no) at row %d: '%s'", row, part_no_raw)
            break

        # Step d: Blank row handling.
        pn_empty = part_no_raw is None or part_no_raw == ""
        qty_empty = qty_raw is None or (isinstance(qty_raw, str) and qty_raw.strip() == "")
        if pn_empty and qty_empty:
            if first_data_extracted:
                logger.debug("Blank row at %d after data; stopping", row)
                break
            continue  # Leading blank row: skip silently.

        # Step e: Header continuation filter.
        if part_no_raw is not None and "part no" in part_no_raw.lower():
            logger.debug("Skipping header continuation row %d", row)
            continue

        # Step f: Read all field values.
        part_no = _require_string(sheet, row, c_part, "part_no", merge_tracker)
        po_no = _require_string(sheet, row, c_po, "po_no", merge_tracker)
        currency = _require_string(sheet, row, c_curr, "currency", merge_tracker)
        # Reason: COO is read as optional first because COD can override it
        # (FR-011). If COO is empty but COD has a meaningful value, COD is
        # used. ERR_030 fires only if BOTH are empty/placeholder.
        coo_raw = _read_optional_string(sheet, row, c_coo, merge_tracker)
        cod_value = _read_optional_string(sheet, row, c_cod, merge_tracker)
        if cod_value is not None and not is_placeholder(cod_value):
            coo = cod_value
        elif coo_raw is not None and not is_placeholder(coo_raw):
            coo = coo_raw
        else:
            raise ProcessingError(
                code=ErrorCode.ERR_030,
                message=f"Empty required field 'coo' at row {row}",
                row=row, field="coo",
            )

        brand = _require_string(sheet, row, c_brand, "brand", merge_tracker)
        brand_type = _read_brand_type(sheet, row, c_btype, merge_tracker, brand)
        model_val = _require_string(sheet, row, c_model, "model", merge_tracker)

        # Optional: serial.
        serial_value = _read_optional_string(sheet, row, c_serial, merge_tracker)

        # inv_no handling.
        row_inv_no: str | None = inv_no
        if c_inv is not None:
            inv_raw = _read_optional_string(sheet, row, c_inv, merge_tracker)
            row_inv_no = _clean_inv_no(inv_raw) if inv_raw is not None else inv_no
        if row_inv_no is not None:
            row_inv_no = _clean_inv_no(row_inv_no)

        # Numeric fields.
        qty = _read_numeric_field(sheet, row, c_qty, "qty", None, merge_tracker)
        price = _read_numeric_field(sheet, row, c_price, "price", 5, merge_tracker)
        amount = _read_numeric_field(sheet, row, c_amount, "amount", 2, merge_tracker)

        # Step g: Build InvoiceItem.
        items.append(InvoiceItem(
            part_no=part_no, po_no=po_no, qty=qty, price=price,
            amount=amount, currency=currency, coo=coo, cod=cod_value,
            brand=brand, brand_type=brand_type, model=model_val,
            inv_no=row_inv_no, serial=serial_value, allocated_weight=None,
        ))
        first_data_extracted = True

    # Verbatim log output per spec.
    if items:
        logger.info(
            "Invoice sheet extracted %d items (rows %d-%d)",
            len(items), start_row, start_row + len(items) - 1,
        )
    else:
        logger.info("Invoice sheet extracted 0 items")

    return items
