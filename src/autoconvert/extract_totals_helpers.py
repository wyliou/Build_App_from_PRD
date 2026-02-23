"""Total row detection and totals extraction helpers for AutoConvert.

Split from extract_packing.py to stay under the 500-line file limit.
Implements FR-014, FR-015, FR-016, FR-017.
"""

from __future__ import annotations

import logging
import re
from decimal import Decimal, InvalidOperation

from openpyxl.worksheet.worksheet import Worksheet

from autoconvert.errors import ErrorCode, ProcessingError, WarningCode
from autoconvert.merge_tracker import MergeTracker
from autoconvert.models import ColumnMapping, PackingTotals
from autoconvert.utils import (
    detect_cell_precision,
    is_cell_empty,
    is_stop_keyword,
    round_half_up,
    strip_unit_suffix,
)

logger = logging.getLogger(__name__)

# FR-017 compiled patterns
_BREAKDOWN_RE = re.compile(r"^(\d+)\s*[（(]")
_UNIT_SUFFIX_RE = re.compile(r"^(\d+)\s*(?:托|箱|件|CTNS)\b", re.IGNORECASE)
_EMBEDDED_CHINESE_RE = re.compile(r"共\s*(\d+)\s*(?:托|箱|件)")
_PALLET_RANGE_RE = re.compile(r"PLT\s*#\s*(\d+)", re.IGNORECASE)
_PLT_INDICATOR_RE = re.compile(r"PLT(?:\.G)?", re.IGNORECASE)
_JIANSHU_RE = re.compile(r"件[数數]")
_JIANSHU_EMBEDDED_RE = re.compile(r"件[数數]\s*[:：]\s*(\d+)")
_PALLET_PRIORITY_RE = re.compile(r"(\d+)\s*托")


def _parse_numeric_safe(value: object, field: str, row: int) -> Decimal:
    """Parse a cell value to Decimal, stripping units if string.

    Args:
        value: Raw cell value.
        field: Field name for error context.
        row: Row number for error context.

    Returns:
        Decimal value.

    Raises:
        ProcessingError: ERR_031 on parse failure.
    """
    if isinstance(value, str):
        cleaned = strip_unit_suffix(value.strip())
        try:
            return Decimal(cleaned)
        except (InvalidOperation, ValueError) as exc:
            raise ProcessingError(
                code=ErrorCode.ERR_031,
                message=f"Invalid numeric '{value}' at row {row}, field '{field}'",
                row=row, field=field,
            ) from exc
    if isinstance(value, bool):
        raise ProcessingError(
            code=ErrorCode.ERR_031,
            message=f"Invalid numeric '{value}' at row {row}, field '{field}'",
            row=row, field=field,
        )
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    raise ProcessingError(
        code=ErrorCode.ERR_031,
        message=f"Invalid numeric '{value}' at row {row}, field '{field}'",
        row=row, field=field,
    )


def _round_with_precision(
    value: Decimal, raw: object, number_format: str,
) -> tuple[Decimal, int]:
    """Round a value using cell format precision detection.

    For General/empty format: round to 5 dp, normalize trailing zeros.

    Args:
        value: Parsed Decimal value.
        raw: Original cell value (for detect_cell_precision).
        number_format: Cell number_format string.

    Returns:
        Tuple of (rounded Decimal, detected precision).
    """
    precision = detect_cell_precision(raw, number_format)
    fmt = (number_format or "").strip()
    if fmt == "" or fmt.lower() == "general":
        normalized = round_half_up(value, 5).normalize()
        str_val = str(normalized)
        visible = len(str_val.split(".")[1]) if "." in str_val else 0
        return round_half_up(value, visible), visible
    return round_half_up(value, precision), precision


# ---------------------------------------------------------------------------
# FR-014 — Detect Total Row
# ---------------------------------------------------------------------------


def detect_total_row(
    sheet: Worksheet,
    last_data_row: int,
    column_map: ColumnMapping,
    merge_tracker: MergeTracker,
) -> int:
    """Detect the total row using two strategies: keyword then implicit.

    Args:
        sheet: The packing worksheet.
        last_data_row: Row number of the last extracted PackingItem.
        column_map: Column mapping with field_map.
        merge_tracker: MergeTracker for merge-aware checks.

    Returns:
        1-based row number of the detected total row.

    Raises:
        ProcessingError: ERR_032 when no total row is found.
    """
    fm = column_map.field_map
    part_col, nw_col, gw_col = fm["part_no"], fm["nw"], fm["gw"]
    start = last_data_row + 1
    end = min(last_data_row + 15, sheet.max_row or last_data_row + 15)  # type: ignore[operator]

    # Strategy 1: keyword search in columns A-J
    for row in range(start, end + 1):
        for col in range(1, 11):
            val = sheet.cell(row=row, column=col).value
            if isinstance(val, str) and is_stop_keyword(val):
                return row

    # Strategy 2: implicit total row
    for row in range(start, end + 1):
        part_raw = sheet.cell(row=row, column=part_col).value
        if not is_cell_empty(part_raw):
            continue
        # Exclude merge continuations for part_no
        if merge_tracker.is_in_merge(row, part_col) and not merge_tracker.is_merge_anchor(row, part_col):
            continue
        nw_raw = sheet.cell(row=row, column=nw_col).value
        gw_raw = sheet.cell(row=row, column=gw_col).value
        if is_cell_empty(nw_raw) or is_cell_empty(gw_raw):
            continue
        try:
            nw_v = _parse_numeric_safe(nw_raw, "nw", row)
            gw_v = _parse_numeric_safe(gw_raw, "gw", row)
        except (ProcessingError, InvalidOperation, ValueError, TypeError):
            continue
        if nw_v > 0 and gw_v > 0:
            return row

    raise ProcessingError(
        code=ErrorCode.ERR_032,
        message=f"Total row not found in rows {start}-{end}",
        row=last_data_row,
    )


# ---------------------------------------------------------------------------
# FR-015 — Extract total_nw
# ---------------------------------------------------------------------------


def _extract_total_nw(
    sheet: Worksheet, total_row: int, nw_col: int,
) -> tuple[Decimal, int]:
    """Extract total_nw from the total row with precision detection.

    Args:
        sheet: The packing worksheet.
        total_row: 1-based row of the total row.
        nw_col: 1-based column index for NW.

    Returns:
        Tuple of (total_nw, precision).

    Raises:
        ProcessingError: ERR_033 for non-numeric or missing NW value.
    """
    raw = sheet.cell(row=total_row, column=nw_col).value
    if is_cell_empty(raw):
        raise ProcessingError(
            code=ErrorCode.ERR_033,
            message=f"Missing total NW value at row {total_row}",
            row=total_row, field="nw",
        )
    try:
        nw_val = _parse_numeric_safe(raw, "nw", total_row)
    except ProcessingError:
        raise ProcessingError(
            code=ErrorCode.ERR_033,
            message=f"Invalid total NW value '{raw}' at row {total_row}",
            row=total_row, field="nw",
        )
    fmt = sheet.cell(row=total_row, column=nw_col).number_format
    return _round_with_precision(nw_val, raw, fmt)


# ---------------------------------------------------------------------------
# FR-016 — Extract total_gw
# ---------------------------------------------------------------------------


def _extract_total_gw(
    sheet: Worksheet, total_row: int, gw_col: int,
) -> tuple[Decimal, int]:
    """Extract total_gw with packaging weight augmentation (+1/+2 rows).

    Args:
        sheet: The packing worksheet.
        total_row: 1-based row of the total row.
        gw_col: 1-based column index for GW.

    Returns:
        Tuple of (total_gw, precision).

    Raises:
        ProcessingError: ERR_034 for non-numeric or missing GW value.
    """
    raw = sheet.cell(row=total_row, column=gw_col).value
    if is_cell_empty(raw):
        raise ProcessingError(
            code=ErrorCode.ERR_034,
            message=f"Missing total GW value at row {total_row}",
            row=total_row, field="gw",
        )
    try:
        gw_val = _parse_numeric_safe(raw, "gw", total_row)
    except ProcessingError:
        raise ProcessingError(
            code=ErrorCode.ERR_034,
            message=f"Invalid total GW value '{raw}' at row {total_row}",
            row=total_row, field="gw",
        )
    fmt = sheet.cell(row=total_row, column=gw_col).number_format
    primary_gw, precision = _round_with_precision(gw_val, raw, fmt)

    # Packaging weight augmentation: check +1 and +2 rows
    max_row = sheet.max_row or total_row  # type: ignore[assignment]
    plus1_ok = _is_row_numeric(sheet, total_row + 1, gw_col, max_row)  # type: ignore[arg-type]
    plus2_val = _try_parse_row(sheet, total_row + 2, gw_col, max_row)  # type: ignore[arg-type]

    if plus1_ok and plus2_val is not None:
        fmt2 = sheet.cell(row=total_row + 2, column=gw_col).number_format
        raw2 = sheet.cell(row=total_row + 2, column=gw_col).value
        return _round_with_precision(plus2_val, raw2, fmt2)

    return primary_gw, precision


def _is_row_numeric(
    sheet: Worksheet, row: int, col: int, max_row: int,
) -> bool:
    """Check if a cell contains a parseable numeric value."""
    if row > max_row or row < 1:
        return False
    raw = sheet.cell(row=row, column=col).value
    if is_cell_empty(raw):
        return False
    try:
        _parse_numeric_safe(raw, "", row)
        return True
    except (ProcessingError, InvalidOperation, ValueError, TypeError):
        return False


def _try_parse_row(
    sheet: Worksheet, row: int, col: int, max_row: int,
) -> Decimal | None:
    """Try to parse a cell as Decimal, return None on failure."""
    if row > max_row or row < 1:
        return None
    raw = sheet.cell(row=row, column=col).value
    if is_cell_empty(raw):
        return None
    try:
        return _parse_numeric_safe(raw, "", row)
    except (ProcessingError, InvalidOperation, ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# FR-017 — Extract total_packets
# ---------------------------------------------------------------------------


def _extract_total_packets(
    sheet: Worksheet, total_row: int, column_map: ColumnMapping,
) -> tuple[int | None, ProcessingError | None]:
    """Extract total_packets using 3-priority search.

    Args:
        sheet: The packing worksheet.
        total_row: 1-based total row.
        column_map: Column mapping with field_map.

    Returns:
        Tuple of (packets or None, ATT_002 warning or None).
    """
    nw_col = column_map.field_map["nw"]
    max_col = max(11, nw_col + 3)
    max_row: int = sheet.max_row or total_row  # type: ignore[assignment]

    for search_fn in (
        lambda: _search_jianshu(sheet, total_row, max_col, max_row),
        lambda: _search_plt_indicator(sheet, total_row, max_col),
        lambda: _search_below_total_patterns(sheet, total_row, max_col, max_row),
    ):
        result = search_fn()
        if result is not None:
            return result, None

    warning = ProcessingError(
        code=WarningCode.ATT_002,
        message="Total packets not found in packing sheet",
        row=total_row, field="total_packets",
    )
    return None, warning


def _validate_packets(value: int) -> int | None:
    """Validate packets in range 1-1000; return None if out of range."""
    return value if 1 <= value <= 1000 else None


def _search_jianshu(
    sheet: Worksheet, total_row: int, max_col: int, max_row: int,
) -> int | None:
    """Priority 1: Search for 件数/件數 label in rows total_row+1 to +3."""
    for row in range(total_row + 1, min(total_row + 4, max_row + 1)):
        for col in range(1, max_col):
            cell_val = sheet.cell(row=row, column=col).value
            if not isinstance(cell_val, str) or not _JIANSHU_RE.search(cell_val):
                continue
            # Embedded value: "件数: 3"
            embedded = _JIANSHU_EMBEDDED_RE.search(cell_val)
            if embedded:
                val = _validate_packets(int(embedded.group(1)))
                if val is not None:
                    return val
            # Adjacent right cells (up to +3 cols)
            for adj_col in range(col + 1, min(col + 4, max_col)):
                adj_val = sheet.cell(row=row, column=adj_col).value
                if adj_val is None:
                    continue
                adj_str = strip_unit_suffix(str(adj_val).strip())
                if not adj_str:
                    continue
                try:
                    val = _validate_packets(int(Decimal(adj_str)))
                    if val is not None:
                        return val
                except (InvalidOperation, ValueError, TypeError):
                    continue
    return None


def _search_plt_indicator(
    sheet: Worksheet, total_row: int, max_col: int,
) -> int | None:
    """Priority 2: Search for PLT/PLT.G indicator in rows total_row-1 and -2."""
    for row in (total_row - 1, total_row - 2):
        if row < 1:
            continue
        for col in range(1, max_col):
            cell_val = sheet.cell(row=row, column=col).value
            if cell_val is None:
                continue
            # Number-before-PLT: numeric cell with PLT in adjacent
            if isinstance(cell_val, (int, float)) and not isinstance(cell_val, bool):
                for adj_col in (col + 1, col - 1):
                    if adj_col < 1 or adj_col >= max_col:
                        continue
                    adj = sheet.cell(row=row, column=adj_col).value
                    if isinstance(adj, str) and _PLT_INDICATOR_RE.search(adj.strip()):
                        val = _validate_packets(int(cell_val))
                        if val is not None:
                            return val
            # PLT-before-number: PLT text, number in adjacent right
            if isinstance(cell_val, str) and _PLT_INDICATOR_RE.search(cell_val.strip()):
                for adj_col in range(col + 1, min(col + 3, max_col)):
                    adj_val = sheet.cell(row=row, column=adj_col).value
                    if adj_val is None:
                        continue
                    if isinstance(adj_val, (int, float)) and not isinstance(adj_val, bool):
                        val = _validate_packets(int(adj_val))
                        if val is not None:
                            return val
                    if isinstance(adj_val, str):
                        try:
                            val = _validate_packets(int(Decimal(adj_val.strip())))
                            if val is not None:
                                return val
                        except (InvalidOperation, ValueError, TypeError):
                            continue
    return None


def _search_below_total_patterns(
    sheet: Worksheet, total_row: int, max_col: int, max_row: int,
) -> int | None:
    """Priority 3: Search below-total rows for pattern-based packet counts.

    Pallet count (托) wins over box count (件) when both appear.
    """
    for row in range(total_row + 1, min(total_row + 4, max_row + 1)):
        for col in range(1, max_col):
            cell_val = sheet.cell(row=row, column=col).value
            if not isinstance(cell_val, str):
                continue
            text = cell_val.strip()
            if not text:
                continue
            # Pallet priority: "共7托（172件）" → 7
            m = _PALLET_PRIORITY_RE.search(text)
            if m:
                val = _validate_packets(int(m.group(1)))
                if val is not None:
                    return val
            # (a) Total-with-breakdown: "348（256胶框+92纸箱）"
            m = _BREAKDOWN_RE.match(text)
            if m:
                val = _validate_packets(int(m.group(1)))
                if val is not None:
                    return val
            # (b) Unit-suffix: "7托", "30箱", "55 CTNS"
            m = _UNIT_SUFFIX_RE.match(text)
            if m:
                val = _validate_packets(int(m.group(1)))
                if val is not None:
                    return val
            # (c) Embedded Chinese: "共7托"
            m = _EMBEDDED_CHINESE_RE.search(text)
            if m:
                val = _validate_packets(int(m.group(1)))
                if val is not None:
                    return val
            # (d) Pallet range: "PLT#1(1~34)"
            m = _PALLET_RANGE_RE.search(text)
            if m:
                val = _validate_packets(int(m.group(1)))
                if val is not None:
                    return val
    return None


# ---------------------------------------------------------------------------
# Combined extract_totals entry point
# ---------------------------------------------------------------------------


def extract_totals(
    sheet: Worksheet, total_row: int, column_map: ColumnMapping,
) -> tuple[PackingTotals, list[ProcessingError]]:
    """Extract total_nw, total_gw, total_packets from the packing total area.

    Args:
        sheet: The packing worksheet.
        total_row: 1-based row of the detected total row.
        column_map: Column mapping with field_map.

    Returns:
        Tuple of (PackingTotals, list of warnings such as ATT_002).

    Raises:
        ProcessingError: ERR_033 for invalid total_nw, ERR_034 for invalid total_gw.
    """
    fm = column_map.field_map
    warnings: list[ProcessingError] = []
    total_nw, nw_prec = _extract_total_nw(sheet, total_row, fm["nw"])
    total_gw, gw_prec = _extract_total_gw(sheet, total_row, fm["gw"])
    total_packets, warning = _extract_total_packets(sheet, total_row, column_map)

    if warning is not None:
        logger.warning("%s: %s", warning.code, warning.message)
        warnings.append(warning)

    logger.info(
        "Packing total row at row %d, NW= %s, GW= %s, Packets= %s",
        total_row, total_nw, total_gw, total_packets,
    )

    return PackingTotals(
        total_nw=total_nw, total_nw_precision=nw_prec,
        total_gw=total_gw, total_gw_precision=gw_prec,
        total_packets=total_packets,
    ), warnings
