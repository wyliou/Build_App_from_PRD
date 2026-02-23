"""Output template generation for AutoConvert.

Populates the 40-column customs template with extracted and transformed
data, saves output files. Implements FR-029, FR-030.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from autoconvert.errors import ErrorCode, ProcessingError
from autoconvert.models import AppConfig, InvoiceItem, PackingTotals

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column index constants (1-based, openpyxl convention)
# ---------------------------------------------------------------------------

_COL_PART_NO = 1    # A — 企业料号
_COL_PO_NO = 2      # B — 采购订单号
_COL_ZHENGMIAN = 3  # C — 征免方式 (fixed "3")
_COL_CURRENCY = 4   # D — 币制
_COL_QTY = 5        # E — 申报数量
_COL_PRICE = 6      # F — 申报单价
_COL_AMOUNT = 7     # G — 申报总价
_COL_COO = 8        # H — 原产国
# I=9, J=10, K=11  — reserved, empty
_COL_SERIAL = 12        # L — 报关单商品序号
_COL_NET_WEIGHT = 13    # M — 净重 (allocated_weight)
_COL_INV_NO = 14        # N — 发票号码
# O=15             — reserved, empty
_COL_TOTAL_GW = 16      # P — 毛重 (row 5 only)
# Q=17             — reserved, empty
_COL_DOMESTIC_DEST = 18  # R — 境内目的地代码 (fixed "32052")
_COL_ADMIN_DIST = 19     # S — 行政区划代码 (fixed "320506")
_COL_FINAL_DEST = 20     # T — 最终目的国 (fixed "142")
# U=21 … AJ=36    — reserved, empty (16 columns)
_COL_TOTAL_PACKETS = 37  # AK — 件数 (row 5 only)
_COL_BRAND = 38          # AL — 品牌
_COL_BRAND_TYPE = 39     # AM — 品牌类型
_COL_MODEL = 40          # AN — 型号

# Fixed string values written to every data row
_FIXED_ZHENGMIAN = "3"
_FIXED_DOMESTIC_DEST = "32052"
_FIXED_ADMIN_DIST = "320506"
_FIXED_FINAL_DEST = "142"

# Sheet name (must match template exactly)
_SHEET_NAME = "工作表1"

# First data row (template rows 1-4 are header rows)
_FIRST_DATA_ROW = 5


def write_template(
    invoice_items: list[InvoiceItem],
    packing_totals: PackingTotals,
    config: AppConfig,
    output_path: Path,
) -> None:
    """Load the output template, populate rows 5+ with invoice data, and save.

    Writes one row per InvoiceItem starting at row 5, preserving header rows
    1-4 unchanged. Fixed values are written to every row. The total_gw and
    total_packets fields from PackingTotals are written to row 5 only.

    Args:
        invoice_items: Validated items with allocated weights, in output order.
        packing_totals: Extracted packing totals (total_gw, total_packets).
        config: Application configuration containing output_template_path.
        output_path: Full path where the populated workbook should be saved.

    Raises:
        ProcessingError: ERR_051 if the template workbook cannot be loaded.
        ProcessingError: ERR_052 if the output file cannot be saved.
    """
    # --- Step 1: Load template ---
    try:
        workbook = openpyxl.load_workbook(config.output_template_path)
    except Exception as exc:
        raise ProcessingError(
            code=ErrorCode.ERR_051,
            message=(
                f"TEMPLATE_LOAD_FAILED: Could not load output template "
                f"'{config.output_template_path}': {exc}"
            ),
        ) from exc

    # --- Step 2: Select sheet ---
    if _SHEET_NAME not in workbook.sheetnames:
        raise ProcessingError(
            code=ErrorCode.ERR_051,
            message=(
                f"TEMPLATE_LOAD_FAILED: Sheet '{_SHEET_NAME}' not found in "
                f"template '{config.output_template_path}'. "
                f"Available sheets: {workbook.sheetnames}"
            ),
        )
    sheet: Worksheet = workbook[_SHEET_NAME]  # type: ignore[assignment]

    # --- Steps 3-6: Write data rows ---
    for row_idx, item in enumerate(invoice_items, start=_FIRST_DATA_ROW):
        _write_item_row(sheet, row_idx, item)

        # Fixed values written to every row
        sheet.cell(row=row_idx, column=_COL_ZHENGMIAN).value = _FIXED_ZHENGMIAN  # type: ignore[assignment]
        sheet.cell(row=row_idx, column=_COL_DOMESTIC_DEST).value = _FIXED_DOMESTIC_DEST  # type: ignore[assignment]
        sheet.cell(row=row_idx, column=_COL_ADMIN_DIST).value = _FIXED_ADMIN_DIST  # type: ignore[assignment]
        sheet.cell(row=row_idx, column=_COL_FINAL_DEST).value = _FIXED_FINAL_DEST  # type: ignore[assignment]

        # P and AK written to row 5 only
        if row_idx == _FIRST_DATA_ROW:
            sheet.cell(row=row_idx, column=_COL_TOTAL_GW).value = float(  # type: ignore[assignment]
                packing_totals.total_gw
            )
            if packing_totals.total_packets is not None:
                sheet.cell(row=row_idx, column=_COL_TOTAL_PACKETS).value = (  # type: ignore[assignment]
                    packing_totals.total_packets
                )

    # --- Step 7: Save workbook ---
    try:
        workbook.save(output_path)
    except (OSError, PermissionError) as exc:
        raise ProcessingError(
            code=ErrorCode.ERR_052,
            message=(
                f"OUTPUT_WRITE_FAILED: Could not save output file "
                f"'{output_path}': {exc}"
            ),
        ) from exc

    # --- Step 8: Log success ---
    logger.info("Output successfully written to: %s", output_path.name)


def _write_item_row(
    sheet: Worksheet,
    row: int,
    item: InvoiceItem,
) -> None:
    """Write a single InvoiceItem's fields to the given sheet row.

    Writes text columns as strings and numeric columns as their native types.
    Skips reserved/empty columns entirely (no None writes).

    Args:
        sheet: The target openpyxl Worksheet object.
        row: 1-based row index to write data into.
        item: The InvoiceItem whose fields are being written.
    """
    # Text columns — write as str
    sheet.cell(row=row, column=_COL_PART_NO).value = str(item.part_no)  # type: ignore[assignment]
    sheet.cell(row=row, column=_COL_PO_NO).value = str(item.po_no)  # type: ignore[assignment]
    sheet.cell(row=row, column=_COL_CURRENCY).value = str(item.currency)  # type: ignore[assignment]
    sheet.cell(row=row, column=_COL_COO).value = str(item.coo)  # type: ignore[assignment]
    sheet.cell(row=row, column=_COL_BRAND).value = str(item.brand)  # type: ignore[assignment]
    sheet.cell(row=row, column=_COL_BRAND_TYPE).value = str(item.brand_type)  # type: ignore[assignment]
    sheet.cell(row=row, column=_COL_MODEL).value = str(item.model)  # type: ignore[assignment]

    # inv_no may be None — only write if present (None write would clear formatting)
    if item.inv_no is not None:
        sheet.cell(row=row, column=_COL_INV_NO).value = str(item.inv_no)  # type: ignore[assignment]

    # Serial: text column — may be None; skip if absent
    if item.serial is not None:
        sheet.cell(row=row, column=_COL_SERIAL).value = str(item.serial)  # type: ignore[assignment]

    # Numeric columns — write as float (openpyxl accepts Decimal too, but
    # float is safer for cell formatting compatibility)
    sheet.cell(row=row, column=_COL_QTY).value = float(item.qty)  # type: ignore[assignment]
    sheet.cell(row=row, column=_COL_PRICE).value = float(item.price)  # type: ignore[assignment]
    sheet.cell(row=row, column=_COL_AMOUNT).value = float(item.amount)  # type: ignore[assignment]

    # Net weight: numeric, may be None (caller-contract bug if so)
    if item.allocated_weight is not None:
        sheet.cell(row=row, column=_COL_NET_WEIGHT).value = float(  # type: ignore[assignment]
            item.allocated_weight
        )
    # Reason: if allocated_weight is None, we leave the cell unwritten to
    # avoid overwriting template formatting with None (per spec gotchas).
