"""Internal helpers for config loading and validation.

Extracted from config.py to stay under the 500-line limit.
These functions are not part of the public API.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, cast

import openpyxl  # type: ignore[import-untyped]
import yaml

from autoconvert.errors import ConfigError, ErrorCode
from autoconvert.models import FieldPattern, InvNoCellConfig

# Valid field type strings for column definitions.
_VALID_FIELD_TYPES: frozenset[str] = frozenset({"string", "numeric", "currency"})

# Required top-level keys in field_patterns.yaml.
_REQUIRED_YAML_KEYS: frozenset[str] = frozenset(
    {
        "invoice_sheet",
        "packing_sheet",
        "invoice_columns",
        "packing_columns",
        "inv_no_cell",
    }
)


def load_yaml(yaml_path: Path) -> dict[str, Any]:
    """Load and structurally validate field_patterns.yaml.

    Args:
        yaml_path: Path to field_patterns.yaml.

    Returns:
        Parsed YAML data as a dict.

    Raises:
        ConfigError: On missing required top-level keys (ERR_004).
    """
    with open(yaml_path, encoding="utf-8") as f:
        config_data: dict[str, Any] = cast(dict[str, Any], yaml.safe_load(f))

    for key in _REQUIRED_YAML_KEYS:
        if key not in config_data:
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=f"Missing required key '{key}' in {yaml_path.name}",
                path=str(yaml_path),
            )

    return config_data


def compile_pattern(
    pattern_str: str, context_name: str, file_path: str
) -> re.Pattern[str]:
    """Compile a single regex pattern with IGNORECASE flag.

    Args:
        pattern_str: The regex pattern string to compile.
        context_name: Descriptive name for error messages (e.g. field name).
        file_path: Config file path for error messages.

    Returns:
        Compiled re.Pattern.

    Raises:
        ConfigError: On invalid regex (ERR_002).
    """
    try:
        return re.compile(pattern_str, re.IGNORECASE)
    except re.error as exc:
        raise ConfigError(
            code=ErrorCode.ERR_002,
            message=(
                f"Invalid regex pattern in '{context_name}': "
                f"'{pattern_str}' — {exc}"
            ),
            path=file_path,
        ) from exc


def compile_pattern_list(
    patterns: list[str], context_name: str, file_path: str
) -> list[re.Pattern[str]]:
    """Compile a list of regex pattern strings.

    Args:
        patterns: List of regex strings to compile.
        context_name: Descriptive name for error messages.
        file_path: Config file path for error messages.

    Returns:
        List of compiled re.Pattern objects.

    Raises:
        ConfigError: On any invalid regex (ERR_002).
    """
    return [compile_pattern(p, context_name, file_path) for p in patterns]


def build_field_patterns(
    section: dict[str, Any],
    expected_fields: frozenset[str],
    section_name: str,
    file_path: str,
) -> dict[str, FieldPattern]:
    """Build FieldPattern objects from a YAML column section.

    Validates that each expected field is present, and each field entry
    has the required keys (patterns, type, required) with correct types.

    Args:
        section: The parsed YAML dict for invoice_columns or packing_columns.
        expected_fields: Set of required field names.
        section_name: Section name for error messages.
        file_path: Config file path for error messages.

    Returns:
        Dict mapping field name to FieldPattern.

    Raises:
        ConfigError: On missing fields or malformed entries (ERR_004),
            invalid regex (ERR_002).
    """
    result: dict[str, FieldPattern] = {}

    # Check all expected fields are present.
    for field_name in expected_fields:
        if field_name not in section:
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=(
                    f"Missing field '{field_name}' in "
                    f"'{section_name}' in {Path(file_path).name}"
                ),
                path=file_path,
            )

    for field_name in expected_fields:
        entry = section[field_name]
        _validate_field_entry(entry, field_name, section_name, file_path)

        # Compile patterns.
        compiled = compile_pattern_list(
            entry["patterns"],
            f"{section_name}.{field_name}",
            file_path,
        )

        result[field_name] = FieldPattern(
            patterns=compiled,
            field_type=entry["type"],
            required=entry["required"],
        )

    return result


def _validate_field_entry(
    entry: Any,
    field_name: str,
    section_name: str,
    file_path: str,
) -> None:
    """Validate a single field entry dict has correct structure.

    Args:
        entry: The YAML dict for one field (e.g. part_no).
        field_name: Name of the field for error messages.
        section_name: Parent section name for error messages.
        file_path: Config file path for error messages.

    Raises:
        ConfigError: On structural issues (ERR_004).
    """
    if not isinstance(entry, dict):
        raise ConfigError(
            code=ErrorCode.ERR_004,
            message=(
                f"Field '{field_name}' in '{section_name}' must be a "
                f"mapping, got {type(entry).__name__}"
            ),
            path=file_path,
        )

    for required_key in ("patterns", "type", "required"):
        if required_key not in entry:
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=(
                    f"Missing key '{required_key}' in "
                    f"'{section_name}.{field_name}' in "
                    f"{Path(file_path).name}"
                ),
                path=file_path,
            )

    if not isinstance(entry["patterns"], list):
        raise ConfigError(
            code=ErrorCode.ERR_004,
            message=(
                f"'{section_name}.{field_name}.patterns' must be a list"
            ),
            path=file_path,
        )

    field_type = entry["type"]
    if field_type not in _VALID_FIELD_TYPES:
        raise ConfigError(
            code=ErrorCode.ERR_004,
            message=(
                f"Invalid type '{field_type}' for "
                f"'{section_name}.{field_name}', "
                f"expected one of {sorted(_VALID_FIELD_TYPES)}"
            ),
            path=file_path,
        )

    if not isinstance(entry["required"], bool):
        raise ConfigError(
            code=ErrorCode.ERR_004,
            message=(
                f"'{section_name}.{field_name}.required' must be bool"
            ),
            path=file_path,
        )


def build_inv_no_cell_config(
    section: dict[str, Any], file_path: str
) -> InvNoCellConfig:
    """Build InvNoCellConfig from the inv_no_cell YAML section.

    Args:
        section: The parsed YAML dict for inv_no_cell.
        file_path: Config file path for error messages.

    Returns:
        InvNoCellConfig with all patterns compiled.

    Raises:
        ConfigError: On missing keys (ERR_004) or invalid regex (ERR_002).
    """
    for key in ("patterns", "label_patterns", "exclude_patterns"):
        if key not in section:
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=(
                    f"Missing key '{key}' in 'inv_no_cell' "
                    f"in {Path(file_path).name}"
                ),
                path=file_path,
            )

    return InvNoCellConfig(
        patterns=compile_pattern_list(
            section["patterns"], "inv_no_cell.patterns", file_path
        ),
        label_patterns=compile_pattern_list(
            section["label_patterns"],
            "inv_no_cell.label_patterns",
            file_path,
        ),
        exclude_patterns=compile_pattern_list(
            section["exclude_patterns"],
            "inv_no_cell.exclude_patterns",
            file_path,
        ),
    )


def normalize_lookup_key(value: str) -> str:
    """Normalize a lookup table key to UPPERCASE with collapsed comma spacing.

    Shared normalization contract: uppercase the key, then collapse
    ``", "`` (comma-space) to ``","`` (comma only).

    Args:
        value: Raw source value string.

    Returns:
        Normalized key string.
    """
    # Reason: The PRD requires case-insensitive lookup and comma normalization.
    # "Taiwan, China" -> "TAIWAN,CHINA"
    return value.strip().upper().replace(", ", ",")


def load_lookup_table(
    xlsx_path: Path, sheet_name: str, display_filename: str
) -> dict[str, str]:
    """Load a two-column lookup table from an Excel file.

    Reads the header row to locate Source_Value and Target_Code columns,
    then reads all data rows. Keys are normalized to UPPERCASE with
    collapsed comma spacing. All Target_Code values are cast to str.
    Duplicate normalized keys raise ConfigError(ERR_003).

    Args:
        xlsx_path: Path to the Excel file.
        sheet_name: Sheet name to read (case-sensitive).
        display_filename: Filename for error messages.

    Returns:
        Dict mapping normalized Source_Value to str Target_Code.

    Raises:
        ConfigError: On duplicate keys (ERR_003) or structural issues (ERR_005).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        ws = wb[sheet_name]
    except KeyError:
        wb.close()
        raise ConfigError(
            code=ErrorCode.ERR_005,
            message=f"Sheet '{sheet_name}' not found in {display_filename}",
            path=str(xlsx_path),
        )

    source_col, target_col = _find_lookup_columns(ws, display_filename, xlsx_path)
    lookup = _read_lookup_rows(ws, source_col, target_col, display_filename, xlsx_path)

    wb.close()
    return lookup


def _find_lookup_columns(
    ws: Any, display_filename: str, xlsx_path: Path
) -> tuple[int, int]:
    """Locate Source_Value and Target_Code column indices in the header row.

    Args:
        ws: openpyxl worksheet.
        display_filename: Filename for error messages.
        xlsx_path: Path for error messages.

    Returns:
        Tuple of (source_col, target_col) as 1-based indices.

    Raises:
        ConfigError: If either column header is missing (ERR_005).
    """
    source_col: int | None = None
    target_col: int | None = None

    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col_idx).value
        if isinstance(header_val, str):
            header_val = header_val.strip()
            if header_val == "Source_Value":
                source_col = col_idx
            elif header_val == "Target_Code":
                target_col = col_idx

    if source_col is None or target_col is None:
        raise ConfigError(
            code=ErrorCode.ERR_005,
            message=(
                f"Missing 'Source_Value' or 'Target_Code' column header "
                f"in {display_filename}"
            ),
            path=str(xlsx_path),
        )

    return source_col, target_col


def _read_lookup_rows(
    ws: Any,
    source_col: int,
    target_col: int,
    display_filename: str,
    xlsx_path: Path,
) -> dict[str, str]:
    """Read data rows from a lookup sheet and build the normalized dict.

    Args:
        ws: openpyxl worksheet.
        source_col: 1-based column index for Source_Value.
        target_col: 1-based column index for Target_Code.
        display_filename: Filename for error messages.
        xlsx_path: Path for error messages.

    Returns:
        Dict mapping normalized keys to string target codes.

    Raises:
        ConfigError: On duplicate normalized keys (ERR_003).
    """
    lookup: dict[str, str] = {}

    for row_idx in range(2, ws.max_row + 1):
        source_raw = ws.cell(row=row_idx, column=source_col).value
        target_raw = ws.cell(row=row_idx, column=target_col).value

        # Skip rows where source is empty.
        if source_raw is None or (
            isinstance(source_raw, str) and source_raw.strip() == ""
        ):
            continue

        key = normalize_lookup_key(str(source_raw))

        # Normalize Target_Code to str.
        # Reason: openpyxl may read integer cell values as int; the PRD
        # requires all Target_Code values to be stored as str.
        if isinstance(target_raw, (int, float)):
            target_value = str(int(target_raw))
        else:
            target_value = str(target_raw).strip() if target_raw else ""

        if key in lookup:
            raise ConfigError(
                code=ErrorCode.ERR_003,
                message=(
                    f"Duplicate Source_Value '{key}' in {display_filename}"
                ),
                path=str(xlsx_path),
            )

        lookup[key] = target_value

    return lookup


def validate_template(template_path: Path) -> None:
    """Validate the output template Excel file structure.

    Checks that the sheet named ``工作表1`` exists, has at least 40 columns
    (A through AN), and has at least 4 rows (header metadata rows).

    Args:
        template_path: Path to output_template.xlsx.

    Raises:
        ConfigError: On missing sheet, too few columns, or too few rows
            (ERR_005).
    """
    wb = openpyxl.load_workbook(template_path, data_only=True)

    sheet_name = "\u5de5\u4f5c\u88681"  # 工作表1
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ConfigError(
            code=ErrorCode.ERR_005,
            message=(
                f"Sheet '{sheet_name}' not found in "
                f"{template_path.name}"
            ),
            path=str(template_path),
        )

    ws = wb[sheet_name]

    if ws.max_column < 40:
        wb.close()
        raise ConfigError(
            code=ErrorCode.ERR_005,
            message=(
                f"Template sheet '{sheet_name}' has {ws.max_column} columns, "
                f"expected at least 40 (A-AN)"
            ),
            path=str(template_path),
        )

    if ws.max_row < 4:
        wb.close()
        raise ConfigError(
            code=ErrorCode.ERR_005,
            message=(
                f"Template sheet '{sheet_name}' has {ws.max_row} rows, "
                f"expected at least 4 header metadata rows"
            ),
            path=str(template_path),
        )

    wb.close()
