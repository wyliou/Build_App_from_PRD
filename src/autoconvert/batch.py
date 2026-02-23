"""Batch processing orchestrator for AutoConvert.

Manages file discovery, directory setup, per-file processing pipeline,
and diagnostic mode. Implements FR-001, FR-003, FR-028, FR-034.

Error codes owned by this module:
    ERR_010 (FILE_LOCKED) — PermissionError when opening a file.
    ERR_011 (FILE_CORRUPT) — Parse error or unexpected exception.
    ERR_021 (INVOICE_NUMBER_NOT_FOUND) — after all fallbacks fail.
"""

from __future__ import annotations

import logging
import os
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from autoconvert.column_map import (
    detect_header_row,
    extract_inv_no_from_header,
    map_columns,
)
from autoconvert.errors import ErrorCode, ProcessingError
from autoconvert.extract_invoice import extract_invoice_items
from autoconvert.extract_packing import (
    detect_total_row,
    extract_packing_items,
    extract_totals,
    validate_merged_weights,
)
from autoconvert.merge_tracker import MergeTracker
from autoconvert.models import (
    AppConfig,
    BatchResult,
    FileResult,
    InvoiceItem,
    PackingItem,
    PackingTotals,
)
from autoconvert.output import write_template
from autoconvert.sheet_detect import detect_sheets as _detect_sheets
from autoconvert.transform import clean_po_number, convert_country, convert_currency
from autoconvert.validate import determine_file_status
from autoconvert.weight_alloc import allocate_weights

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def run_batch(
    config: AppConfig,
    data_dir: Path,
    diagnostic_file: str | None = None,
) -> BatchResult:
    """Orchestrate full batch processing of vendor Excel files.

    Creates required directories, optionally clears the output directory,
    scans for input files, processes each file independently via
    ``process_file()``, and returns a BatchResult with all outcomes.
    In diagnostic mode, processes only the specified file without clearing
    the output directory.

    Args:
        config: Validated application configuration from config.py.
        data_dir: Path to the data/ directory containing input files.
        diagnostic_file: Filename or path for single-file diagnostic mode.
            None for normal batch mode.

    Returns:
        BatchResult with per-file results, counts, and processing time.
    """
    finished_dir = data_dir / "finished"

    # FR-001: Create data/ and data/finished/ if they do not exist.
    _create_directories(data_dir, finished_dir)

    # FR-028: Clear data/finished/ before processing (skip in diagnostic mode).
    if diagnostic_file is None:
        _clear_finished_directory(finished_dir)

    # FR-003: Scan for processable files.
    if diagnostic_file is not None:
        files = _resolve_diagnostic_file(diagnostic_file, data_dir)
    else:
        files = _scan_input_files(data_dir)

    if not files:
        logger.info("No processable files found in %s", data_dir)
        return BatchResult(
            total_files=0,
            success_count=0,
            attention_count=0,
            failed_count=0,
            processing_time=0.0,
            file_results=[],
            log_path=str(data_dir.parent / "process_log.txt"),
        )

    # Process each file independently.
    start_time = time.time()
    file_results: list[FileResult] = []

    for idx, filepath in enumerate(files, start=1):
        logger.info("-" * 65)
        logger.info("[%d/%d] Processing: %s ...", idx, len(files), filepath.name)
        result = process_file(filepath, config)
        file_results.append(result)

        # Log per-file status.
        if result.status == "Success":
            logger.info("\u2705 SUCCESS")
        elif result.status == "Attention":
            for w in result.warnings:
                logger.warning("[%s] %s", w.code, w.message)
            logger.warning("\u26a0\ufe0f ATTENTION")
        else:
            for e in result.errors:
                logger.error("[%s] %s", e.code, e.message)
            logger.error("\u274c FAILED")

    processing_time = time.time() - start_time

    success_count = sum(1 for r in file_results if r.status == "Success")
    attention_count = sum(1 for r in file_results if r.status == "Attention")
    failed_count = sum(1 for r in file_results if r.status == "Failed")

    return BatchResult(
        total_files=len(file_results),
        success_count=success_count,
        attention_count=attention_count,
        failed_count=failed_count,
        processing_time=processing_time,
        file_results=file_results,
        log_path=str(data_dir.parent / "process_log.txt"),
    )


def process_file(filepath: Path, config: AppConfig) -> FileResult:
    """Process a single vendor Excel file through the full pipeline.

    Pipeline phases: sheet detection, column mapping, data extraction,
    transformation, weight allocation, validation, and output. Returns
    a FileResult with status, errors, warnings, and extracted data.

    Args:
        filepath: Absolute path to the input Excel file.
        config: Validated application configuration.

    Returns:
        FileResult with status, errors, warnings, and extracted data.
    """
    errors: list[ProcessingError] = []
    warnings: list[ProcessingError] = []
    invoice_items: list[InvoiceItem] | None = None
    packing_items: list[PackingItem] | None = None
    packing_totals: PackingTotals | None = None
    status = "Failed"

    try:
        # Phase 1: Open workbook.
        workbook = _open_workbook(filepath)

        # Phase 2: Sheet detection.
        try:
            sheet_pair = _detect_sheets(workbook, config)
        except ProcessingError as e:
            e.filename = filepath.name
            errors.append(e)
            status = determine_file_status(errors, warnings)
            return _build_result(
                filepath, status, errors, warnings, None, None, None,
            )

        inv_sheet = sheet_pair.invoice_sheet
        pack_sheet = sheet_pair.packing_sheet

        # Phase 3: Column mapping.
        phase3_errors: list[ProcessingError] = []

        # 3a/3b: MergeTracker (must happen BEFORE header detection).
        inv_merge = MergeTracker(inv_sheet)
        pack_merge = MergeTracker(pack_sheet)

        # 3c/3d: Detect header rows and map columns for invoice.
        inv_col_map = None
        try:
            inv_header_row = detect_header_row(inv_sheet, "invoice", config)
            inv_col_map = map_columns(inv_sheet, inv_header_row, "invoice", config)
        except ProcessingError as e:
            e.filename = filepath.name
            phase3_errors.append(e)

        # 3e/3f: Detect header rows and map columns for packing.
        pack_col_map = None
        try:
            pack_header_row = detect_header_row(pack_sheet, "packing", config)
            pack_col_map = map_columns(
                pack_sheet, pack_header_row, "packing", config,
            )
        except ProcessingError as e:
            e.filename = filepath.name
            phase3_errors.append(e)

        # 3g: Invoice number fallback.
        inv_no: str | None = None
        if inv_col_map is not None:
            if "inv_no" in inv_col_map.field_map:
                inv_no = None  # Will be read per-row by extract_invoice_items.
            else:
                inv_no = extract_inv_no_from_header(inv_sheet, config)
                if inv_no is None:
                    phase3_errors.append(
                        ProcessingError(
                            code=ErrorCode.ERR_021,
                            message=(
                                "Invoice number not found: neither column "
                                "mapping nor header area extraction returned "
                                "a value"
                            ),
                            filename=filepath.name,
                        )
                    )
                else:
                    logger.info(
                        "Inv_No extracted (header): %s", inv_no,
                    )

        if phase3_errors:
            errors.extend(phase3_errors)
            status = determine_file_status(errors, warnings)
            return _build_result(
                filepath, status, errors, warnings, None, None, None,
            )

        # Guaranteed non-None after Phase 3 success.
        assert inv_col_map is not None
        assert pack_col_map is not None

        # Phase 4: Extraction.
        try:
            # 4a: Extract invoice items.
            invoice_items = extract_invoice_items(
                inv_sheet, inv_col_map, inv_merge, inv_no,
            )

            # 4b: Extract packing items.
            packing_items_raw, last_data_row = extract_packing_items(
                pack_sheet, pack_col_map, pack_merge,
            )
            packing_items = packing_items_raw

            # 4c: Validate merged weights.
            validate_merged_weights(packing_items_raw, pack_merge, pack_col_map)

            # 4d: Detect total row.
            total_row = detect_total_row(
                pack_sheet, last_data_row, pack_col_map, pack_merge,
            )

            # 4e: Extract totals.
            packing_totals, totals_warnings = extract_totals(
                pack_sheet, total_row, pack_col_map,
            )
            warnings.extend(totals_warnings)
        except ProcessingError as e:
            e.filename = filepath.name
            errors.append(e)
            status = determine_file_status(errors, warnings)
            return _build_result(
                filepath, status, errors, warnings,
                invoice_items, packing_items, packing_totals,
            )

        # Phase 5: Transformation.
        assert invoice_items is not None

        # 5a: Convert currency.
        invoice_items, currency_warnings = convert_currency(invoice_items, config)
        warnings.extend(currency_warnings)

        # 5b: Convert country.
        invoice_items, coo_warnings = convert_country(invoice_items, config)
        warnings.extend(coo_warnings)

        # 5c: Clean PO numbers.
        invoice_items = clean_po_number(invoice_items)

        # Phase 6: Weight allocation.
        assert packing_items is not None
        assert packing_totals is not None
        try:
            invoice_items = allocate_weights(
                invoice_items, packing_items, packing_totals,
            )
        except ProcessingError as e:
            e.filename = filepath.name
            errors.append(e)
            status = determine_file_status(errors, warnings)
            return _build_result(
                filepath, status, errors, warnings,
                invoice_items, packing_items, packing_totals,
            )

        # Phase 7: Validation.
        status = determine_file_status(errors, warnings)

        # Phase 8: Output.
        if status in ("Success", "Attention"):
            finished_dir = filepath.parent / "finished"
            output_path = finished_dir / f"{filepath.stem}_template.xlsx"
            try:
                write_template(invoice_items, packing_totals, config, output_path)
                logger.info(
                    "Output successfully written to: %s_template.xlsx",
                    filepath.stem,
                )
            except ProcessingError as e:
                e.filename = filepath.name
                errors.append(e)
                status = determine_file_status(errors, warnings)

    except PermissionError:
        errors.append(
            ProcessingError(
                code=ErrorCode.ERR_010,
                message=f"File is locked or inaccessible: {filepath.name}",
                filename=filepath.name,
            )
        )
        status = "Failed"
    except InvalidFileException as e:
        errors.append(
            ProcessingError(
                code=ErrorCode.ERR_011,
                message=f"File is corrupt or invalid: {filepath.name} ({e})",
                filename=filepath.name,
            )
        )
        status = "Failed"
    except Exception as e:  # noqa: BLE001
        # Outer guard: catch any unexpected exception.
        errors.append(
            ProcessingError(
                code=ErrorCode.ERR_011,
                message=f"Unexpected error processing {filepath.name}: {e}",
                filename=filepath.name,
            )
        )
        status = "Failed"

    return _build_result(
        filepath, status, errors, warnings,
        invoice_items if status != "Failed" else None,
        packing_items if status != "Failed" else None,
        packing_totals if status != "Failed" else None,
    )


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _open_workbook(filepath: Path) -> openpyxl.Workbook:
    """Open an Excel workbook, dispatching by extension.

    For .xlsx: uses openpyxl with data_only=True, read_only=False.
    For .xls: uses xlrd with the XlrdSheetAdapter wrapper.

    Args:
        filepath: Path to the Excel file.

    Returns:
        openpyxl Workbook (or an adapter wrapping xlrd).

    Raises:
        PermissionError: If the file is locked (ERR_010).
        InvalidFileException: If the file is corrupt (ERR_011).
    """
    ext = filepath.suffix.lower()

    if ext == ".xls":
        return _open_xls_workbook(filepath)

    # Default: .xlsx
    return openpyxl.load_workbook(filepath, data_only=True, read_only=False)


def _open_xls_workbook(filepath: Path) -> openpyxl.Workbook:
    """Open a .xls file via xlrd and wrap sheets in adapters.

    Creates an openpyxl-compatible Workbook-like object by wrapping
    each xlrd sheet in an XlrdSheetAdapter.

    Args:
        filepath: Path to the .xls file.

    Returns:
        An openpyxl Workbook with adapted sheets.

    Raises:
        PermissionError: If the file is locked.
    """
    import xlrd  # type: ignore[import-untyped]

    from autoconvert.xlrd_adapter import XlrdSheetAdapter

    try:
        xls_book = xlrd.open_workbook(str(filepath))
    except xlrd.XLRDError as e:  # type: ignore[attr-defined]
        raise InvalidFileException(str(e)) from e

    # Build an openpyxl Workbook and replace sheets with adapters.
    wb = openpyxl.Workbook()

    # Remove default sheet.
    default = wb.active
    if default is not None:
        wb.remove(default)

    for sheet_idx in range(xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_idx)
        adapter = XlrdSheetAdapter(xls_sheet)

        # Create an openpyxl sheet placeholder for the adapter.
        wb.create_sheet(title=xls_sheet.name)

        # Reason: Replace the openpyxl worksheet with our adapter in the
        # workbook's internal mapping. detect_sheets uses workbook[name].
        wb._sheets[wb.sheetnames.index(xls_sheet.name)] = adapter  # type: ignore[assignment]

    logger.debug("Opened .xls file via xlrd adapter: %s", filepath.name)
    return wb


def _create_directories(data_dir: Path, finished_dir: Path) -> None:
    """Create data/ and data/finished/ directories if they do not exist.

    Args:
        data_dir: Path to the data directory.
        finished_dir: Path to the data/finished directory.
    """
    for dir_path in (data_dir, finished_dir):
        try:
            os.makedirs(dir_path, exist_ok=True)
        except PermissionError:
            logger.error("Permission denied creating directory: %s", dir_path)
            sys.exit(2)


def _clear_finished_directory(finished_dir: Path) -> None:
    """Remove all files in data/finished/ before processing.

    Does NOT remove subdirectories.

    Args:
        finished_dir: Path to the finished output directory.
    """
    if not finished_dir.exists():
        return

    try:
        for item in finished_dir.iterdir():
            if item.is_file():
                os.remove(item)
    except PermissionError:
        logger.error(
            "Permission denied clearing directory: %s", finished_dir,
        )
        sys.exit(2)


def _scan_input_files(data_dir: Path) -> list[Path]:
    """Scan data_dir for processable .xlsx and .xls files.

    Excludes ~$ temp files and hidden files (name starts with '.').
    Returns files sorted alphabetically by name.

    Args:
        data_dir: Path to the data directory.

    Returns:
        Sorted list of file paths to process.
    """
    xlsx_files = list(data_dir.glob("*.xlsx"))
    xls_files = list(data_dir.glob("*.xls"))

    # Combine and filter: remove .xlsx from xls glob matches
    # (*.xls also matches *.xlsx on some platforms)
    all_files = xlsx_files + [f for f in xls_files if f.suffix.lower() == ".xls"]

    filtered = [
        f for f in all_files
        if not f.name.startswith("~$") and not f.name.startswith(".")
    ]

    return sorted(filtered, key=lambda p: p.name)


def _resolve_diagnostic_file(
    diagnostic_file: str, data_dir: Path,
) -> list[Path]:
    """Resolve the diagnostic file path and return as a single-item list.

    Args:
        diagnostic_file: Filename or path from --diagnostic argument.
        data_dir: Path to the data directory for relative resolution.

    Returns:
        Single-item list containing the resolved file path.
    """
    path = Path(diagnostic_file)
    if not path.is_absolute():
        path = data_dir / path

    return [path] if path.exists() else []


def _build_result(
    filepath: Path,
    status: str,
    errors: list[ProcessingError],
    warnings: list[ProcessingError],
    invoice_items: list[InvoiceItem] | None,
    packing_items: list[PackingItem] | None,
    packing_totals: PackingTotals | None,
) -> FileResult:
    """Build a FileResult from processing outcomes.

    Args:
        filepath: Path to the processed file.
        status: Processing status string.
        errors: Collected error list.
        warnings: Collected warning list.
        invoice_items: Extracted invoice items or None.
        packing_items: Extracted packing items or None.
        packing_totals: Extracted packing totals or None.

    Returns:
        FileResult instance.
    """
    return FileResult(
        filename=filepath.name,
        status=status,
        errors=errors,
        warnings=warnings,
        invoice_items=invoice_items,
        packing_items=packing_items,
        packing_totals=packing_totals,
    )
